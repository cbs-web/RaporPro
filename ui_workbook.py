import datetime
import os
import tkinter as tk
from tkinter import Toplevel, filedialog, messagebox, ttk

from performans import log_exception, perf_timer, perf_tracked
from sabitler import *
from workbook_motoru import (
    WORKBOOK_SHEET_DEFS,
    apply_rows_to_veri as wb_apply_rows_to_veri,
    build_initial_rows as wb_build_initial_rows,
    calc_n30 as wb_calc_n30,
    header_map as wb_header_map,
    rows_to_dicts as wb_rows_to_dicts,
    validate_rows as wb_validate_rows,
)
from karot_motoru import derinlik_baslangic
from yardimcilar import litoloji_yazim_uyarilari, safe_float, temizle_baslik
from widgets import UndoRedoEntry
from ui_workbook_eski import WorkbookEskiMixin


class WorkbookMixin(WorkbookEskiMixin):
    @perf_tracked("workbook.tksheet_open")
    def veri_giris_workbook_tksheet_ac(self):
        try:
            with perf_timer("workbook.tksheet_import"):
                from tksheet import Sheet
        except Exception as exc:
            import sys
            messagebox.showwarning(
                "Workbook",
                f"tksheet yüklenemedi, eski workbook açılıyor:\n{exc}\n\nPython:\n{sys.executable}\n\nÇözüm: Bu Python için `pip install tksheet` çalıştırın."
            )
            return self.veri_giris_workbook_ac()

        self.sondaj_verilerini_kaydet()
        win = Toplevel(self.root)
        self.pencere_hazirla(win, "Veri Giriş Workbook - Excel Modu", "1320x760", (980, 620))

        sheet_defs = WORKBOOK_SHEET_DEFS

        sheets = {}
        sheet_frames = {}
        pending_rows = {}
        pending_source_nos = {}
        frame_to_key = {}
        validate_after = {"id": None}
        internal_update = {"active": False}
        validation_messages = {}
        validation_cells = []
        filter_var = tk.StringVar(value="Tümü")
        wb_info_var = tk.StringVar(value="Hazır")
        filter_combo_ref = {"widget": None}

        top = ttk.Frame(win, padding=8)
        top.pack(fill="x")
        nb = ttk.Notebook(win)
        nb.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        def yeni_sondaj_sablonu(idx):
            bugun = datetime.datetime.now()
            bugun_str = bugun.strftime("%d.%m.%Y")
            t2_str = (bugun + datetime.timedelta(days=10)).strftime("%d.%m.%Y")
            return {
                "no": f"SK-{idx + 1}", "der": "15.0", "y": "", "x": "", "k": "",
                "bas_tar": bugun_str, "bit_tar": bugun_str,
                "yass_d1": "", "yass_t1": bugun_str, "yass_d2": "", "yass_t2": t2_str,
                "litoloji": [], "spt": [], "pmt": [], "kaya": [], "numuneler": []
            }

        def normalize_header(cell):
            import unicodedata
            text = str(cell).strip().lower()
            text = text.replace("\u0131", "i").replace("\u0130", "i")
            text = unicodedata.normalize("NFKD", text)
            text = "".join(ch for ch in text if not unicodedata.combining(ch))
            return temizle_baslik(text)

        def get_active_key():
            return frame_to_key.get(nb.select(), "sondajlar")

        def row_has_data(values, ignored=None):
            ignored = ignored or set()
            return any(str(value).strip() for key, value in values.items() if key not in ignored)

        def row_values_to_list(sheet_key, values):
            return [values.get(col_key, "") for _, col_key in sheet_defs[sheet_key]["columns"]]

        def normalize_data_rows(sheet_key, data):
            ncols = len(sheet_defs[sheet_key]["columns"])
            normalized = []
            for row in data:
                values = ["" if cell is None else str(cell) for cell in list(row)[:ncols]]
                if len(values) < ncols:
                    values.extend([""] * (ncols - len(values)))
                normalized.append(values)
            return normalized

        def sync_source_nos(sheet_key):
            if sheet_key not in sheets:
                total = len(pending_rows.get(sheet_key, []))
                source_nos = pending_source_nos.setdefault(sheet_key, [])
                if len(source_nos) < total:
                    source_nos.extend([""] * (total - len(source_nos)))
                elif len(source_nos) > total:
                    del source_nos[total:]
                return
            total = sheets[sheet_key]["widget"].get_total_rows()
            source_nos = sheets[sheet_key].setdefault("source_nos", [])
            if len(source_nos) < total:
                source_nos.extend([""] * (total - len(source_nos)))
            elif len(source_nos) > total:
                del source_nos[total:]

        def collect_rows(sheet_key):
            sync_source_nos(sheet_key)
            if sheet_key not in sheets:
                return wb_rows_to_dicts(sheet_key, pending_rows.get(sheet_key, []), sheet_defs)
            sheet = sheets[sheet_key]["widget"]
            data = normalize_data_rows(sheet_key, sheet.get_sheet_data())
            keys = [col_key for _, col_key in sheet_defs[sheet_key]["columns"]]
            return [{key: str(row[idx]).strip() for idx, key in enumerate(keys)} for row in data]

        def default_row_values(sheet_key):
            if sheet_key == "sondajlar":
                row_count = sheets[sheet_key]["widget"].get_total_rows() if sheet_key in sheets else len(pending_rows.get(sheet_key, []))
                return yeni_sondaj_sablonu(row_count)
            vals = {}
            data = collect_rows(sheet_key)
            if data:
                last = data[-1]
                vals["sondaj_no"] = last.get("sondaj_no", "")
                if sheet_key == "litoloji":
                    vals["top"] = last.get("bot", "")
                elif sheet_key == "spt":
                    vals["der"] = f"{safe_float(last.get('der')) + 1.5:.2f}" if last.get("der") else ""
            else:
                sondajlar = collect_rows("sondajlar") if "sondajlar" in sheets else []
                vals["sondaj_no"] = next((r.get("no") for r in sondajlar if r.get("no")), "SK-1")
            return vals

        def set_sheet_rows(sheet_key, rows, source_nos=None):
            data = normalize_data_rows(sheet_key, rows)
            if not data:
                data = [row_values_to_list(sheet_key, default_row_values(sheet_key))]
            if sheet_key not in sheets:
                pending_rows[sheet_key] = data
                pending_source_nos[sheet_key] = list(source_nos or [""] * len(data))
                sync_source_nos(sheet_key)
                return
            sheet = sheets[sheet_key]["widget"]
            internal_update["active"] = True
            sheet.set_sheet_data(data, reset_col_positions=False, reset_row_positions=True, reset_highlights=True)
            sheet.row_index([str(i + 1) for i in range(len(data))], redraw=False)
            sheet.refresh()
            internal_update["active"] = False
            sheets[sheet_key]["source_nos"] = list(source_nos or [""] * len(data))
            sync_source_nos(sheet_key)

        def selected_row_index(sheet):
            rows = sorted(sheet.get_selected_rows(return_tuple=True))
            if rows:
                return rows[0]
            r1, c1, r2, c2 = sheet.get_selected_min_max()
            if r1 is not None:
                return r1
            selected = sheet.get_currently_selected()
            return getattr(selected, "row", None) if selected else None

        def selected_rows(sheet):
            rows = sorted(sheet.get_selected_rows(return_tuple=True))
            if rows:
                return rows
            r1, c1, r2, c2 = sheet.get_selected_min_max()
            if r1 is None:
                current = selected_row_index(sheet)
                return [current] if current is not None else []
            return list(range(r1, max(r1 + 1, r2)))

        def current_cell(sheet, row=None, col=None):
            total_rows, total_cols = sheet.get_total_rows(), sheet.get_total_columns()
            if total_rows <= 0 or total_cols <= 0:
                return None, None
            if row is None or col is None:
                selected = sheet.get_currently_selected()
                row = getattr(selected, "row", None) if selected else None
                col = getattr(selected, "column", None) if selected else None
            if row is None or col is None:
                r1, c1, r2, c2 = sheet.get_selected_min_max()
                row = r1 if r1 is not None else 0
                col = c1 if c1 is not None else 0
            row = max(0, min(int(row), total_rows - 1))
            col = max(0, min(int(col), total_cols - 1))
            return row, col

        def select_horizontal_cell(sheet_key, row=None, col=None, step=1):
            sheet = sheets[sheet_key]["widget"]
            row, col = current_cell(sheet, row, col)
            if row is None or col is None:
                return "break"
            last_col = max(0, sheet.get_total_columns() - 1)
            next_col = max(0, min(col + step, last_col))
            sheet.select_cell(row, next_col)
            sheet.see(row, next_col, bottom_right_corner=step > 0)
            try:
                sheet.MT.focus_set()
            except Exception:
                pass
            return "break"

        def workbook_tab_key(sheet_key, step=1):
            return select_horizontal_cell(sheet_key, step=step)

        def workbook_text_editor_tab(sheet_key, event=None, step=1):
            sheet = sheets[sheet_key]["widget"]
            row, col = None, None
            try:
                row, col = sheet.MT.text_editor.coords
            except Exception:
                pass
            try:
                sheet.MT.close_text_editor(event)
            except Exception:
                pass
            return select_horizontal_cell(sheet_key, row, col, step)

        def ensure_sheet(sheet_key):
            if sheet_key not in sheets:
                create_sheet(sheet_key)
            return sheets[sheet_key]["widget"]

        def ensure_all_sheets():
            for key in sheet_defs:
                ensure_sheet(key)

        def ensure_active_sheet():
            return ensure_sheet(get_active_key())

        def sondaj_filter_values():
            values = ["Tümü"]
            for row in collect_rows("sondajlar"):
                no = row.get("no", "").strip()
                if no and no not in values:
                    values.append(no)
            return values

        def refresh_filter_values():
            values = sondaj_filter_values()
            if filter_var.get() not in values:
                filter_var.set("Tümü")
            combo = filter_combo_ref.get("widget")
            if combo is not None:
                combo.configure(values=values)

        def row_sondaj_no(sheet_key, values):
            return values.get("no", "").strip() if sheet_key == "sondajlar" else values.get("sondaj_no", "").strip()

        def selected_sondaj_no(sheet_key=None):
            sheet_key = sheet_key or get_active_key()
            try:
                sheet = ensure_sheet(sheet_key)
                idx = selected_row_index(sheet)
                rows = collect_rows(sheet_key)
                if idx is not None and idx < len(rows):
                    no = row_sondaj_no(sheet_key, rows[idx])
                    if no:
                        return no
            except Exception:
                pass
            selected_filter = filter_var.get().strip()
            if selected_filter and selected_filter != "Tümü":
                return selected_filter
            return next((row.get("no") for row in collect_rows("sondajlar") if row.get("no")), "")

        def apply_sondaj_filter(show_status=True):
            sheet_key = get_active_key()
            sheet = ensure_sheet(sheet_key)
            target_no = filter_var.get().strip()
            validate_workbook(show_status=False)
            if not target_no or target_no == "Tümü":
                wb_info_var.set("Filtre kapalı.")
                if show_status:
                    self.set_status("Workbook sondaj filtresi kapatıldı.", level="info")
                return
            rows = collect_rows(sheet_key)
            matches = [idx for idx, values in enumerate(rows) if row_sondaj_no(sheet_key, values) == target_no]
            for row_idx in matches:
                try:
                    sheet.highlight_cells(row=row_idx, bg="#D6EAF8", redraw=False)
                except Exception as exc:
                    log_exception("ui_workbook.litoloji_auto_next_row", exc_value=exc)
            sheet.redraw()
            if matches:
                sheet.select_cell(matches[0], 0)
                sheet.see(matches[0], 0)
                msg = f"{target_no}: aktif sayfada {len(matches)} satır vurgulandı."
                wb_info_var.set(msg)
                if show_status:
                    self.set_status(msg, level="info")
            else:
                msg = f"{target_no}: aktif sayfada eşleşen satır yok."
                wb_info_var.set(msg)
                if show_status:
                    self.set_status(msg, level="warning")

        def validation_message(sheet_key, col_key, level):
            prefix = "Hata" if level == "error" else "Uyarı"
            messages = {
                "no": "Sondaj no boş ya da tekrarlı olabilir.",
                "sondaj_no": "Bu satırdaki sondaj no, Sondajlar sayfasında tanımlı değil.",
                "der": "Derinlik pozitif olmalı ve sondaj derinliğini aşmamalı.",
                "top": "Litoloji başlangıç değeri önceki tabaka ile uyumlu olmalı.",
                "bot": "Litoloji bitiş değeri başlangıçtan büyük olmalı ve sondaj derinliğini aşmamalı.",
                "tanim": "Litoloji tanımı yazım açısından kontrol edilmeli.",
                "y": "Koordinat boş ya da 0 görünüyor.",
                "x": "Koordinat boş ya da 0 görünüyor.",
            }
            return f"{prefix}: {messages.get(col_key, 'Bu hücre kontrol edilmeli.')}"

        def show_current_cell_message(sheet_key):
            try:
                sheet = sheets[sheet_key]["widget"]
                row, col = current_cell(sheet)
                if row is None or col is None:
                    return
                msg = validation_messages.get((sheet_key, row, col))
                if msg:
                    wb_info_var.set(msg)
                    self.set_status(msg, level="warning" if msg.startswith("Uyarı") else "error")
            except Exception:
                pass

        def active_add_row():
            sheet_key = get_active_key()
            sheet = ensure_sheet(sheet_key)
            sheet.insert_rows([row_values_to_list(sheet_key, default_row_values(sheet_key))], idx="end", undo=True)
            sheets[sheet_key]["source_nos"].append("")
            sync_source_nos(sheet_key)
            sheet.row_index([str(i + 1) for i in range(sheet.get_total_rows())])
            refresh_filter_values()
            schedule_validate()

        def active_insert_row():
            sheet_key = get_active_key()
            sheet = ensure_sheet(sheet_key)
            idx = selected_row_index(sheet)
            if idx is None:
                idx = sheet.get_total_rows()
            else:
                idx += 1
            sheet.insert_rows([row_values_to_list(sheet_key, default_row_values(sheet_key))], idx=idx, undo=True)
            sheets[sheet_key]["source_nos"].insert(idx, "")
            sync_source_nos(sheet_key)
            sheet.row_index([str(i + 1) for i in range(sheet.get_total_rows())])
            sheet.select_cell(idx, 0)
            refresh_filter_values()
            schedule_validate()

        def active_delete_rows():
            sheet_key = get_active_key()
            sheet = ensure_sheet(sheet_key)
            rows = selected_rows(sheet)
            if not rows:
                return
            sheet.delete_rows(rows, undo=True)
            for row_idx in sorted(rows, reverse=True):
                if row_idx < len(sheets[sheet_key]["source_nos"]):
                    del sheets[sheet_key]["source_nos"][row_idx]
            if sheet.get_total_rows() == 0:
                sheet.insert_rows([row_values_to_list(sheet_key, default_row_values(sheet_key))], idx="end", undo=True)
                sheets[sheet_key]["source_nos"].append("")
            sheet.row_index([str(i + 1) for i in range(sheet.get_total_rows())])
            refresh_filter_values()
            schedule_validate()

        def active_clear_column():
            sheet_key = get_active_key()
            sheet = ensure_sheet(sheet_key)
            row_idx, col_idx = current_cell(sheet)
            if col_idx is None:
                return
            columns = sheet_defs[sheet_key]["columns"]
            if col_idx >= len(columns):
                return
            label, _ = columns[col_idx]
            total = sheet.get_total_rows()
            if total <= 0:
                return
            if not messagebox.askyesno("Sütunu Temizle", f"{label} sütunundaki {total} hücre temizlensin mi?", parent=win):
                return
            internal_update["active"] = True
            for row_no in range(total):
                sheet.set_cell_data(row_no, col_idx, "", redraw=False)
            internal_update["active"] = False
            sheet.refresh()
            self.set_status(f"{sheet_defs[sheet_key]['title']} / {label} sütunu temizlendi.", level="info")
            schedule_validate()

        def active_duplicate_rows():
            sheet_key = get_active_key()
            sheet = ensure_sheet(sheet_key)
            rows_idx = selected_rows(sheet)
            if not rows_idx:
                return
            rows = collect_rows(sheet_key)
            copied = [row_values_to_list(sheet_key, rows[idx]) for idx in rows_idx if idx < len(rows)]
            if not copied:
                return
            insert_idx = max(rows_idx) + 1
            sheet.insert_rows(copied, idx=insert_idx, undo=True)
            for offset in range(len(copied)):
                sheets[sheet_key]["source_nos"].insert(insert_idx + offset, "")
            sheet.row_index([str(i + 1) for i in range(sheet.get_total_rows())])
            sheet.select_cell(insert_idx, 0)
            refresh_filter_values()
            schedule_validate()
            self.set_status(f"{len(copied)} satır çoğaltıldı.", level="success")

        def active_smart_row():
            sheet_key = get_active_key()
            if sheet_key != "litoloji":
                active_add_row()
                return
            sheet = ensure_sheet(sheet_key)
            row_idx = selected_row_index(sheet)
            if row_idx is None:
                row_idx = max(0, sheet.get_total_rows() - 1)
            rows = collect_rows("litoloji")
            base = rows[row_idx] if row_idx < len(rows) else {}
            new_row = {"sondaj_no": base.get("sondaj_no", ""), "top": base.get("bot", ""), "bot": "", "tanim": base.get("tanim", "")}
            insert_idx = row_idx + 1
            sheet.insert_rows([row_values_to_list("litoloji", new_row)], idx=insert_idx, undo=True)
            sheets[sheet_key]["source_nos"].insert(insert_idx, "")
            sheet.select_cell(insert_idx, 2)
            sheet.see(insert_idx, 2)
            schedule_validate()

        def generate_spt_rows():
            target_no = selected_sondaj_no(get_active_key())
            if not target_no:
                messagebox.showwarning("SPT", "SPT üretmek için önce bir sondaj seçin.")
                return
            sondaj = next((row for row in collect_rows("sondajlar") if row.get("no") == target_no), None)
            if not sondaj:
                messagebox.showwarning("SPT", f"{target_no} sondajı bulunamadı.")
                return
            depth = safe_float(sondaj.get("der"))
            if depth <= 0:
                messagebox.showwarning("SPT", f"{target_no} için geçerli sondaj derinliği yok.")
                return
            sheet = ensure_sheet("spt")
            existing = {
                round(safe_float(row.get("der")), 2)
                for row in collect_rows("spt")
                if row.get("sondaj_no") == target_no and safe_float(row.get("der")) > 0
            }
            new_rows = []
            d = 1.5
            while d <= depth + 0.01:
                key = round(d, 2)
                if key not in existing:
                    new_rows.append(row_values_to_list("spt", {
                        "sondaj_no": target_no,
                        "der": f"{d:.2f}",
                        "v15": "",
                        "v30": "",
                        "v45": "",
                        "n30": "",
                    }))
                d += 1.5
            if not new_rows:
                self.set_status(f"{target_no} için eksik SPT satırı yok.", level="info")
                return
            insert_idx = sheet.get_total_rows()
            sheet.insert_rows(new_rows, idx="end", undo=True)
            sheets["spt"]["source_nos"].extend([""] * len(new_rows))
            sheet.row_index([str(i + 1) for i in range(sheet.get_total_rows())])
            nb.select(sheet_frames["spt"])
            sheet.select_cell(insert_idx, 0)
            sheet.see(insert_idx, 0)
            filter_var.set(target_no)
            refresh_filter_values()
            schedule_validate()
            self.set_status(f"{target_no} için {len(new_rows)} SPT satırı üretildi.", level="success")

        def generate_litoloji_missing_rows():
            sheet = ensure_sheet("litoloji")
            existing_nos = {row.get("sondaj_no") for row in collect_rows("litoloji") if row_has_data(row, {"sondaj_no"})}
            new_rows = []
            for sondaj in collect_rows("sondajlar"):
                no = sondaj.get("no", "").strip()
                if not no or no in existing_nos:
                    continue
                depth = safe_float(sondaj.get("der"))
                if depth <= 0:
                    continue
                new_rows.append(row_values_to_list("litoloji", {
                    "sondaj_no": no,
                    "top": "0.00",
                    "bot": f"{depth:.2f}",
                    "tanim": "",
                }))
            if not new_rows:
                self.set_status("Eksik litoloji başlangıç satırı yok.", level="info")
                return
            insert_idx = sheet.get_total_rows()
            sheet.insert_rows(new_rows, idx="end", undo=True)
            sheets["litoloji"]["source_nos"].extend([""] * len(new_rows))
            sheet.row_index([str(i + 1) for i in range(sheet.get_total_rows())])
            nb.select(sheet_frames["litoloji"])
            sheet.select_cell(insert_idx, 0)
            sheet.see(insert_idx, 0)
            refresh_filter_values()
            schedule_validate()
            self.set_status(f"{len(new_rows)} litoloji başlangıç satırı üretildi.", level="success")

        def generate_selected_sondaj_package():
            target_no = selected_sondaj_no(get_active_key())
            if not target_no:
                messagebox.showwarning("Workbook", "Önce bir sondaj seçin.")
                return
            sondaj = next((row for row in collect_rows("sondajlar") if row.get("no") == target_no), None)
            if not sondaj:
                messagebox.showwarning("Workbook", f"{target_no} sondajı bulunamadı.")
                return
            depth = safe_float(sondaj.get("der"))
            if depth <= 0:
                messagebox.showwarning("Workbook", f"{target_no} için geçerli sondaj derinliği yok.")
                return

            created = 0
            lit_sheet = ensure_sheet("litoloji")
            if not any(row.get("sondaj_no") == target_no for row in collect_rows("litoloji")):
                lit_sheet.insert_rows([row_values_to_list("litoloji", {"sondaj_no": target_no, "top": "0.00", "bot": f"{depth:.2f}", "tanim": ""})], idx="end", undo=True)
                sheets["litoloji"]["source_nos"].append("")
                created += 1

            spt_sheet = ensure_sheet("spt")
            existing_spt = {
                round(safe_float(row.get("der")), 2)
                for row in collect_rows("spt")
                if row.get("sondaj_no") == target_no and safe_float(row.get("der")) > 0
            }
            spt_rows = []
            d = 1.5
            while d <= depth + 0.01:
                key = round(d, 2)
                if key not in existing_spt:
                    spt_rows.append(row_values_to_list("spt", {"sondaj_no": target_no, "der": f"{d:.2f}", "v15": "", "v30": "", "v45": "", "n30": ""}))
                d += 1.5
            if spt_rows:
                spt_sheet.insert_rows(spt_rows, idx="end", undo=True)
                sheets["spt"]["source_nos"].extend([""] * len(spt_rows))
                created += len(spt_rows)

            for key in ("litoloji", "spt"):
                if key in sheets:
                    sheets[key]["widget"].row_index([str(i + 1) for i in range(sheets[key]["widget"].get_total_rows())])
            filter_var.set(target_no)
            refresh_filter_values()
            nb.select(sheet_frames["litoloji"])
            apply_sondaj_filter(show_status=False)
            schedule_validate()
            self.set_status(f"{target_no} için {created} hızlı veri satırı hazırlandı.", level="success" if created else "info")

        def auto_spt_n30():
            if "spt" not in sheets:
                rows = normalize_data_rows("spt", pending_rows.get("spt", []))
                changed = False
                for row in rows:
                    if len(row) >= 6 and not str(row[5]).strip():
                        calculated = wb_calc_n30(row[3] if len(row) > 3 else "", row[4] if len(row) > 4 else "")
                        if calculated:
                            row[5] = calculated
                            changed = True
                if changed:
                    pending_rows["spt"] = rows
                return
            sheet = sheets["spt"]["widget"]
            rows = collect_rows("spt")
            changed = False
            internal_update["active"] = True
            for row_idx, values in enumerate(rows):
                if values.get("n30", "").strip():
                    continue
                joined = " ".join([values.get("v30", ""), values.get("v45", "")]).lower()
                if "50/" in joined or "-" in joined:
                    calculated = "R"
                else:
                    total = safe_float(values.get("v30")) + safe_float(values.get("v45"))
                    calculated = str(int(total)) if total and float(total).is_integer() else (str(total) if total else "")
                if calculated:
                    sheet.set_cell_data(row_idx, 5, calculated, redraw=False)
                    changed = True
            internal_update["active"] = False
            if changed:
                sheet.refresh()

        def header_map(sheet_key, cells):
            aliases = {
                "sondajno": "sondaj_no", "sondaj": "sondaj_no", "sk": "sondaj_no", "kuyuno": "sondaj_no",
                "no": "no", "sondajadi": "no", "derinlik": "der", "der": "der", "derinlikm": "der",
                "enlem": "y", "lat": "y", "latitude": "y", "y": "y", "boylam": "x", "lon": "x", "longitude": "x", "x": "x",
                "tur": "sondaj_turu", "turu": "sondaj_turu", "sondajturu": "sondaj_turu", "zeminkaya": "sondaj_turu",
                "delgicapi": "delgi_capi", "delgicap": "delgi_capi", "cap": "delgi_capi", "capi": "delgi_capi",
                "kot": "k", "bastarih": "bas_tar", "bastarihi": "bas_tar", "baslangictarihi": "bas_tar",
                "bittarih": "bit_tar", "bittarihi": "bit_tar", "bitistarihi": "bit_tar",
                "yassilk": "yass_d1", "yassd1": "yass_d1", "yass1": "yass_d1", "yasst1": "yass_t1", "yassilktarih": "yass_t1",
                "yassson": "yass_d2", "yassd2": "yass_d2", "yass2": "yass_d2", "yasst2": "yass_t2", "yasssontarih": "yass_t2",
                "baslangic": "top", "bas": "top", "ust": "top", "top": "top", "bitis": "bot", "bit": "bot", "alt": "bot", "bot": "bot",
                "tanim": "tanim", "litoloji": "tanim", "birim": "tanim",
                "15": "v15", "n15": "v15", "30": "v30", "n30vurus": "v30", "45": "v45", "n45": "v45", "n30": "n30",
                "em": "em", "pl": "pl", "tcr": "tcr", "scr": "scr", "rqd": "rqd",
                "aralik": "aralik", "derinlikaralik": "aralik", "tur": "tur", "turu": "tur", "turuno": "tur", "numune": "tur",
            }
            allowed = {key for _, key in sheet_defs[sheet_key]["columns"]}
            mapped = []
            for cell in cells:
                normalized = normalize_header(cell)
                key = aliases.get(normalized)
                if sheet_key == "sondajlar" and normalized in ("tur", "turu", "sondajturu", "zeminkaya"):
                    key = "sondaj_turu"
                if sheet_key == "sondajlar" and key == "sondaj_no":
                    key = "no"
                elif sheet_key != "sondajlar" and key == "no":
                    key = "sondaj_no"
                mapped.append(key if key in allowed else None)
            return mapped if sum(1 for item in mapped if item) >= 2 else None

        @perf_tracked("workbook.tksheet_validate")
        def validate_workbook(show_status=True):
            if internal_update["active"]:
                return 0, 0
            auto_spt_n30()
            for sheet_info in sheets.values():
                sheet_info["widget"].dehighlight_all(redraw=False)
            rows_by_sheet = {sheet_key: collect_rows(sheet_key) for sheet_key in sheet_defs}
            result = wb_validate_rows(rows_by_sheet)
            validation_messages.clear()
            validation_cells.clear()

            def to_col_indexes(items, level):
                indexed = []
                for sheet_key, row_idx, col_key in items:
                    col_keys = [key for _, key in sheet_defs[sheet_key]["columns"]]
                    if col_key in col_keys:
                        col_idx = col_keys.index(col_key)
                        indexed.append((sheet_key, row_idx, col_idx))
                        validation_messages[(sheet_key, row_idx, col_idx)] = validation_message(sheet_key, col_key, level)
                return indexed

            invalid = to_col_indexes(result["errors"], "error")
            warnings = to_col_indexes(result["warnings"], "warning")
            validation_cells.extend([("error", *item) for item in invalid])
            validation_cells.extend([("warning", *item) for item in warnings])
            for sheet_key, row_idx, col_idx in warnings:
                if sheet_key in sheets:
                    sheets[sheet_key]["widget"].highlight_cells(row=row_idx, column=col_idx, bg="#FCF3CF", redraw=False)
            for sheet_key, row_idx, col_idx in invalid:
                if sheet_key in sheets:
                    sheets[sheet_key]["widget"].highlight_cells(row=row_idx, column=col_idx, bg="#FADBD8", redraw=False)
            for sheet_info in sheets.values():
                sheet_info["widget"].redraw()
            if show_status:
                if invalid:
                    self.set_status(f"Workbook kontrolü: {len(invalid)} hata, {len(warnings)} uyarı.", level="error")
                elif warnings:
                    self.set_status(f"Workbook kontrolü: {len(warnings)} uyarı.", level="warning")
                else:
                    self.set_status("Workbook kontrolü temiz.", level="success")
            return len(invalid), len(warnings)

        def go_to_first_issue():
            validate_workbook(show_status=True)
            if not validation_cells:
                self.set_status("Workbook kontrolü temiz.", level="success")
                return
            level, sheet_key, row_idx, col_idx = validation_cells[0]
            sheet = ensure_sheet(sheet_key)
            nb.select(sheet_frames[sheet_key])
            sheet.select_cell(row_idx, col_idx)
            sheet.see(row_idx, col_idx)
            show_current_cell_message(sheet_key)
            if not validation_messages.get((sheet_key, row_idx, col_idx)):
                self.set_status(f"İlk {level}: {sheet_defs[sheet_key]['title']} satır {row_idx + 1}.", level="error" if level == "error" else "warning")

        def schedule_validate(event=None):
            if internal_update["active"]:
                return
            if validate_after["id"]:
                try: win.after_cancel(validate_after["id"])
                except Exception: pass
            validate_after["id"] = win.after(350, lambda: validate_workbook(show_status=False))

        def on_sheet_edit(sheet_key, event=None):
            if sheet_key == "litoloji":
                try:
                    row_idx, col_idx = int(event.row), int(event.column)
                    if sheet_defs[sheet_key]["columns"][col_idx][1] == "bot":
                        rows = collect_rows("litoloji")
                        if row_idx < len(rows) and rows[row_idx].get("bot", "").strip():
                            next_idx = row_idx + 1
                            if next_idx >= len(rows) or not row_has_data(rows[next_idx]):
                                base = rows[row_idx]
                                new_row = {"sondaj_no": base.get("sondaj_no", ""), "top": base.get("bot", ""), "bot": "", "tanim": base.get("tanim", "")}
                                sheets[sheet_key]["widget"].insert_rows([row_values_to_list("litoloji", new_row)], idx=next_idx, undo=True)
                                sheets[sheet_key]["source_nos"].insert(next_idx, "")
                except Exception as exc:
                    log_exception("ui_workbook.dropdown_sondaj_names", exc_value=exc)
            schedule_validate()

        def show_workbook_context_menu(sheet_key, event=None):
            sheet = ensure_sheet(sheet_key)
            nb.select(sheet_frames[sheet_key])
            try:
                row_idx = sheet.identify_row(event, allow_end=False) if event is not None else None
                col_idx = sheet.identify_column(event, allow_end=False) if event is not None else None
                if row_idx is not None and col_idx is not None:
                    sheet.select_cell(row_idx, col_idx)
                elif row_idx is not None:
                    sheet.select_row(row_idx)
            except Exception:
                pass
            menu = tk.Menu(win, tearoff=False)
            menu.add_command(label="Satır ekle", command=active_insert_row)
            menu.add_command(label="Satır çoğalt", command=active_duplicate_rows)
            menu.add_command(label="Satır sil", command=active_delete_rows)
            menu.add_command(label="Sütunu temizle", command=active_clear_column)
            menu.add_separator()
            menu.add_command(label="Akıllı satır", command=active_smart_row)
            menu.add_command(label="Sondaj paketi", command=generate_selected_sondaj_package)
            if sheet_key == "spt":
                menu.add_command(label="N30 hesapla", command=lambda: (auto_spt_n30(), validate_workbook(show_status=True)))
            elif sheet_key == "litoloji":
                menu.add_command(label="Eksik litoloji", command=generate_litoloji_missing_rows)
            menu.add_separator()
            menu.add_command(label="Kontrol et", command=lambda: validate_workbook(show_status=True))
            menu.add_command(label="İlk soruna git", command=go_to_first_issue)
            menu.add_command(label="Boyuta sigdir", command=lambda: ensure_sheet(sheet_key).set_all_cell_sizes_to_text())
            try:
                x_root = getattr(event, "x_root", win.winfo_pointerx())
                y_root = getattr(event, "y_root", win.winfo_pointery())
                menu.tk_popup(x_root, y_root)
            finally:
                try:
                    menu.grab_release()
                except Exception:
                    pass
            return "break"

        @perf_tracked("workbook.tksheet_export_excel")
        def export_workbook():
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                from openpyxl.utils import get_column_letter
            except Exception as exc:
                messagebox.showerror("Excel", f"openpyxl yüklenemedi:\n{exc}"); return
            path = filedialog.asksaveasfilename(title="Workbook Excel'e Aktar", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
            if not path: return
            wb = Workbook(); wb.remove(wb.active)
            for sheet_key, spec in sheet_defs.items():
                ws = wb.create_sheet(spec["title"])
                ws.append([label for label, _ in spec["columns"]])
                for cell in ws[1]:
                    cell.font = Font(bold=True); cell.fill = PatternFill("solid", fgColor="D9EAF7")
                for values in collect_rows(sheet_key):
                    if row_has_data(values, {"sondaj_no"}):
                        ws.append([values.get(col_key, "") for _, col_key in spec["columns"]])
                for col_idx, width in enumerate(spec["widths"], start=1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = max(10, width / 7)
            try:
                wb.save(path); self.set_status(f"Workbook Excel'e aktarıldı: {os.path.basename(path)}", level="success")
            except Exception as exc:
                messagebox.showerror("Excel", f"Excel kaydedilemedi:\n{exc}")

        @perf_tracked("workbook.tksheet_import_excel")
        def import_workbook():
            try:
                from openpyxl import load_workbook
            except Exception as exc:
                messagebox.showerror("Excel", f"openpyxl yüklenemedi:\n{exc}"); return
            path = filedialog.askopenfilename(title="Workbook Excel'den Al", filetypes=[("Excel", "*.xlsx")])
            if not path: return
            try:
                wb = load_workbook(path, data_only=True)
            except Exception as exc:
                messagebox.showerror("Excel", f"Excel okunamadı:\n{exc}"); return
            imported = 0
            for sheet_key, spec in sheet_defs.items():
                if spec["title"] not in wb.sheetnames:
                    continue
                raw_rows = []
                for row in wb[spec["title"]].iter_rows(values_only=True):
                    cells = ["" if cell is None else str(cell) for cell in row]
                    if any(cell.strip() for cell in cells):
                        raw_rows.append(cells)
                if not raw_rows:
                    continue
                mapping = wb_header_map(sheet_key, raw_rows[0], sheet_defs)
                data_rows = raw_rows[1:] if mapping else raw_rows
                new_rows = []
                for raw in data_rows:
                    values = {}
                    if mapping:
                        for idx, value in enumerate(raw):
                            if idx < len(mapping) and mapping[idx]:
                                values[mapping[idx]] = value.strip()
                    else:
                        for idx, value in enumerate(raw):
                            if idx < len(spec["columns"]):
                                values[spec["columns"][idx][1]] = value.strip()
                    if row_has_data(values, {"sondaj_no"}):
                        new_rows.append(row_values_to_list(sheet_key, values))
                if new_rows:
                    set_sheet_rows(sheet_key, new_rows, [row[0] if row else "" for row in new_rows] if sheet_key == "sondajlar" else [""] * len(new_rows))
                    imported += len(new_rows)
            validate_workbook(show_status=True)
            refresh_filter_values()
            self.set_status(f"Excel'den workbook'a {imported} satır aktarıldı.", level="success")

        @perf_tracked("workbook.tksheet_apply")
        def apply_workbook(close=False):
            error_count, warning_count = validate_workbook(show_status=True)
            if error_count and not messagebox.askyesno("Workbook Kontrol", f"{error_count} hata görünüyor. Yine de uygulansın mı?"):
                return
            if warning_count and not error_count and not messagebox.askyesno("Workbook Kontrol", f"{warning_count} uyarı var. Yine de uygulansın mı?"):
                return
            rows_by_sheet = {sheet_key: collect_rows(sheet_key) for sheet_key in sheet_defs}
            new_sondajlar, warnings = wb_apply_rows_to_veri(self.veri, rows_by_sheet, sheets["sondajlar"].get("source_nos", []))
            if not new_sondajlar:
                messagebox.showwarning("Workbook", "Sondajlar sayfasında uygulanacak veri yok."); return
            self.veri["sondaj"] = new_sondajlar
            self.sondaj_tablosunu_ciz(); self.ozet_yenile(collect=False)
            self.set_status(f"Excel mod workbook uygulandı: {len(new_sondajlar)} sondaj güncellendi.", level="success")
            for warning in warnings[:5]: self.set_status(warning, level="warning")
            if len(warnings) > 5: self.set_status(f"{len(warnings) - 5} ek workbook uyarısı daha var.", level="warning")
            if close: win.destroy()

        def create_sheet(sheet_key):
            if sheet_key in sheets:
                return sheets[sheet_key]["widget"]
            spec = sheet_defs[sheet_key]
            frame = sheet_frames.get(sheet_key)
            if frame is None:
                frame = ttk.Frame(nb)
                nb.add(frame, text=spec["title"])
                sheet_frames[sheet_key] = frame
                frame_to_key[str(frame)] = sheet_key
            for child in frame.winfo_children():
                child.destroy()
            sheet = Sheet(
                frame, headers=[label for label, _ in spec["columns"]], data=[],
                show_row_index=True, show_header=True, theme="light blue",
                paste_can_expand_y=True, paste_can_expand_x=False,
                edit_cell_return="down", edit_cell_tab="",
                default_column_width=110, default_row_index_width=52,
                column_drag_and_drop_perform=False, row_drag_and_drop_perform=True,
            )
            sheet.pack(fill="both", expand=True)
            sheet.enable_bindings("all")
            def safe_sheet_bind(binding, func):
                try:
                    sheet.bind(binding, func)
                except tk.TclError:
                    pass

            def safe_editor_bind(binding, func):
                try:
                    sheet.bind_key_text_editor(binding, func)
                except tk.TclError:
                    pass

            safe_sheet_bind("<Tab>", lambda event=None, sk=sheet_key: workbook_tab_key(sk, 1))
            safe_sheet_bind("<Shift-Tab>", lambda event=None, sk=sheet_key: workbook_tab_key(sk, -1))
            safe_sheet_bind("<ButtonRelease-1>", lambda event=None, sk=sheet_key: show_current_cell_message(sk))
            safe_sheet_bind("<Button-3>", lambda event=None, sk=sheet_key: show_workbook_context_menu(sk, event))
            safe_sheet_bind("<Button-2>", lambda event=None, sk=sheet_key: show_workbook_context_menu(sk, event))
            for context_target in (getattr(sheet, "MT", None), getattr(sheet, "CH", None), getattr(sheet, "RI", None), getattr(sheet, "TL", None)):
                if context_target is None:
                    continue
                try:
                    context_target.bind("<Button-3>", lambda event=None, sk=sheet_key: show_workbook_context_menu(sk, event), add="+")
                    context_target.bind("<Button-2>", lambda event=None, sk=sheet_key: show_workbook_context_menu(sk, event), add="+")
                except tk.TclError:
                    pass
            safe_editor_bind("<Tab>", lambda event=None, sk=sheet_key: workbook_text_editor_tab(sk, event, 1))
            safe_editor_bind("<Shift-Tab>", lambda event=None, sk=sheet_key: workbook_text_editor_tab(sk, event, -1))
            sheet.extra_bindings("all_modified_events", lambda event=None, sk=sheet_key: on_sheet_edit(sk, event))
            sheet.set_column_widths(spec["widths"])
            sheets[sheet_key] = {"widget": sheet, "source_nos": []}
            rows = pending_rows.get(sheet_key, [])
            source = pending_source_nos.get(sheet_key, [""] * len(rows))
            set_sheet_rows(sheet_key, rows or [row_values_to_list(sheet_key, default_row_values(sheet_key))], source)
            refresh_filter_values()
            if sheet_key in ("litoloji", "spt", "pmt", "kaya", "numune"):
                try:
                    sondaj_names = [row.get("no") for row in collect_rows("sondajlar") if row.get("no")]
                    if sondaj_names:
                        sheet.dropdown_column(0, values=sondaj_names, set_value=None, edit_data=False)
                except Exception:
                    pass
            return sheet

        with perf_timer("workbook.tksheet_create_sheets"):
            for key, spec in sheet_defs.items():
                frame = ttk.Frame(nb)
                nb.add(frame, text=spec["title"])
                frame_to_key[str(frame)] = key
                sheet_frames[key] = frame
                ttk.Label(frame, text=f"{spec['title']} sayfası ilk açılışta yüklenecek.", padding=20).pack(anchor="center", expand=True)
        with perf_timer("workbook.tksheet_load_initial_data"):
            initial, source_nos = wb_build_initial_rows(self.veri, sheet_defs)
            for sheet_key in sheet_defs:
                rows = initial[sheet_key] or [row_values_to_list(sheet_key, default_row_values(sheet_key))]
                pending_rows[sheet_key] = normalize_data_rows(sheet_key, rows)
                pending_source_nos[sheet_key] = source_nos if sheet_key == "sondajlar" else [""] * len(rows)
            create_sheet("sondajlar")
        with perf_timer("workbook.tksheet_dropdown_setup"):
            pass
        with perf_timer("workbook.tksheet_initial_validate"):
            validate_workbook(show_status=False)

        def on_workbook_tab_changed(event=None):
            ensure_sheet(get_active_key())
            apply_sondaj_filter(show_status=False)

        nb.bind("<<NotebookTabChanged>>", on_workbook_tab_changed, add="+")

        filter_group = ttk.LabelFrame(top, text="Filtre", padding=(4, 2))
        excel_group = ttk.LabelFrame(top, text="Excel", padding=(4, 2))
        row_group = ttk.LabelFrame(top, text="Satır", padding=(4, 2))
        generate_group = ttk.LabelFrame(top, text="Uret", padding=(4, 2))
        control_group = ttk.LabelFrame(top, text="Kontrol", padding=(4, 2))
        apply_group = ttk.LabelFrame(top, text="Uygula", padding=(4, 2))
        for group in (filter_group, excel_group, row_group, generate_group, control_group):
            group.pack(side="left", padx=3, pady=1)
        apply_group.pack(side="right", padx=3, pady=1)

        ttk.Label(filter_group, text="Sondaj").pack(side="left", padx=(0, 3))
        filter_combo = ttk.Combobox(filter_group, textvariable=filter_var, values=sondaj_filter_values(), width=12, state="readonly")
        filter_combo.pack(side="left", padx=3)
        filter_combo_ref["widget"] = filter_combo
        filter_combo.bind("<<ComboboxSelected>>", lambda event=None: apply_sondaj_filter(show_status=True))
        tk.Button(filter_group, text="Filtrele", command=lambda: apply_sondaj_filter(show_status=True), bg="#D6EAF8", font=FONT_BOLD).pack(side="left", padx=3)
        tk.Button(excel_group, text="Al", command=import_workbook, bg="#2E86C1", fg="white", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(excel_group, text="Aktar", command=export_workbook, bg="#1E8449", fg="white", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(row_group, text="+", command=active_add_row, bg=COLOR_ACCENT, fg="white", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(row_group, text="Ekle", command=active_insert_row, bg="#5499C7", fg="white", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(row_group, text="Çoğalt", command=active_duplicate_rows, bg="#85C1E9", fg="#111", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(row_group, text="Sil", command=active_delete_rows, bg=COLOR_DANGER, fg="white", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(row_group, text="Sütun Temizle", command=active_clear_column, bg="#7F8C8D", fg="white", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(row_group, text="Akıllı", command=active_smart_row, bg="#AF7AC5", fg="white", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(generate_group, text="SPT", command=generate_spt_rows, bg="#F7DC6F", fg="#111", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(generate_group, text="Litoloji", command=generate_litoloji_missing_rows, bg="#FADBD8", fg="#111", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(control_group, text="Boyut", command=lambda: ensure_active_sheet().set_all_cell_sizes_to_text(), bg="#D5DBDB", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(control_group, text="Kontrol", command=lambda: validate_workbook(show_status=True), bg=COLOR_WARNING, fg="white", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(generate_group, text="Paket", command=generate_selected_sondaj_package, bg="#7DCEA0", fg="#111", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(generate_group, text="N30", command=lambda: (auto_spt_n30(), validate_workbook(show_status=True)), bg="#F9E79F", fg="#111", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(control_group, text="İlk Sorun", command=go_to_first_issue, bg="#F5B7B1", fg="#111", font=FONT_BOLD).pack(side="left", padx=2)
        ttk.Label(top, textvariable=wb_info_var, foreground="#1F618D").pack(side="left", padx=8)
        tk.Button(apply_group, text="Uygula", command=lambda: apply_workbook(False), bg=COLOR_SUCCESS, fg="white", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(apply_group, text="Kapat", command=lambda: apply_workbook(True), bg=COLOR_PRIMARY, fg="white", font=FONT_BOLD).pack(side="left", padx=2)
        self.set_status("Excel mod workbook acildi.", level="success")
