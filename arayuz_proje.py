# Dosya: RaporPro/arayuz_proje.py
import datetime
import copy
import hashlib
import json
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, Listbox, Toplevel

from harita_cikti import eski_paylasimli_temp_harita_yolu_mu
from sabitler import *
from yardimcilar import *
from performans import ERROR_LOG_PATH, PERF_LOG_PATH, log_exception, perf_timer, perf_tracked
from proje_motoru import format_hesap_ozeti, hesap_ozeti, proje_saglik_ozeti
from proje_arsiv import (
    arsiv_kaydi_ekle,
    arsiv_kaydi_sil,
    arsiv_kayitlari_yukle,
    biten_isler_kml_yaz,
    proje_merkez_koordinati,
)
from proje_surumleri import (
    VARSAYILAN_SURUM_SINIRI,
    surum_deposunu_kopyala,
    surum_kaydi_olustur,
)
from proje_sema import PROJE_SEMA_SURUMU, proje_verisini_migre_et
from kalite_kontrol import backup_project_file
from workbook_motoru import (
    WORKBOOK_SHEET_DEFS,
    build_initial_rows as wb_build_initial_rows,
    yeni_sondaj_sablonu as wb_yeni_sondaj_sablonu,
)
from uygulama_yollari import SOURCE_DIR, kullanici_yolu


APP_DIR = str(SOURCE_DIR)
RECENT_PROJECTS_PATH = str(
    kullanici_yolu("recent_projects.json", legacy=SOURCE_DIR / "recent_projects.json")
)


class ArayuzProjeMixin:
    def proje_kayit_imzasi(self, veri=None):
        payload = veri if veri is not None else self.veri
        try:
            text = json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
        except Exception:
            text = repr(payload)
        return hashlib.sha256(text.encode("utf-8", errors="replace")).hexdigest()

    def kayit_imzasi_guncelle(self, collect=False):
        if collect and hasattr(self, "e_kunye"):
            try:
                self.guncelle_veri_objesi(silent=True)
            except Exception as exc:
                log_exception("project.signature.collect", exc_value=exc)
        self._son_kayit_imzasi = self.proje_kayit_imzasi()
        return self._son_kayit_imzasi

    def proje_degisti_mi(self):
        if getattr(self, "_son_kayit_imzasi", None) is None:
            return True
        try:
            if hasattr(self, "e_kunye"):
                self.guncelle_veri_objesi(silent=True)
        except Exception as exc:
            log_exception("project.dirty.collect", exc_value=exc)
        return self.proje_kayit_imzasi() != getattr(self, "_son_kayit_imzasi", None)

    def proje_kaydedilmemis_yeni_mi(self):
        if self.aktif_dosya_yolu:
            return False
        try:
            varsayilan = self.varsayilan_veri_olustur()
            current = copy.deepcopy(self.veri)
            self.veri_eksikleri_tamamla(current, varsayilan)
            return self.proje_kayit_imzasi(current) != self.proje_kayit_imzasi(varsayilan)
        except Exception:
            return True

    def kaydedilmemis_degisiklik_onayi(self, hedef="programdan çıkış"):
        if not self.proje_degisti_mi():
            return True
        hedef_metni = str(hedef or "devam etme").strip()
        if self.proje_kilitli_mi():
            msg = (
                "Projede kaydedilmemiş değişiklikler var ancak proje kilitli olduğu için kaydedilemez.\n\n"
                f"{hedef_metni.capitalize()} işlemine kaydetmeden devam edilsin mi?"
            )
            return bool(messagebox.askyesno("Kaydedilmemiş Değişiklikler", msg))
        secim = messagebox.askyesnocancel(
            "Kaydedilmemiş Değişiklikler",
            "Projede kaydedilmemiş değişiklikler var.\n\n"
            f"{hedef_metni.capitalize()} işleminden önce kaydedilsin mi?\n\n"
            "Evet: Kaydet ve devam et\n"
            "Hayır: Kaydetmeden devam et\n"
            "İptal: Programa dön",
        )
        if secim is None:
            return False
        if secim is False:
            return True
        return bool(self.veri_kaydet())

    def recent_projects_yukle(self):
        try:
            if not os.path.exists(RECENT_PROJECTS_PATH):
                return []
            with open(RECENT_PROJECTS_PATH, "r", encoding="utf-8") as f:
                payload = json.load(f)
            items = payload.get("projects", payload) if isinstance(payload, dict) else payload
            if not isinstance(items, list):
                return []
            recent = []
            seen = set()
            for item in items:
                path = item.get("path") if isinstance(item, dict) else item
                if not path:
                    continue
                path = os.path.abspath(str(path))
                key = os.path.normcase(path)
                if key in seen or not os.path.exists(path):
                    continue
                seen.add(key)
                recent.append(path)
            return recent[:8]
        except Exception as exc:
            log_exception("recent_projects.load", exc_value=exc)
            return []

    def recent_projects_kaydet(self):
        try:
            atomic_json_dump({"projects": self.recent_projects[:8]}, RECENT_PROJECTS_PATH, indent=2, ensure_ascii=False)
        except Exception as exc:
            log_exception("recent_projects.save", exc_value=exc)

    def recent_project_ekle(self, path):
        if not path:
            return
        path = os.path.abspath(str(path))
        current = getattr(self, "recent_projects", [])
        key = os.path.normcase(path)
        cleaned = []
        seen = {key}
        for item in current:
            item_path = os.path.abspath(str(item))
            item_key = os.path.normcase(item_path)
            if item_key in seen or not os.path.exists(item_path):
                continue
            seen.add(item_key)
            cleaned.append(item_path)
        self.recent_projects = [path] + cleaned
        self.recent_projects_kaydet()

    def proje_dosyasi_yukle(self, dosya_yolu):
        with perf_timer("project.open_read_apply"):
            with open(dosya_yolu, 'r', encoding='utf-8') as f:
                yuklenen_veri = json.load(f)
            yuklenen_veri, migrasyon = self.proje_verisini_hazirla(yuklenen_veri)
            self.veri = yuklenen_veri
            self.aktif_dosya_yolu = dosya_yolu
            self.doldur_arayuz()
            if migrasyon.degisti:
                self._son_kayit_imzasi = None
            else:
                self.kayit_imzasi_guncelle(collect=True)
        self.proje_baslik_guncelle()
        self.recent_project_ekle(dosya_yolu)
        if migrasyon.degisti:
            self.set_status(
                f"Eski proje v{migrasyon.onceki_surum} → v{migrasyon.yeni_surum} olarak hazırlandı; "
                "kalıcılaştırmak için Kaydet'i kullanın.",
                level="warning",
            )
            self.set_save_indicator("Proje yapısı güncellendi: kaydedilmedi", "warning")
        else:
            self.set_status(f"Proje açıldı: {dosya_yolu}", level="success")
        self.proje_kilit_durumunu_goster()

    def son_projeler_penceresi(self):
        self.recent_projects = self.recent_projects_yukle()
        win = Toplevel(self.root)
        self.pencere_hazirla(win, "Son Projeler", "720x360", (620, 300), modal=True)

        ttk.Label(win, text="Son açılan projeler", font=FONT_HEADER).pack(anchor="w", padx=14, pady=(12, 6))
        list_frame = ttk.Frame(win)
        list_frame.pack(fill="both", expand=True, padx=14, pady=6)
        lb = Listbox(list_frame, height=9, font=("Consolas", 9))
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=scrollbar.set)
        lb.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def refresh():
            lb.delete(0, tk.END)
            if not self.recent_projects:
                lb.insert(tk.END, "Son proje yok.")
                return
            for path in self.recent_projects:
                status = "" if os.path.exists(path) else "  [bulunamadı]"
                lb.insert(tk.END, f"{os.path.basename(path):<32} {path}{status}")

        def selected_path():
            if not self.recent_projects:
                return None
            sel = lb.curselection()
            if not sel:
                return None
            idx = sel[0]
            return self.recent_projects[idx] if idx < len(self.recent_projects) else None

        def open_selected(event=None):
            path = selected_path()
            if not path:
                return "break"
            if not os.path.exists(path):
                messagebox.showwarning("Son Projeler", "Bu proje dosyası artık bulunamıyor. Listeden kaldırıldı.")
                self.recent_projects = [p for p in self.recent_projects if os.path.normcase(os.path.abspath(p)) != os.path.normcase(os.path.abspath(path))]
                self.recent_projects_kaydet()
                refresh()
                return "break"
            if not self.kaydedilmemis_degisiklik_onayi("başka proje açma"):
                return "break"
            try:
                self.proje_dosyasi_yukle(path)
                win.destroy()
            except Exception as exc:
                messagebox.showerror("Hata", f"Dosya açılamadı:\n{exc}")
            return "break"

        def remove_selected():
            path = selected_path()
            if not path:
                return
            self.recent_projects = [p for p in self.recent_projects if os.path.normcase(os.path.abspath(p)) != os.path.normcase(os.path.abspath(path))]
            self.recent_projects_kaydet()
            refresh()

        lb.bind("<Double-Button-1>", open_selected)
        lb.bind("<Return>", open_selected)
        btns = ttk.Frame(win)
        btns.pack(fill="x", padx=14, pady=(4, 12))
        tk.Button(btns, text="Aç", command=open_selected, bg=COLOR_PRIMARY, fg="white", font=FONT_BOLD).pack(side="left", padx=(0, 6))
        tk.Button(btns, text="Listeden Kaldır", command=remove_selected, bg="#ECF0F1").pack(side="left", padx=6)
        tk.Button(btns, text="Kapat", command=win.destroy, bg="#ECF0F1").pack(side="right")
        refresh()
        if self.recent_projects:
            lb.selection_set(0)

    def kisayol_calistir(self, command):
        try:
            command()
        except Exception as exc:
            log_exception("shortcut.run", exc_value=exc)
            self.set_status(f"Kısayol hatası: {exc}", level="error")
        return "break"

    def kur_kisayollar(self):
        bindings = {
            "<Control-s>": self.veri_kaydet,
            "<Control-o>": self.proje_ac,
            "<Control-n>": self.yeni_proje,
            "<Control-Shift-S>": self.proje_farkli_kaydet,
            "<Control-w>": self.veri_giris_workbook_tksheet_ac,
            "<Control-r>": self.rapor_on_kontrol,
            "<Control-k>": self.kesit_secim_penceresi,
            "<Control-h>": self.son_projeler_penceresi,
            "<Control-Shift-H>": self.surum_gecmisi_penceresi,
            "<F5>": self.ozet_yenile,
            "<F9>": self.final_kontrol_penceresi,
        }
        for sequence, command in bindings.items():
            try:
                self.root.bind_all(sequence, lambda event, cmd=command: self.kisayol_calistir(cmd), add="+")
            except Exception as exc:
                log_exception("shortcut.bind", exc_value=exc)

    def varsayilan_veri_olustur(self):
        default = {
            "schema_version": PROJE_SEMA_SURUMU,
            "kunye": {"sahibi":"", "il":"", "ilce":"", "mah":"", "mev":"", "paf":"", "ada":"", "par":""},
            "bina": {"kul":"", "sinif":"", "onem":"", "malz":"", "bod":"", "kat":"", "plan":"", "yukseklik":"", "yukseklik_sinif":"", "temel_alan":"", "ins":"", "der":"", "gqe_min":"", "gqe_max":"", "gqe_ort":"", "comb_min":"", "comb_max":"", "comb_ort":"", "ysinif":"", "tem":"", "coklu_blok": False, "bloklar": []},
            "arazi": {"kot":"", "yon":"", "egim":"", "min":"", "max":"", "ort":"", "imar_alani":"", "imar_durumu":"", "zemin":"", "kategori": "", "pga":"", "alan_y": "", "alan_x": ""},
            "sondaj": [],
            "jeofizik": {"tarih": "", "ss_list": [], "mt_list": []},
            "harita_cizimleri": {"vaziyet": {}, "jeoloji": {}, "yerbuldurur": {}},
            "lab_sheet": {"rows": []},
            "jeofizik_sheet": {"rows": []},
            "kesit_ayarlari": {},
            "ek_icerikleri": {"normal": {}, "arazi_deneyli": {}},
            "proje_durumu": {"tamamlandi": False, "kilitli": False, "tamamlanma_tarihi": "", "arsiv_notu": ""},
            "ayarlar": {
                "firma_adi": "UB ZEMIN MUHENDISLIK",
                "log_baslik": "SONDAJ LOGU",
                "sorumlu_muhendis_unvan": "Sorumlu Jeoloji Muhendisi",
                "sorumlu_muhendis": "Gökalp DOĞAN",
                "sondor_belge_baslik": "Sondor Belge No",
                "sondor_belge": "Murat ERÇELİK 3629",
                "makine_metodu": "Rotary / Burgusuz",
                "spt_sahmerdan": "Otomatik",
                "sondaj_turu": "Zemin",
                "delgi_capi": "76mm",
                "varsayilan_word_path": "",
                "varsayilan_cikti_klasor": "",
                "log_export_klasor": "",
                "log_export_format": "JPG",
                "log_export_dpi": "300",
                "log_export_prefix": "Log",
                "cikti_merkezi_klasor": "",
                "cikti_merkezi_format": "JPG",
                "cikti_merkezi_dpi": "300",
                "rapor_buyuk_baslik_yeni_sayfa": "1",
                "taahhut_ilgili_idare": "",
                "taahhut_tarih": "",
                "ek_tutanak_path": "",
                "ek_arazi_deneyli_path": "",
                "tutanak_sablon_path": "",
                "tutanak_sondaj_firma": "Kale Detay Sondaj",
                "tutanak_uygulama_sekli": "Burgusuz/Sulu",
                "tutanak_sondaj_makinesi": "SMK-500",
                "tutanak_jeofizik_cihaz": "GEODE",
                "tutanak_jeofon": "3,0m - 4,5 Hz",
                "tutanak_offset": "3,0m",
                "tutanak_kanal_sayisi": "12",
                "tutanak_kaynak": "Balyoz",
                "taahhut_jeoloji_ad": "Gökalp DOĞAN",
                "taahhut_jeoloji_sicil": "7400",
                "taahhut_jeoloji_unvan": "JEOLOJİ MÜHENDİSİ",
                "taahhut_jeoloji_imza_unvan": "Jeoloji Mühendisi",
                "taahhut_jeoloji_adres": "İsmetpaşa Mh. Hasan Mevsuf Sk. No :4 Da:5",
                "taahhut_jeoloji_telefon": "0 545 639 90 62",
                "taahhut_jeofizik_ad": "Suat ERGİN",
                "taahhut_jeofizik_sicil": "1982",
                "taahhut_jeofizik_unvan": "JEOFİZİK MÜHENDİSİ",
                "taahhut_jeofizik_imza_unvan": "Jeofizik Mühendisi",
                "taahhut_jeofizik_adres": "İsmetpaşa Mh. Hasan Mevsuf Sk. No :4 Da:5",
                "taahhut_jeofizik_telefon": "0 532 281 12 95",
                "yedek_sayisi": "10",
                "surum_gecmisi_sayisi": str(VARSAYILAN_SURUM_SINIRI),
                "spt_guven_esigi": "90",
                "spt_auto_pro": "1"
            },
            "dosyalar": {"kml_path": None, "word_path": None, "lab_excel_path": None, "jeo_excel_path": None, "img_yer": None, "img_tkgm": None, "img_pga": None, "img_mjh": None, "word_img_sondaj": None, "word_img_jeofizik": None}
        }
        return default

    def proje_verisini_hazirla(self, veri):
        return proje_verisini_migre_et(veri, self.varsayilan_veri_olustur())

    def veri_eksikleri_tamamla(self, hedef, varsayilan):
        for key, value in varsayilan.items():
            if key not in hedef:
                hedef[key] = value
            elif isinstance(value, dict) and isinstance(hedef[key], dict):
                self.veri_eksikleri_tamamla(hedef[key], value)
        return hedef

    def _dosya_yolu_coz(self, path):
        if path is None:
            return None
        path = str(path).strip().strip('"')
        if path in {"", "-", "None", "none", "null"}:
            return None
        path = os.path.expandvars(os.path.expanduser(path))
        if os.path.exists(path):
            return path
        if self.aktif_dosya_yolu:
            proje_klasoru = os.path.dirname(os.path.abspath(self.aktif_dosya_yolu))
            if not os.path.isabs(path):
                aday = os.path.join(proje_klasoru, path)
                if os.path.exists(aday):
                    return aday
            basename = os.path.basename(path)
            if basename:
                aday = os.path.join(proje_klasoru, basename)
                if os.path.exists(aday):
                    return aday
        return path

    def _dosya_yolu_al(self, dosyalar, key, *legacy_keys):
        kaynaklar = [dosyalar if isinstance(dosyalar, dict) else {}, self.veri if isinstance(self.veri, dict) else {}]
        for kaynak in kaynaklar:
            for aday_key in (key, *legacy_keys):
                value = kaynak.get(aday_key)
                resolved = self._dosya_yolu_coz(value)
                if resolved:
                    return resolved
        return None

    def _harita_gorsel_yolu_al(self, dosyalar, key, *legacy_keys):
        path = self._dosya_yolu_al(dosyalar, key, *legacy_keys)
        if eski_paylasimli_temp_harita_yolu_mu(path):
            return None
        return path

    def _harita_gorsel_yolu_kaydet(self, path):
        return None if eski_paylasimli_temp_harita_yolu_mu(path) else path

    def kml_etiket_guncelle(self):
        path = getattr(self, "kml_path", None)
        secili = bool(path)
        mevcut = bool(secili and os.path.exists(path))
        if mevcut:
            text = f"KML Sınır: {os.path.basename(path)}"
            color = "#27AE60"
        elif secili:
            text = f"KML Sınır: {os.path.basename(path)} (bulunamadı)"
            color = COLOR_WARNING
        else:
            text = "KML Sınır: Seçilmedi"
            color = "#333"
        if hasattr(self, 'lbl_kml_top'):
            self.lbl_kml_top.config(text=text, fg=color)
        if hasattr(self, "harita_durum_yenile"):
            self.harita_durum_yenile()

    @perf_tracked("project.load_default")
    def veri_yukle(self):
        default = self.varsayilan_veri_olustur()
        return default

    def get_yedek_sayisi(self):
        try:
            keep = int(str(self.veri.get("ayarlar", {}).get("yedek_sayisi", "10")).strip())
            return max(1, keep)
        except Exception:
            return 10

    def get_surum_gecmisi_sayisi(self):
        try:
            keep = int(str(self.veri.get("ayarlar", {}).get("surum_gecmisi_sayisi", VARSAYILAN_SURUM_SINIRI)).strip())
            return max(5, min(keep, 250))
        except Exception:
            return VARSAYILAN_SURUM_SINIRI

    def proje_surum_kaydi_yaz(self, neden, force=False, source="manual"):
        """Kayıt başarılı olduktan sonra sürüm kopyasını oluşturur; ana kaydı engellemez."""
        if not self.aktif_dosya_yolu:
            return None, False
        try:
            record, created = surum_kaydi_olustur(
                self.aktif_dosya_yolu,
                self.veri,
                reason=neden,
                keep=self.get_surum_gecmisi_sayisi(),
                force=force,
                source=source,
            )
            if created:
                self.set_status("Yeni proje sürümü geçmişe eklendi.", level="info")
            return record, created
        except Exception as exc:
            log_exception("project.version.write", exc_value=exc)
            self.set_status(f"Sürüm geçmişi uyarısı: {exc}", level="warning")
            return None, False

    def proje_durumu(self):
        return self.veri.setdefault("proje_durumu", {"tamamlandi": False, "kilitli": False, "tamamlanma_tarihi": "", "arsiv_notu": ""})

    def proje_kilitli_mi(self):
        return bool(self.proje_durumu().get("kilitli"))

    def proje_baslik_guncelle(self):
        if self.aktif_dosya_yolu:
            title = f"Zemin Rapor Pro - {os.path.basename(self.aktif_dosya_yolu)}"
        else:
            title = "Zemin Rapor Pro - Yeni Proje"
        if self.proje_kilitli_mi():
            title += " [KİLİTLİ]"
        self.root.title(title)

    def proje_kilit_durumunu_goster(self):
        durum = self.proje_durumu()
        if durum.get("kilitli"):
            tarih = durum.get("tamamlanma_tarihi") or "-"
            self.set_save_indicator(f"Proje kilitli: {tarih}", "warning")
        else:
            self.set_save_indicator("Proje açıldı", "info")

    def proje_tamamlandi_kilitle(self):
        self.guncelle_veri_objesi()
        if not self.aktif_dosya_yolu:
            if not messagebox.askyesno("Proje Kilitle", "Projeyi kilitlemeden önce kaydetmek gerekir. Şimdi kaydedilsin mi?"):
                return
            self.proje_farkli_kaydet()
            if not self.aktif_dosya_yolu:
                return
        if self.proje_kilitli_mi():
            messagebox.showinfo("Proje Kilitli", "Bu proje zaten tamamlandı olarak kilitli.")
            return
        lat, lon = proje_merkez_koordinati(self.veri)
        tarih = datetime.datetime.now().isoformat(timespec="seconds")
        self.proje_durumu().update({
            "tamamlandi": True,
            "kilitli": True,
            "tamamlanma_tarihi": tarih,
        })
        try:
            arsiv_kaydi_ekle(self.veri, self.aktif_dosya_yolu, kml_path=getattr(self, "kml_path", None))
        except Exception as exc:
            log_exception("project.archive.add", exc_value=exc)
            self.set_status(f"Arşiv kaydı oluşturulamadı: {exc}", level="warning")
        try:
            self._kilitli_kayda_izin_ver = True
            self.veri_kaydet()
        finally:
            self._kilitli_kayda_izin_ver = False
        self.proje_baslik_guncelle()
        msg = "Proje tamamlandı olarak kilitlendi ve arşiv listesine eklendi."
        if not lat or not lon:
            msg += "\n\nNot: KML haritasında görünmesi için Arazi sekmesinde proje merkezi veya sondaj koordinatı gerekir."
        messagebox.showinfo("Proje Kilitlendi", msg)

    def proje_kilidini_kaldir(self):
        if not self.proje_kilitli_mi():
            messagebox.showinfo("Proje Kilidi", "Bu proje kilitli değil.")
            return
        if not messagebox.askyesno("Kilidi Kaldır", "Proje kilidi kaldırılsın mı? Bundan sonra proje tekrar düzenlenip kaydedilebilir."):
            return
        self.proje_durumu().update({"tamamlandi": False, "kilitli": False})
        try:
            arsiv_kaydi_sil(self.aktif_dosya_yolu)
        except Exception as exc:
            log_exception("project.archive.remove", exc_value=exc)
        try:
            self._kilitli_kayda_izin_ver = True
            self.veri_kaydet()
        finally:
            self._kilitli_kayda_izin_ver = False
        self.proje_baslik_guncelle()
        self.set_save_indicator("Kilit kaldırıldı", "info")
        messagebox.showinfo("Kilidi Kaldır", "Proje kilidi kaldırıldı.")

    def biten_isler_kml_olustur(self):
        records = arsiv_kayitlari_yukle()
        if not records:
            messagebox.showinfo("Biten İşler KML", "Henüz tamamlandı olarak kilitlenmiş proje yok.")
            return
        path = filedialog.asksaveasfilename(
            title="Biten işler KML kaydet",
            defaultextension=".kml",
            initialfile="RaporPro_Biten_Isler.kml",
            filetypes=[("KML Dosyası", "*.kml"), ("Tüm Dosyalar", "*.*")],
        )
        if not path:
            return
        try:
            info = biten_isler_kml_yaz(records, path)
            self.set_status(f"Biten işler KML oluşturuldu: {os.path.basename(path)}", level="success")
            messagebox.showinfo(
                "Biten İşler KML",
                f"KML oluşturuldu:\n{path}\n\nHaritaya eklenen iş: {info['written']}\nKoordinatı olmadığı için atlanan: {info['skipped']}",
            )
        except Exception as exc:
            log_exception("project.archive.kml", exc_value=exc)
            messagebox.showerror("Biten İşler KML", f"KML oluşturulamadı:\n{exc}")

    @perf_tracked("project.save")
    def veri_kaydet(self):
        self.guncelle_veri_objesi()
        if self.proje_kilitli_mi() and not getattr(self, "_kilitli_kayda_izin_ver", False):
            messagebox.showwarning("Proje Kilitli", "Bu proje tamamlandı olarak kilitli. Kaydetmek için önce Proje > Proje Kilidini Kaldır komutunu kullanın.")
            self.set_save_indicator("Kilitli: kaydedilmedi", "warning")
            return False
        if self.aktif_dosya_yolu:
            try:
                backup_path, backup_error = backup_project_file(self.aktif_dosya_yolu, keep=self.get_yedek_sayisi())
                if backup_error:
                    self.set_status(f"Yedekleme uyarısı: {backup_error}", level="warning")
                atomic_json_dump(self.veri, self.aktif_dosya_yolu, indent=4, ensure_ascii=False)
                self.proje_surum_kaydi_yaz("Proje kaydedildi")
                self.kayit_imzasi_guncelle()
                self.set_status(f"Kaydedildi: {os.path.basename(self.aktif_dosya_yolu)}", level="success")
                self.last_save_time = datetime.datetime.now()
                self.set_save_indicator(f"Son kayıt: {self.last_save_time.strftime('%H:%M')}", "success")
                self.recent_project_ekle(self.aktif_dosya_yolu)
                if backup_path:
                    self.set_status(f"Yedek oluşturuldu: {os.path.basename(backup_path)}", level="info")
                return True
            except Exception as e:
                self.set_status(f"Kayıt Hatası: {str(e)}", level="error")
                self.set_save_indicator("Kayıt hatası", "error")
                return False
        else:
            return self.proje_farkli_kaydet()

    @perf_tracked("project.save_as")
    def proje_farkli_kaydet(self):
        self.guncelle_veri_objesi()
        onceki_proje_yolu = self.aktif_dosya_yolu
        if self.proje_kilitli_mi() and not getattr(self, "_kilitli_kayda_izin_ver", False):
            messagebox.showwarning("Proje Kilitli", "Bu proje kilitli. Farklı kaydetmek için önce kilidi kaldırın.")
            self.set_save_indicator("Kilitli: farklı kaydedilmedi", "warning")
            return False
        proje_adi = self.veri["kunye"].get("sahibi", "Yeni_Proje")
        if not proje_adi: proje_adi = "Zemin_Etud_Projesi"
        varsayilan_isim = f"{proje_adi}.json"

        dosya_yolu = filedialog.asksaveasfilename(
            defaultextension=".json",
            initialfile=varsayilan_isim,
            filetypes=[("RaporPro Projesi", "*.json"), ("Tüm Dosyalar", "*.*")],
            title="Projeyi Farklı Kaydet"
        )
        if dosya_yolu:
            try:
                backup_path, backup_error = backup_project_file(dosya_yolu, keep=self.get_yedek_sayisi())
                if backup_error:
                    self.set_status(f"Yedekleme uyarısı: {backup_error}", level="warning")
                atomic_json_dump(self.veri, dosya_yolu, indent=4, ensure_ascii=False)
                if onceki_proje_yolu and os.path.normcase(os.path.abspath(onceki_proje_yolu)) != os.path.normcase(os.path.abspath(dosya_yolu)):
                    try:
                        surum_deposunu_kopyala(onceki_proje_yolu, dosya_yolu)
                    except Exception as exc:
                        log_exception("project.version.copy", exc_value=exc)
                self.aktif_dosya_yolu = dosya_yolu
                self.proje_surum_kaydi_yaz("Farklı kaydet" if onceki_proje_yolu else "İlk proje kaydı")
                self.root.title(f"Zemin Rapor Pro - {os.path.basename(dosya_yolu)}")
                self.kayit_imzasi_guncelle()
                self.set_status(f"Yeni proje olarak kaydedildi: {dosya_yolu}", level="success")
                self.last_save_time = datetime.datetime.now()
                self.set_save_indicator(f"Son kayıt: {self.last_save_time.strftime('%H:%M')}", "success")
                self.recent_project_ekle(dosya_yolu)
                if backup_path:
                    self.set_status(f"Yedek oluşturuldu: {os.path.basename(backup_path)}", level="info")
                return True
            except Exception as e:
                messagebox.showerror("Hata", f"Dosya kaydedilemedi:\n{str(e)}")
                self.set_save_indicator("Kayıt hatası", "error")
                return False
        return False

    @perf_tracked("project.open")
    def proje_ac(self):
        dosya_yolu = filedialog.askopenfilename(
            filetypes=[("RaporPro Projesi", "*.json"), ("Tüm Dosyalar", "*.*")],
            title="Proje Aç"
        )
        if dosya_yolu:
            if not self.kaydedilmemis_degisiklik_onayi("başka proje açma"):
                return
            try:
                self.proje_dosyasi_yukle(dosya_yolu)
            except Exception as e:
                messagebox.showerror("Hata", f"Dosya açılamadı:\n{str(e)}")

    @perf_tracked("project.new")
    def yeni_proje(self):
        if not self.kaydedilmemis_degisiklik_onayi("yeni proje oluşturma"):
            return
        self.yeni_proje_sihirbazi()

    def reset_dosya_baglantilari(self):
        self.aktif_dosya_yolu = None
        self.kml_path = None
        self.word_path = None
        self.lab_excel_path = None
        self.jeo_excel_path = None
        self.img_yer = None
        self.img_tkgm = None
        self.img_pga = None
        self.img_mjh = None
        self.word_img_sondaj = None
        self.word_img_jeofizik = None
        self.ek_tutanak_path = None
        self.ek_arazi_deneyli_path = None
        self.last_output_quality_report = None

    def yeni_proje_sihirbazi(self):
        mevcut_ayarlar = self.veri.get("ayarlar", {}).copy()
        win = Toplevel(self.root)
        self.pencere_hazirla(win, "Yeni Proje Sihirbazı", "560x430", (520, 390), modal=True)
        body = ttk.Frame(win, padding=16)
        body.pack(fill="both", expand=True)

        fields = {}
        defaults = [
            ("Proje adi", "sahibi", ""),
            ("Il", "il", ""),
            ("Ilce", "ilce", ""),
            ("Mahalle", "mah", ""),
            ("Sondaj sayisi", "count", "3"),
            ("Varsayilan derinlik (m)", "depth", "15.0"),
        ]
        for row, (label, key, value) in enumerate(defaults):
            ttk.Label(body, text=label).grid(row=row, column=0, sticky="e", padx=6, pady=6)
            entry = ttk.Entry(body, width=36)
            entry.grid(row=row, column=1, sticky="ew", padx=6, pady=6)
            entry.insert(0, value)
            fields[key] = entry

        kesit_var = tk.StringVar(value="line_projection")
        ttk.Label(body, text="Kesit mantigi").grid(row=len(defaults), column=0, sticky="e", padx=6, pady=6)
        ttk.Combobox(body, textvariable=kesit_var, values=("line_projection", "true_distance", "schematic"), state="readonly", width=34).grid(row=len(defaults), column=1, sticky="ew", padx=6, pady=6)

        word_var = tk.StringVar(value=mevcut_ayarlar.get("varsayilan_word_path", ""))
        ttk.Label(body, text="Word sablonu").grid(row=len(defaults) + 1, column=0, sticky="e", padx=6, pady=6)
        word_entry = ttk.Entry(body, textvariable=word_var, width=36)
        word_entry.grid(row=len(defaults) + 1, column=1, sticky="ew", padx=6, pady=6)
        tk.Button(body, text="Seç", command=lambda: self._ayar_dosya_sec(word_entry, [("Word", "*.docx")]), bg="#ECF0F1").grid(row=len(defaults) + 1, column=2, padx=4, pady=6)

        demo_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(body, text="Örnek litoloji ve SPT satırlarıyla başlat", variable=demo_var).grid(row=len(defaults) + 2, column=0, columnspan=2, sticky="w", pady=(8, 2))
        body.columnconfigure(1, weight=1)

        def olustur():
            count = max(0, int(safe_float(fields["count"].get()) or 0))
            depth = safe_float(fields["depth"].get()) or 15.0
            self.veri = self.varsayilan_veri_olustur()
            self.veri["ayarlar"].update(mevcut_ayarlar)
            self.veri["kunye"].update({
                "sahibi": fields["sahibi"].get().strip(),
                "il": fields["il"].get().strip(),
                "ilce": fields["ilce"].get().strip(),
                "mah": fields["mah"].get().strip(),
            })
            self.veri["sondaj"] = []
            for idx in range(count):
                sondaj = wb_yeni_sondaj_sablonu(idx)
                sondaj["der"] = f"{depth:.1f}"
                if demo_var.get():
                    split = max(1.0, depth * 0.45)
                    sondaj["k"] = f"{100 - idx * 1.2:.2f}"
                    sondaj["litoloji"] = [[0, f"{split:.2f}", "Killi kum"], [f"{split:.2f}", f"{depth:.2f}", "Siltli kil"]]
                    d = 1.5
                    while d <= depth + 0.01:
                        sondaj["spt"].append([f"{d:.2f}", "2", "4", "5", "9"])
                        d += 1.5
                self.veri["sondaj"].append(sondaj)
            self.veri["kesit_ayarlari"] = {"mode": kesit_var.get(), "selected_sondajlar": [s.get("no", "") for s in self.veri["sondaj"]]}
            if word_var.get().strip():
                self.veri["dosyalar"]["word_path"] = word_var.get().strip()
            self.reset_dosya_baglantilari()
            self.word_path = word_var.get().strip() or None
            self.doldur_arayuz()
            self.root.title("Zemin Rapor Pro - Yeni Proje")
            self.set_save_indicator("Yeni proje: kaydedilmedi", "warning")
            self.set_status("Yeni proje sihirbazla oluşturuldu.", level="success")
            win.destroy()

        btns = ttk.Frame(body)
        btns.grid(row=len(defaults) + 3, column=0, columnspan=3, sticky="ew", pady=(18, 0))
        tk.Button(btns, text="Oluştur", command=olustur, bg=COLOR_SUCCESS, fg="white", font=FONT_BOLD).pack(side="right", padx=4)
        tk.Button(btns, text="Vazgeç", command=win.destroy, bg="#ECF0F1").pack(side="right", padx=4)

    def ornek_proje_yukle(self):
        if not messagebox.askyesno("Örnek Proje", "Mevcut çalışma kaydedilmemişse kaybolabilir. Örnek proje yüklensin mi?"):
            return
        mevcut_ayarlar = self.veri.get("ayarlar", {}).copy()
        self.veri = self.varsayilan_veri_olustur()
        self.veri["ayarlar"].update(mevcut_ayarlar)
        self.veri["kunye"].update({"sahibi": "Örnek Zemin Etüdü", "il": "İstanbul", "ilce": "Kadıköy", "mah": "Örnek Mahalle", "ada": "123", "par": "45"})
        self.veri["bina"].update({"kul": "Konut", "kat": "5", "bod": "1", "temel_alan": "450", "der": "3.0"})
        self.veri["arazi"].update({"yon": "Guney", "egim": "5", "min": "98", "max": "104", "ort": "101", "pga": "0.40"})
        coords = [("40.990100", "29.030100"), ("40.990240", "29.030420"), ("40.990420", "29.030760")]
        lithologies = [
            [[0, 1.2, "Bitkisel toprak"], [1.2, 6.0, "Killi kum"], [6.0, 15.0, "Siltli kil"]],
            [[0, 1.0, "Dolgu"], [1.0, 5.5, "Kumlu kil"], [5.5, 15.0, "Cakilli kum"]],
            [[0, 1.4, "Bitkisel toprak"], [1.4, 7.0, "Siltli kil"], [7.0, 15.0, "Kum"]],
        ]
        self.veri["sondaj"] = []
        for idx in range(3):
            sondaj = wb_yeni_sondaj_sablonu(idx)
            sondaj.update({"der": "15.0", "k": f"{102 - idx * 1.4:.2f}", "y": coords[idx][0], "x": coords[idx][1]})
            sondaj["litoloji"] = lithologies[idx]
            sondaj["spt"] = [["1.50", "2", "3", "4", "7"], ["3.00", "3", "5", "6", "11"], ["4.50", "4", "6", "8", "14"], ["6.00", "5", "8", "10", "18"], ["7.50", "6", "10", "12", "22"], ["9.00", "8", "12", "15", "27"], ["10.50", "10", "15", "18", "33"], ["12.00", "12", "18", "20", "38"], ["13.50", "15", "20", "25", "45"]]
            self.veri["sondaj"].append(sondaj)
        self.veri["jeofizik"]["ss_list"] = [{"ad": "SS-1", "coords": ["40.990100", "29.030100", "40.990240", "29.030420", "40.990420", "29.030760"], "layers": [{"h": "6", "vp": "650", "vs": "220"}, {"h": "", "vp": "1200", "vs": "420"}]}]
        self.veri["jeofizik"]["mt_list"] = [{"no": "MT-1", "y": "40.990240", "x": "29.030420", "freq": "2.5", "to": "0.40", "ta": "0.28", "tb": "0.62", "hv": "3.2", "sure": "30"}]
        self.veri["kesit_ayarlari"] = {"mode": "line_projection", "selected_sondajlar": ["SK-1", "SK-2", "SK-3"], "line_start_no": "SK-1", "line_start_y": coords[0][0], "line_start_x": coords[0][1], "line_end_no": "SK-3", "line_end_y": coords[2][0], "line_end_x": coords[2][1]}
        self.reset_dosya_baglantilari()
        self.doldur_arayuz()
        self.root.title("Zemin Rapor Pro - Örnek Proje")
        self.set_save_indicator("Örnek proje: kaydedilmedi", "warning")
        self.set_status("Örnek proje yüklendi.", level="success")

    def proje_sablon_penceresi(self):
        win = Toplevel(self.root)
        self.pencere_hazirla(win, "Proje Şablonları", "520x360", (480, 320), modal=True)
        ttk.Label(win, text="Yeni proje şablonu seçin", font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=16, pady=(16, 8))
        templates = [
            ("Standart 3x15", 3, 15.0, "3 sondaj, 15 m"),
            ("Orta 5x20", 5, 20.0, "5 sondaj, 20 m"),
            ("Detay 8x25", 8, 25.0, "8 sondaj, 25 m"),
            ("Sadece Rapor", 0, 0.0, "Sondajsız rapor dosya bağlantıları"),
            ("Jeofizikli Proje", 3, 15.0, "3 sondaj + jeofizik hazır alan"),
        ]
        selected = tk.StringVar(value=templates[0][0])
        for label, count, depth, desc in templates:
            ttk.Radiobutton(win, text=f"{label} - {desc}", variable=selected, value=label).pack(anchor="w", padx=22, pady=5)

        def apply_template():
            template = next(item for item in templates if item[0] == selected.get())
            self.proje_sablonu_uygula(template)
            win.destroy()

        tk.Button(win, text="Şablonu Uygula", command=apply_template, bg=COLOR_PRIMARY, fg="white", font=FONT_BOLD).pack(fill="x", padx=18, pady=18)

    @perf_tracked("project.template_apply")
    def proje_sablonu_uygula(self, template):
        label, count, depth, desc = template
        mevcut_ayarlar = self.veri.get("ayarlar", {}).copy()
        self.veri = self.varsayilan_veri_olustur()
        self.veri["ayarlar"].update(mevcut_ayarlar)
        self.veri["sondaj"] = []
        for idx in range(count):
            sondaj = wb_yeni_sondaj_sablonu(idx)
            sondaj["der"] = f"{depth:.1f}"
            self.veri["sondaj"].append(sondaj)
        if label == "Jeofizikli Proje":
            self.veri["jeofizik"]["ss_list"] = [{"ad": "SS-1", "coords": [""] * 6, "layers": []}]
            self.veri["jeofizik"]["mt_list"] = [{"no": "MT-1", "y": "", "x": ""}]
        self.aktif_dosya_yolu = None
        self.doldur_arayuz()
        self.root.title(f"Zemin Rapor Pro - {label}")
        self.set_status(f"Proje şablonu uygulandı: {label}", level="success")

    def _dosya_map(self):
        return {
            "word_path": self.word_path,
            "lab_excel_path": self.lab_excel_path,
            "jeo_excel_path": self.jeo_excel_path,
            "kml_path": self.kml_path,
            "img_yer": self.img_yer,
            "img_tkgm": self.img_tkgm,
            "img_pga": self.img_pga,
            "img_mjh": self.img_mjh,
            "word_img_sondaj": self.word_img_sondaj,
            "word_img_jeofizik": self.word_img_jeofizik,
            "ek_tutanak_path": getattr(self, "ek_tutanak_path", None) or self.veri.get("ayarlar", {}).get("ek_tutanak_path"),
            "ek_arazi_deneyli_path": getattr(self, "ek_arazi_deneyli_path", None) or self.veri.get("ayarlar", {}).get("ek_arazi_deneyli_path"),
        }

    @perf_tracked("project.export_data")
    def veri_disari_aktar(self):
        self.guncelle_veri_objesi(silent=True)
        path = filedialog.asksaveasfilename(title="Proje Verisini Excel'e Aktar", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            wb.remove(wb.active)
            rows, _ = wb_build_initial_rows(self.veri, WORKBOOK_SHEET_DEFS)
            for sheet_key, spec in WORKBOOK_SHEET_DEFS.items():
                ws = wb.create_sheet(spec["title"])
                ws.append([label for label, _ in spec["columns"]])
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="D9EAF7")
                for row in rows.get(sheet_key, []):
                    ws.append(row)
                for idx, width in enumerate(spec["widths"], start=1):
                    ws.column_dimensions[get_column_letter(idx)].width = max(10, width / 7)
            health = proje_saglik_ozeti(self.veri, self._dosya_map())
            ws = wb.create_sheet("Saglik")
            ws.append(["Durum", health["state"], health["score"]])
            ws.append(["Başlık", "Sonuç", "Detay"])
            for item in health["items"]:
                ws.append([item["label"], "OK" if item["ok"] else "EKSIK", item["detail"]])
            ws = wb.create_sheet("Hesap Ozeti")
            for line in format_hesap_ozeti(hesap_ozeti(self.veri)).splitlines():
                ws.append([line])
            wb.save(path)
            self.set_status(f"Proje verisi aktarıldı: {os.path.basename(path)}", level="success")
        except Exception as exc:
            log_exception("project.export_data", exc_value=exc)
            messagebox.showerror("Dışa Aktar", f"Veri aktarılamadı:\n{exc}")

    def gunluk_penceresi(self):
        win = Toplevel(self.root)
        self.pencere_hazirla(win, "RaporPro Günlükleri", "900x620", (760, 480), modal=True)
        notebook = ttk.Notebook(win)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        def read_tail(path, max_lines=500):
            if not os.path.exists(path):
                return "Henüz kayıt yok."
            try:
                with open(path, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                return "".join(lines[-max_lines:]) or "Henüz kayıt yok."
            except Exception as exc:
                return f"Günlük okunamadı:\n{exc}"

        for title, path in (("Performans", PERF_LOG_PATH), ("Hatalar", ERROR_LOG_PATH)):
            frame = ttk.Frame(notebook)
            notebook.add(frame, text=title)
            txt = tk.Text(frame, wrap="none", font=("Consolas", 9))
            txt.pack(fill="both", expand=True, padx=8, pady=8)
            txt.insert("1.0", read_tail(path))
            txt.config(state="disabled")

        tk.Button(win, text="Kapat", command=win.destroy, bg=COLOR_PRIMARY, fg="white").pack(pady=(0, 10))
