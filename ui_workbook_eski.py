# Dosya: RaporPro/ui_workbook_eski.py
import datetime
import os
import tkinter as tk
from tkinter import Toplevel, filedialog, messagebox, ttk

from performans import log_exception, perf_timer, perf_tracked
from sabitler import *
from workbook_motoru import (
    WORKBOOK_SHEET_DEFS,
    apply_rows_to_veri as wb_apply_rows_to_veri,
    build_initial_rows as wb_build_initial_rows,
    calc_n30 as wb_calc_n30,
    header_map as wb_header_map,
    rows_to_dicts as wb_rows_to_dicts,
    validate_rows as wb_validate_rows,
)
from karot_motoru import derinlik_baslangic
from yardimcilar import litoloji_yazim_uyarilari, safe_float, temizle_baslik
from widgets import UndoRedoEntry


class WorkbookEskiMixin:
    def veri_giris_workbook_ac(self):
        self.sondaj_verilerini_kaydet()
        win = Toplevel(self.root)
        self.pencere_hazirla(win, "Veri Giriş Workbook", "1250x720", (980, 620))

        sheet_defs = {
            "sondajlar": {
                "title": "Sondajlar",
                "columns": [
                    ("SondajNo", "no"), ("Derinlik", "der"), ("Enlem", "y"), ("Boylam", "x"), ("Kot", "k"),
                    ("Baş.Tarih", "bas_tar"), ("Bit.Tarih", "bit_tar"), ("YASS İlk", "yass_d1"),
                    ("YASS T1", "yass_t1"), ("YASS Son", "yass_d2"), ("YASS T2", "yass_t2")
                ],
                "widths": {"no": 11, "der": 9, "y": 14, "x": 14, "k": 9, "bas_tar": 12, "bit_tar": 12, "yass_d1": 9, "yass_t1": 12, "yass_d2": 9, "yass_t2": 12},
            },
            "litoloji": {
                "title": "Litoloji",
                "columns": [("SondajNo", "sondaj_no"), ("Baslangic", "top"), ("Bitis", "bot"), ("Tanim", "tanim")],
                "widths": {"sondaj_no": 11, "top": 10, "bot": 10, "tanim": 42},
            },
            "spt": {
                "title": "SPT",
                "columns": [("SondajNo", "sondaj_no"), ("Derinlik", "der"), ("15", "v15"), ("30", "v30"), ("45", "v45"), ("N30", "n30")],
                "widths": {"sondaj_no": 11, "der": 10, "v15": 8, "v30": 8, "v45": 8, "n30": 8},
            },
            "pmt": {
                "title": "PMT",
                "columns": [("SondajNo", "sondaj_no"), ("Derinlik", "der"), ("Em", "em"), ("Pl", "pl")],
                "widths": {"sondaj_no": 11, "der": 10, "em": 12, "pl": 12},
            },
            "kaya": {
                "title": "Kaya",
                "columns": [("SondajNo", "sondaj_no"), ("Derinlik", "der"), ("TCR", "tcr"), ("SCR", "scr"), ("RQD", "rqd")],
                "widths": {"sondaj_no": 11, "der": 10, "tcr": 9, "scr": 9, "rqd": 9},
            },
            "numune": {
                "title": "Numune",
                "columns": [("SondajNo", "sondaj_no"), ("Derinlik/Aralık", "aralik"), ("Türü/No", "tur")],
                "widths": {"sondaj_no": 11, "aralik": 18, "tur": 22},
            },
        }

        top = ttk.Frame(win, padding=8)
        top.pack(fill="x")
        nb = ttk.Notebook(win)
        nb.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        sheets = {}
        frame_to_key = {}
        undo_stack = []
        redo_stack = []
        validate_after_id = {"id": None}
        style = ttk.Style()
        style.configure("Workbook.TEntry", fieldbackground="white")
        style.configure("WorkbookSelected.TEntry", fieldbackground="#D6EAF8")
        style.configure("WorkbookError.TEntry", fieldbackground="#FADBD8")
        style.configure("WorkbookWarning.TEntry", fieldbackground="#FCF3CF")

        def yeni_sondaj_sablonu(idx):
            bugun = datetime.datetime.now()
            bugun_str = bugun.strftime("%d.%m.%Y")
            t2_str = (bugun + datetime.timedelta(days=10)).strftime("%d.%m.%Y")
            return {
                "no": f"SK-{idx + 1}", "der": "15.0", "y": "", "x": "", "k": "",
                "bas_tar": bugun_str, "bit_tar": bugun_str,
                "yass_d1": "", "yass_t1": bugun_str, "yass_d2": "", "yass_t2": t2_str,
                "litoloji": [], "spt": [], "pmt": [], "kaya": [], "numuneler": []
            }

        def normalize_header(cell):
            import unicodedata
            text = str(cell).strip().lower()
            text = text.replace("\u0131", "i").replace("\u0130", "i")
            text = unicodedata.normalize("NFKD", text)
            text = "".join(ch for ch in text if not unicodedata.combining(ch))
            return temizle_baslik(text)

        def split_clipboard_line(line):
            if "\t" in line:
                return line.split("\t")
            if ";" in line:
                return line.split(";")
            return line.split()

        def get_active_key():
            return frame_to_key.get(nb.select(), "sondajlar")

        def sheet_row_values(sheet, row):
            return {col_key: row["entries"][col_key].get().strip() for _, col_key in sheet["columns"]}

        def row_has_data(values, ignored=None):
            ignored = ignored or set()
            return any(str(value).strip() for key, value in values.items() if key not in ignored)

        def workbook_snapshot():
            snap = {}
            for sheet_key, sheet in sheets.items():
                snap[sheet_key] = {
                    "rows": [sheet_row_values(sheet, row) for row in sheet["rows"]],
                    "source_nos": [row.get("source_no", "") for row in sheet["rows"]],
                    "widths": sheet["widths"].copy(),
                }
            return snap

        def push_undo():
            if sheets:
                undo_stack.append(workbook_snapshot())
                redo_stack.clear()
                if len(undo_stack) > 40:
                    undo_stack.pop(0)

        def clear_sheet(sheet_key):
            sheet = sheets[sheet_key]
            for widget in sheet["table"].winfo_children():
                widget.destroy()
            sheet["rows"] = []
            sheet["selected"] = set()
            sheet["anchor"] = None
            tk.Label(sheet["table"], text="#", bg="#D5DBDB", font=FONT_BOLD, width=4).grid(row=0, column=0, padx=1, pady=2, sticky="nsew")
            sheet["header_widgets"] = {}
            for col_idx, (label, col_key) in enumerate(sheet["columns"], start=1):
                header = tk.Label(sheet["table"], text=label, bg="#D5DBDB", font=FONT_BOLD, width=sheet["widths"].get(col_key, 12))
                header.grid(row=0, column=col_idx, padx=1, pady=2, sticky="nsew")
                sheet["header_widgets"][col_key] = header

        def load_snapshot(snapshot):
            for sheet_key, data in snapshot.items():
                if sheet_key not in sheets:
                    continue
                sheets[sheet_key]["widths"].update(data.get("widths", {}))
                clear_sheet(sheet_key)
                source_nos = data.get("source_nos", [])
                for idx, row_values in enumerate(data.get("rows", [])):
                    add_row(sheet_key, row_values, source_no=source_nos[idx] if idx < len(source_nos) else "")
            validate_workbook(show_status=False)

        def undo_workbook(event=None):
            if not undo_stack:
                return "break"
            redo_stack.append(workbook_snapshot())
            load_snapshot(undo_stack.pop())
            self.set_status("Workbook islemi geri alindi.", level="info")
            return "break"

        def redo_workbook(event=None):
            if not redo_stack:
                return "break"
            undo_stack.append(workbook_snapshot())
            load_snapshot(redo_stack.pop())
            self.set_status("Workbook islemi ileri alindi.", level="info")
            return "break"

        def default_row_values(sheet_key):
            sheet = sheets[sheet_key]
            if sheet_key == "sondajlar":
                return yeni_sondaj_sablonu(len(sheet["rows"]))
            vals = {}
            if sheet["rows"]:
                last = sheet_row_values(sheet, sheet["rows"][-1])
                vals["sondaj_no"] = last.get("sondaj_no", "")
                if sheet_key == "litoloji":
                    vals["top"] = last.get("bot", "")
                elif sheet_key == "spt":
                    vals["der"] = f"{safe_float(last.get('der')) + 1.5:.2f}" if last.get("der") else ""
            else:
                sondajlar = self.veri.get("sondaj", [])
                vals["sondaj_no"] = sondajlar[0].get("no", "SK-1") if sondajlar else "SK-1"
            return vals

        def focus_cell(sheet_key, row_idx, col_idx):
            sheet = sheets[sheet_key]
            if row_idx >= len(sheet["rows"]):
                add_row(sheet_key, default_row_values(sheet_key))
            row_idx = max(0, min(row_idx, len(sheet["rows"]) - 1))
            col_idx = max(0, min(col_idx, len(sheet["columns"]) - 1))
            key = sheet["columns"][col_idx][1]
            entry = sheet["rows"][row_idx]["entries"][key]
            entry.focus_set()
            entry.selection_range(0, tk.END)
            return "break"

        def copy_down(sheet_key, row_idx, col_idx):
            push_undo()
            sheet = sheets[sheet_key]
            if row_idx + 1 >= len(sheet["rows"]):
                add_row(sheet_key, default_row_values(sheet_key))
            key = sheet["columns"][col_idx][1]
            value = sheet["rows"][row_idx]["entries"][key].get()
            target = sheet["rows"][row_idx + 1]["entries"][key]
            target.delete(0, tk.END)
            target.insert(0, value)
            validate_workbook(show_status=False)
            return focus_cell(sheet_key, row_idx + 1, col_idx)

        def entry_style(sheet, row_idx, col_key):
            cell = (row_idx, col_key)
            if cell in sheet.get("selected", set()):
                return "WorkbookSelected.TEntry"
            if cell in sheet.get("invalid", set()):
                return "WorkbookError.TEntry"
            if cell in sheet.get("warning", set()):
                return "WorkbookWarning.TEntry"
            return "Workbook.TEntry"

        def refresh_sheet_styles(sheet_key):
            sheet = sheets[sheet_key]
            for row_idx, row in enumerate(sheet["rows"]):
                for _, col_key in sheet["columns"]:
                    row["entries"][col_key].configure(style=entry_style(sheet, row_idx, col_key))

        def refresh_all_styles():
            for sheet_key in sheets:
                refresh_sheet_styles(sheet_key)

        def schedule_validate(event=None):
            if validate_after_id["id"]:
                try:
                    win.after_cancel(validate_after_id["id"])
                except Exception:
                    pass
            validate_after_id["id"] = win.after(350, lambda: validate_workbook(show_status=False))

        def select_range(sheet_key, start, end):
            sheet = sheets[sheet_key]
            r1, c1 = start
            r2, c2 = end
            r_min, r_max = sorted((r1, r2))
            c_min, c_max = sorted((c1, c2))
            selected = set()
            for row_idx in range(r_min, r_max + 1):
                if row_idx >= len(sheet["rows"]):
                    continue
                for col_idx in range(c_min, c_max + 1):
                    if col_idx >= len(sheet["columns"]):
                        continue
                    selected.add((row_idx, sheet["columns"][col_idx][1]))
            sheet["selected"] = selected
            refresh_sheet_styles(sheet_key)

        def on_cell_click(event, sheet_key, row_idx, col_idx):
            sheet = sheets[sheet_key]
            sheet["focused"] = {"row": row_idx, "col": col_idx}
            if event.state & 0x0001 and sheet.get("anchor"):
                select_range(sheet_key, sheet["anchor"], (row_idx, col_idx))
            else:
                sheet["anchor"] = (row_idx, col_idx)
                select_range(sheet_key, (row_idx, col_idx), (row_idx, col_idx))
            return None

        def on_cell_focus(sheet_key, row_idx, col_idx):
            sheet = sheets[sheet_key]
            sheet["focused"] = {"row": row_idx, "col": col_idx}
            col_key = sheet["columns"][col_idx][1]
            if row_idx < len(sheet["rows"]):
                sheet["edit_snapshot"] = workbook_snapshot()
                sheet["edit_cell"] = (row_idx, col_key, sheet["rows"][row_idx]["entries"][col_key].get())

        def on_cell_focus_out(sheet_key, row_idx, col_idx):
            sheet = sheets[sheet_key]
            edit_cell = sheet.get("edit_cell")
            if not edit_cell or row_idx >= len(sheet["rows"]):
                return
            col_key = sheet["columns"][col_idx][1]
            old_row, old_col, old_value = edit_cell
            if old_row == row_idx and old_col == col_key:
                new_value = sheet["rows"][row_idx]["entries"][col_key].get()
                if new_value != old_value and sheet.get("edit_snapshot"):
                    undo_stack.append(sheet["edit_snapshot"])
                    redo_stack.clear()
                    if len(undo_stack) > 40:
                        undo_stack.pop(0)
            sheet["edit_cell"] = None
            sheet["edit_snapshot"] = None

        def selected_bounds(sheet):
            selected = sheet.get("selected", set())
            if not selected:
                row_idx = sheet["focused"]["row"]
                col_idx = sheet["focused"]["col"]
                col_key = sheet["columns"][col_idx][1]
                selected = {(row_idx, col_key)}
            col_index = {col_key: idx for idx, (_, col_key) in enumerate(sheet["columns"])}
            rows = [row_idx for row_idx, _ in selected]
            cols = [col_index[col_key] for _, col_key in selected]
            return min(rows), max(rows), min(cols), max(cols)

        def copy_selection(event=None):
            sheet = sheets[get_active_key()]
            if not sheet["rows"]:
                return "break"
            r_min, r_max, c_min, c_max = selected_bounds(sheet)
            lines = []
            for row_idx in range(r_min, r_max + 1):
                cells = []
                for col_idx in range(c_min, c_max + 1):
                    col_key = sheet["columns"][col_idx][1]
                    if row_idx < len(sheet["rows"]):
                        cells.append(sheet["rows"][row_idx]["entries"][col_key].get())
                    else:
                        cells.append("")
                lines.append("\t".join(cells))
            win.clipboard_clear()
            win.clipboard_append("\n".join(lines))
            self.set_status(f"{sheet['title']} secimi panoya kopyalandi.", level="info")
            return "break"

        def delete_selection(event=None):
            sheet_key = get_active_key()
            sheet = sheets[sheet_key]
            if not sheet["rows"]:
                return "break"
            push_undo()
            for row_idx, col_key in list(sheet.get("selected", set())):
                if row_idx < len(sheet["rows"]):
                    sheet["rows"][row_idx]["entries"][col_key].delete(0, tk.END)
            validate_workbook(show_status=False)
            return "break"

        def fill_down_selection(event=None):
            sheet_key = get_active_key()
            sheet = sheets[sheet_key]
            if not sheet["rows"]:
                return "break"
            r_min, r_max, c_min, c_max = selected_bounds(sheet)
            if r_max <= r_min:
                return copy_down(sheet_key, r_min, c_min)
            push_undo()
            for col_idx in range(c_min, c_max + 1):
                col_key = sheet["columns"][col_idx][1]
                value = sheet["rows"][r_min]["entries"][col_key].get()
                for row_idx in range(r_min + 1, r_max + 1):
                    if row_idx < len(sheet["rows"]):
                        entry = sheet["rows"][row_idx]["entries"][col_key]
                        entry.delete(0, tk.END)
                        entry.insert(0, value)
            validate_workbook(show_status=False)
            self.set_status("Seçili aralık aşağı dolduruldu.", level="success")
            return "break"

        def add_row(sheet_key, values=None, source_no=""):
            sheet = sheets[sheet_key]
            row_idx = len(sheet["rows"])
            values = values or {}
            ttk.Label(sheet["table"], text=str(row_idx + 1), width=4, anchor="center").grid(row=row_idx + 1, column=0, padx=1, pady=2, sticky="nsew")
            entries = {}
            for col_idx, (label, col_key) in enumerate(sheet["columns"], start=1):
                entry = UndoRedoEntry(sheet["table"], width=sheet["widths"].get(col_key, 12), style="Workbook.TEntry")
                entry.insert(0, values.get(col_key, ""))
                entry.grid(row=row_idx + 1, column=col_idx, padx=1, pady=2, sticky="nsew")
                entries[col_key] = entry
                entry.bind("<FocusIn>", lambda event, sk=sheet_key, r=row_idx, c=col_idx - 1: on_cell_focus(sk, r, c))
                entry.bind("<FocusOut>", lambda event, sk=sheet_key, r=row_idx, c=col_idx - 1: on_cell_focus_out(sk, r, c))
                entry.bind("<Button-1>", lambda event, sk=sheet_key, r=row_idx, c=col_idx - 1: on_cell_click(event, sk, r, c), add="+")
                entry.bind("<Return>", lambda event, sk=sheet_key, r=row_idx, c=col_idx - 1: focus_cell(sk, r + 1, c))
                entry.bind("<Control-Down>", lambda event, sk=sheet_key, r=row_idx, c=col_idx - 1: copy_down(sk, r, c))
                entry.bind("<Control-c>", copy_selection)
                entry.bind("<Control-C>", copy_selection)
                entry.bind("<Control-v>", lambda event: paste_active_sheet())
                entry.bind("<Control-V>", lambda event: paste_active_sheet())
                entry.bind("<Control-z>", undo_workbook)
                entry.bind("<Control-Z>", undo_workbook)
                entry.bind("<Control-y>", redo_workbook)
                entry.bind("<Control-Y>", redo_workbook)
                entry.bind("<Control-d>", fill_down_selection)
                entry.bind("<Control-D>", fill_down_selection)
                entry.bind("<Delete>", delete_selection)
                entry.bind("<KeyRelease>", schedule_validate, add="+")
                if sheet_key == "litoloji" and col_key == "bot":
                    entry.bind("<Return>", lambda event, sk=sheet_key, r=row_idx: litoloji_smart_next(sk, r))
                if sheet_key == "spt" and col_key in ("v30", "v45"):
                    entry.bind("<KeyRelease>", lambda event, sk=sheet_key, r=row_idx: auto_spt_n30(sk, r), add="+")
            sheet["rows"].append({"entries": entries, "source_no": source_no})
            sheet["canvas"].update_idletasks()
            sheet["canvas"].configure(scrollregion=sheet["canvas"].bbox("all"))

        def active_add_row():
            push_undo()
            sheet_key = get_active_key()
            add_row(sheet_key, default_row_values(sheet_key))
            validate_workbook(show_status=False)

        def rebuild_sheet(sheet_key, rows, source_nos=None):
            clear_sheet(sheet_key)
            source_nos = source_nos or []
            for idx, row_values in enumerate(rows):
                add_row(sheet_key, row_values, source_no=source_nos[idx] if idx < len(source_nos) else "")
            if not sheets[sheet_key]["rows"]:
                add_row(sheet_key, default_row_values(sheet_key))
            validate_workbook(show_status=False)

        def active_insert_row():
            sheet_key = get_active_key()
            sheet = sheets[sheet_key]
            push_undo()
            row_idx = max(0, min(sheet["focused"]["row"] + 1, len(sheet["rows"])))
            rows = [sheet_row_values(sheet, row) for row in sheet["rows"]]
            source_nos = [row.get("source_no", "") for row in sheet["rows"]]
            rows.insert(row_idx, default_row_values(sheet_key))
            source_nos.insert(row_idx, "")
            rebuild_sheet(sheet_key, rows, source_nos)
            focus_cell(sheet_key, row_idx, 0)

        def active_delete_row():
            sheet_key = get_active_key()
            sheet = sheets[sheet_key]
            if not sheet["rows"]:
                return
            push_undo()
            row_idx = max(0, min(sheet["focused"]["row"], len(sheet["rows"]) - 1))
            rows = [sheet_row_values(sheet, row) for row in sheet["rows"]]
            source_nos = [row.get("source_no", "") for row in sheet["rows"]]
            del rows[row_idx]
            del source_nos[row_idx]
            rebuild_sheet(sheet_key, rows, source_nos)
            focus_cell(sheet_key, min(row_idx, len(sheets[sheet_key]["rows"]) - 1), 0)

        def active_clear_row():
            sheet = sheets[get_active_key()]
            if not sheet["rows"]:
                return
            push_undo()
            row_idx = max(0, min(sheet["focused"]["row"], len(sheet["rows"]) - 1))
            for entry in sheet["rows"][row_idx]["entries"].values():
                entry.delete(0, tk.END)
            validate_workbook(show_status=False)

        def active_clear_column():
            sheet_key = get_active_key()
            sheet = sheets[sheet_key]
            if not sheet["rows"] or not sheet["columns"]:
                return
            col_idx = max(0, min(sheet["focused"]["col"], len(sheet["columns"]) - 1))
            label, col_key = sheet["columns"][col_idx]
            if not messagebox.askyesno("Sütunu Temizle", f"{label} sütunundaki {len(sheet['rows'])} hücre temizlensin mi?", parent=win):
                return
            push_undo()
            for row in sheet["rows"]:
                row["entries"][col_key].delete(0, tk.END)
            validate_workbook(show_status=False)
            self.set_status(f"{sheet['title']} / {label} sütunu temizlendi.", level="info")

        def active_resize_col(delta):
            sheet_key = get_active_key()
            sheet = sheets[sheet_key]
            if not sheet["columns"]:
                return
            col_idx = max(0, min(sheet["focused"]["col"], len(sheet["columns"]) - 1))
            col_key = sheet["columns"][col_idx][1]
            new_width = max(5, min(80, sheet["widths"].get(col_key, 12) + delta))
            sheet["widths"][col_key] = new_width
            if col_key in sheet.get("header_widgets", {}):
                sheet["header_widgets"][col_key].config(width=new_width)
            for row in sheet["rows"]:
                row["entries"][col_key].configure(width=new_width)

        def header_map(sheet_key, cells):
            sheet = sheets[sheet_key]
            aliases = {
                "sondajno": "sondaj_no", "sondaj": "sondaj_no", "sk": "sondaj_no", "kuyuno": "sondaj_no",
                "no": "no", "sondajadi": "no",
                "derinlik": "der", "der": "der", "derinlikm": "der",
                "enlem": "y", "lat": "y", "latitude": "y", "y": "y",
                "boylam": "x", "lon": "x", "longitude": "x", "x": "x",
                "kot": "k",
                "bastarih": "bas_tar", "bastarihi": "bas_tar", "baslangictarihi": "bas_tar",
                "bittarih": "bit_tar", "bittarihi": "bit_tar", "bitistarihi": "bit_tar",
                "yassilk": "yass_d1", "yassd1": "yass_d1", "yass1": "yass_d1",
                "yasst1": "yass_t1", "yassilktarih": "yass_t1",
                "yassson": "yass_d2", "yassd2": "yass_d2", "yass2": "yass_d2",
                "yasst2": "yass_t2", "yasssontarih": "yass_t2",
                "baslangic": "top", "bas": "top", "ust": "top", "top": "top",
                "bitis": "bot", "bit": "bot", "alt": "bot", "bot": "bot",
                "tanim": "tanim", "litoloji": "tanim", "birim": "tanim",
                "15": "v15", "n15": "v15", "30": "v30", "n30vurus": "v30", "45": "v45", "n45": "v45",
                "n30": "n30", "em": "em", "pl": "pl", "tcr": "tcr", "scr": "scr", "rqd": "rqd",
                "aralik": "aralik", "derinlikaralik": "aralik", "tur": "tur", "turu": "tur", "turuno": "tur", "numune": "tur",
            }
            allowed = {key for _, key in sheet["columns"]}
            mapped = []
            for cell in cells:
                key = aliases.get(normalize_header(cell))
                if sheet_key == "sondajlar" and key == "sondaj_no":
                    key = "no"
                elif sheet_key != "sondajlar" and key == "no":
                    key = "sondaj_no"
                mapped.append(key if key in allowed else None)
            return mapped if sum(1 for item in mapped if item) >= 2 else None

        def paste_active_sheet():
            sheet_key = get_active_key()
            sheet = sheets[sheet_key]
            try:
                raw = win.clipboard_get()
            except Exception as exc:
                messagebox.showerror("Pano", f"Pano okunamadı:\n{exc}")
                return
            rows = [split_clipboard_line(line.strip()) for line in raw.splitlines() if line.strip()]
            if not rows:
                return
            mapping = header_map(sheet_key, rows[0])
            start_row = sheet["focused"]["row"]
            start_col = sheet["focused"]["col"]
            data_rows = rows[1:] if mapping else rows
            push_undo()
            for r_offset, cells in enumerate(data_rows):
                target_row = start_row + r_offset
                while target_row >= len(sheet["rows"]):
                    add_row(sheet_key, default_row_values(sheet_key))
                for c_offset, value in enumerate(cells):
                    if mapping:
                        if c_offset >= len(mapping) or not mapping[c_offset]:
                            continue
                        col_key = mapping[c_offset]
                    else:
                        target_col = start_col + c_offset
                        if target_col >= len(sheet["columns"]):
                            continue
                        col_key = sheet["columns"][target_col][1]
                    entry = sheet["rows"][target_row]["entries"][col_key]
                    entry.delete(0, tk.END)
                    entry.insert(0, value.strip())
            validate_workbook(show_status=False)
            self.set_status(f"{sheet['title']} sayfasına {len(data_rows)} satır aktarıldı.", level="success")
            return "break"

        def collect_rows(sheet_key):
            sheet = sheets[sheet_key]
            return [sheet_row_values(sheet, row) for row in sheet["rows"]]

        def calc_n30(v30, v45, existing):
            existing_text = str(existing).strip()
            joined = " ".join([str(v30), str(v45), existing_text]).lower()
            if "50/" in joined or existing_text.lower() == "r" or "-" in joined:
                return existing_text or "R"
            if existing_text:
                return existing_text
            total = safe_float(v30) + safe_float(v45)
            return str(int(total)) if total and float(total).is_integer() else (str(total) if total else "")

        def auto_spt_n30(sheet_key, row_idx):
            sheet = sheets[sheet_key]
            if row_idx >= len(sheet["rows"]):
                return
            entries = sheet["rows"][row_idx]["entries"]
            n30 = entries["n30"].get().strip()
            if n30:
                return
            calculated = calc_n30(entries["v30"].get(), entries["v45"].get(), "")
            if calculated:
                entries["n30"].delete(0, tk.END)
                entries["n30"].insert(0, calculated)

        def litoloji_smart_next(sheet_key, row_idx):
            push_undo()
            sheet = sheets[sheet_key]
            if row_idx >= len(sheet["rows"]):
                return "break"
            values = sheet_row_values(sheet, sheet["rows"][row_idx])
            next_idx = row_idx + 1
            if next_idx >= len(sheet["rows"]):
                add_row(sheet_key, {
                    "sondaj_no": values.get("sondaj_no", ""),
                    "top": values.get("bot", ""),
                    "bot": "",
                    "tanim": values.get("tanim", ""),
                })
            else:
                next_entries = sheet["rows"][next_idx]["entries"]
                if not next_entries["sondaj_no"].get().strip():
                    next_entries["sondaj_no"].insert(0, values.get("sondaj_no", ""))
                if not next_entries["top"].get().strip():
                    next_entries["top"].insert(0, values.get("bot", ""))
            validate_workbook(show_status=False)
            return focus_cell(sheet_key, next_idx, 2)

        @perf_tracked("workbook.legacy_validate")
        def validate_workbook(show_status=True):
            for sheet in sheets.values():
                sheet["invalid"] = set()
                sheet["warning"] = set()

            sondaj_rows = collect_rows("sondajlar") if "sondajlar" in sheets else []
            valid_nos = set()
            seen = {}
            depth_by_no = {}

            for row_idx, values in enumerate(sondaj_rows):
                if not row_has_data(values):
                    continue
                no = values.get("no", "").strip()
                if not no:
                    sheets["sondajlar"]["invalid"].add((row_idx, "no"))
                    continue
                if no in seen:
                    sheets["sondajlar"]["invalid"].add((row_idx, "no"))
                    sheets["sondajlar"]["invalid"].add((seen[no], "no"))
                seen[no] = row_idx
                valid_nos.add(no)
                der = safe_float(values.get("der", ""))
                if der <= 0:
                    sheets["sondajlar"]["warning"].add((row_idx, "der"))
                depth_by_no[no] = der
                for coord_key in ("y", "x"):
                    val = values.get(coord_key, "").strip()
                    if val and safe_float(val) == 0:
                        sheets["sondajlar"]["warning"].add((row_idx, coord_key))

            def mark_unknown_no(sheet_key):
                for row_idx, values in enumerate(collect_rows(sheet_key)):
                    if not row_has_data(values, {"sondaj_no"}):
                        continue
                    no = values.get("sondaj_no", "").strip()
                    if not no or no not in valid_nos:
                        sheets[sheet_key]["invalid"].add((row_idx, "sondaj_no"))

            for sheet_key in ("litoloji", "spt", "pmt", "kaya", "numune"):
                mark_unknown_no(sheet_key)

            lit_by_no = {}
            for row_idx, values in enumerate(collect_rows("litoloji")):
                if not row_has_data(values, {"sondaj_no"}):
                    continue
                no = values.get("sondaj_no", "").strip()
                top = safe_float(values.get("top", ""))
                bot = safe_float(values.get("bot", ""))
                if bot <= top:
                    sheets["litoloji"]["invalid"].add((row_idx, "top"))
                    sheets["litoloji"]["invalid"].add((row_idx, "bot"))
                if no in depth_by_no and depth_by_no[no] > 0 and bot > depth_by_no[no]:
                    sheets["litoloji"]["warning"].add((row_idx, "bot"))
                if litoloji_yazim_uyarilari(values.get("tanim", "")):
                    sheets["litoloji"]["warning"].add((row_idx, "tanim"))
                lit_by_no.setdefault(no, []).append((row_idx, top, bot))

            for no, rows in lit_by_no.items():
                rows = sorted(rows, key=lambda item: item[1])
                prev_bot = None
                for row_idx, top, bot in rows:
                    if prev_bot is not None:
                        if top < prev_bot - 0.01:
                            sheets["litoloji"]["invalid"].add((row_idx, "top"))
                        elif top > prev_bot + 0.01:
                            sheets["litoloji"]["warning"].add((row_idx, "top"))
                    prev_bot = bot

            for sheet_key in ("spt", "pmt", "kaya"):
                for row_idx, values in enumerate(collect_rows(sheet_key)):
                    if not row_has_data(values, {"sondaj_no"}):
                        continue
                    no = values.get("sondaj_no", "").strip()
                    der = derinlik_baslangic(values.get("der", "")) if sheet_key == "kaya" else safe_float(values.get("der", ""))
                    if der <= 0:
                        sheets[sheet_key]["warning"].add((row_idx, "der"))
                    if no in depth_by_no and depth_by_no[no] > 0 and der > depth_by_no[no]:
                        sheets[sheet_key]["invalid"].add((row_idx, "der"))

            refresh_all_styles()
            error_count = sum(len(sheet.get("invalid", set())) for sheet in sheets.values())
            warning_count = sum(len(sheet.get("warning", set())) for sheet in sheets.values())
            if show_status:
                if error_count:
                    self.set_status(f"Workbook kontrolü: {error_count} hata, {warning_count} uyarı.", level="error")
                elif warning_count:
                    self.set_status(f"Workbook kontrolü: {warning_count} uyarı.", level="warning")
                else:
                    self.set_status("Workbook kontrolü temiz.", level="success")
            return error_count, warning_count

        @perf_tracked("workbook.legacy_apply")
        def apply_workbook(close=False):
            error_count, warning_count = validate_workbook(show_status=True)
            if error_count and not messagebox.askyesno("Workbook Kontrol", f"{error_count} hata görünüyor. Yine de uygulansın mı?"):
                return
            if warning_count and not error_count:
                if not messagebox.askyesno("Workbook Kontrol", f"{warning_count} uyarı var. Yine de uygulansın mı?"):
                    return
            old_by_no = {s.get("no", ""): s for s in self.veri.get("sondaj", []) if s.get("no")}
            new_sondajlar = []
            no_alias = {}
            warnings = []

            for idx, values in enumerate(collect_rows("sondajlar")):
                if not row_has_data(values):
                    continue
                old_no = sheets["sondajlar"]["rows"][idx].get("source_no", "")
                no = values.get("no") or f"SK-{len(new_sondajlar) + 1}"
                source = old_by_no.get(old_no) or old_by_no.get(no) or yeni_sondaj_sablonu(len(new_sondajlar))
                sondaj = source.copy()
                sondaj.update(values)
                sondaj["no"] = no
                if old_no and old_no != no:
                    no_alias[old_no] = no
                for key in ("litoloji", "spt", "pmt", "kaya", "numuneler"):
                    sondaj[key] = []
                new_sondajlar.append(sondaj)

            if not new_sondajlar:
                messagebox.showwarning("Workbook", "Sondajlar sayfasında uygulanacak veri yok.")
                return

            by_no = {s.get("no", ""): s for s in new_sondajlar}

            def target_sondaj(values):
                raw_no = values.get("sondaj_no", "")
                no = no_alias.get(raw_no, raw_no)
                return by_no.get(no) if no else None

            for values in collect_rows("litoloji"):
                if not row_has_data(values, {"sondaj_no"}):
                    continue
                sondaj = target_sondaj(values)
                if not sondaj:
                    warnings.append(f"Litoloji satırı atlandı: {values.get('sondaj_no', '')}")
                    continue
                sondaj["litoloji"].append([values.get("top", ""), values.get("bot", ""), values.get("tanim", "")])

            for values in collect_rows("spt"):
                if not row_has_data(values, {"sondaj_no"}):
                    continue
                sondaj = target_sondaj(values)
                if not sondaj:
                    warnings.append(f"SPT satırı atlandı: {values.get('sondaj_no', '')}")
                    continue
                sondaj["spt"].append([values.get("der", ""), values.get("v15", ""), values.get("v30", ""), values.get("v45", ""), calc_n30(values.get("v30", ""), values.get("v45", ""), values.get("n30", ""))])

            for values in collect_rows("pmt"):
                if not row_has_data(values, {"sondaj_no"}):
                    continue
                sondaj = target_sondaj(values)
                if not sondaj:
                    warnings.append(f"PMT satırı atlandı: {values.get('sondaj_no', '')}")
                    continue
                sondaj["pmt"].append([values.get("der", ""), values.get("em", ""), values.get("pl", "")])

            for values in collect_rows("kaya"):
                if not row_has_data(values, {"sondaj_no"}):
                    continue
                sondaj = target_sondaj(values)
                if not sondaj:
                    warnings.append(f"Kaya satırı atlandı: {values.get('sondaj_no', '')}")
                    continue
                sondaj["kaya"].append([values.get("der", ""), values.get("tcr", ""), values.get("scr", ""), values.get("rqd", "")])

            for values in collect_rows("numune"):
                if not row_has_data(values, {"sondaj_no"}):
                    continue
                sondaj = target_sondaj(values)
                if not sondaj:
                    warnings.append(f"Numune satırı atlandı: {values.get('sondaj_no', '')}")
                    continue
                sondaj["numuneler"].append([values.get("aralik", ""), values.get("tur", "")])

            self.veri["sondaj"] = new_sondajlar
            self.sondaj_tablosunu_ciz()
            self.ozet_yenile(collect=False)
            self.set_status(f"Workbook uygulandı: {len(new_sondajlar)} sondaj güncellendi.", level="success")
            for warning in warnings[:5]:
                self.set_status(warning, level="warning")
            if len(warnings) > 5:
                self.set_status(f"{len(warnings) - 5} ek workbook uyarısı daha var.", level="warning")
            if close:
                win.destroy()

        def create_sheet(sheet_key):
            spec = sheet_defs[sheet_key]
            frame = ttk.Frame(nb)
            nb.add(frame, text=spec["title"])
            frame_to_key[str(frame)] = sheet_key
            canvas = tk.Canvas(frame, bg=COLOR_BG, highlightthickness=0)
            scroll_y = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
            scroll_x = ttk.Scrollbar(frame, orient="horizontal", command=canvas.xview)
            table = ttk.Frame(canvas)
            table.bind("<Configure>", lambda e, c=canvas: c.configure(scrollregion=c.bbox("all")))
            canvas.create_window((0, 0), window=table, anchor="nw")
            canvas.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
            scroll_y.pack(side="right", fill="y")
            scroll_x.pack(side="bottom", fill="x")
            canvas.pack(side="left", fill="both", expand=True)
            sheets[sheet_key] = {
                "title": spec["title"], "columns": spec["columns"], "widths": spec["widths"],
                "canvas": canvas, "table": table, "rows": [], "focused": {"row": 0, "col": 0},
                "selected": set(), "invalid": set(), "warning": set(), "anchor": None, "header_widgets": {}
            }
            tk.Label(table, text="#", bg="#D5DBDB", font=FONT_BOLD, width=4).grid(row=0, column=0, padx=1, pady=2, sticky="nsew")
            for col_idx, (label, col_key) in enumerate(spec["columns"], start=1):
                header = tk.Label(table, text=label, bg="#D5DBDB", font=FONT_BOLD, width=spec["widths"].get(col_key, 12))
                header.grid(row=0, column=col_idx, padx=1, pady=2, sticky="nsew")
                sheets[sheet_key]["header_widgets"][col_key] = header

        for key in sheet_defs:
            create_sheet(key)

        for sondaj in self.veri.get("sondaj", []):
            add_row("sondajlar", {key: sondaj.get(key, "") for _, key in sheets["sondajlar"]["columns"]}, source_no=sondaj.get("no", ""))
            no = sondaj.get("no", "")
            for row in sondaj.get("litoloji", []):
                add_row("litoloji", {"sondaj_no": no, "top": row[0] if len(row) > 0 else "", "bot": row[1] if len(row) > 1 else "", "tanim": row[2] if len(row) > 2 else ""})
            for row in sondaj.get("spt", []):
                add_row("spt", {"sondaj_no": no, "der": row[0] if len(row) > 0 else "", "v15": row[1] if len(row) > 1 else "", "v30": row[2] if len(row) > 2 else "", "v45": row[3] if len(row) > 3 else "", "n30": row[4] if len(row) > 4 else ""})
            for row in sondaj.get("pmt", []):
                add_row("pmt", {"sondaj_no": no, "der": row[0] if len(row) > 0 else "", "em": row[1] if len(row) > 1 else "", "pl": row[2] if len(row) > 2 else ""})
            for row in sondaj.get("kaya", []):
                add_row("kaya", {"sondaj_no": no, "der": row[0] if len(row) > 0 else "", "tcr": row[1] if len(row) > 1 else "", "scr": row[2] if len(row) > 2 else "", "rqd": row[3] if len(row) > 3 else ""})
            for row in sondaj.get("numuneler", []):
                add_row("numune", {"sondaj_no": no, "aralik": row[0] if len(row) > 0 else "", "tur": row[1] if len(row) > 1 else ""})

        for key in sheets:
            if not sheets[key]["rows"]:
                add_row(key, default_row_values(key))

        validate_workbook(show_status=False)

        @perf_tracked("workbook.legacy_export_excel")
        def export_workbook():
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill
                from openpyxl.utils import get_column_letter
            except Exception as exc:
                messagebox.showerror("Excel", f"openpyxl yüklenemedi:\n{exc}")
                return
            path = filedialog.asksaveasfilename(
                title="Workbook Excel'e Aktar",
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")]
            )
            if not path:
                return
            wb = Workbook()
            default_ws = wb.active
            wb.remove(default_ws)
            for sheet_key, spec in sheet_defs.items():
                sheet = sheets[sheet_key]
                ws = wb.create_sheet(spec["title"])
                headers = [label for label, _ in sheet["columns"]]
                ws.append(headers)
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="D9EAF7")
                for row in sheet["rows"]:
                    values = sheet_row_values(sheet, row)
                    if row_has_data(values):
                        ws.append([values.get(col_key, "") for _, col_key in sheet["columns"]])
                for col_idx, (_, col_key) in enumerate(sheet["columns"], start=1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = max(10, sheet["widths"].get(col_key, 12) + 2)
            try:
                wb.save(path)
                self.set_status(f"Workbook Excel'e aktarıldı: {os.path.basename(path)}", level="success")
            except Exception as exc:
                messagebox.showerror("Excel", f"Excel kaydedilemedi:\n{exc}")

        @perf_tracked("workbook.legacy_import_excel")
        def import_workbook():
            try:
                from openpyxl import load_workbook
            except Exception as exc:
                messagebox.showerror("Excel", f"openpyxl yüklenemedi:\n{exc}")
                return
            path = filedialog.askopenfilename(
                title="Workbook Excel'den Al",
                filetypes=[("Excel", "*.xlsx")]
            )
            if not path:
                return
            try:
                wb = load_workbook(path, data_only=True)
            except Exception as exc:
                messagebox.showerror("Excel", f"Excel okunamadı:\n{exc}")
                return
            push_undo()
            imported = 0
            for sheet_key, spec in sheet_defs.items():
                ws = wb[spec["title"]] if spec["title"] in wb.sheetnames else None
                if ws is None:
                    continue
                raw_rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = ["" if cell is None else str(cell) for cell in row]
                    if any(cell.strip() for cell in cells):
                        raw_rows.append(cells)
                if not raw_rows:
                    continue
                mapping = header_map(sheet_key, raw_rows[0])
                data_rows = raw_rows[1:] if mapping else raw_rows
                new_rows = []
                allowed_columns = sheets[sheet_key]["columns"]
                for raw in data_rows:
                    row_values = {}
                    if mapping:
                        for idx, value in enumerate(raw):
                            if idx < len(mapping) and mapping[idx]:
                                row_values[mapping[idx]] = value.strip()
                    else:
                        for idx, value in enumerate(raw):
                            if idx < len(allowed_columns):
                                row_values[allowed_columns[idx][1]] = value.strip()
                    if row_has_data(row_values, {"sondaj_no"}):
                        new_rows.append(row_values)
                if new_rows:
                    source_nos = [row.get("no", "") for row in new_rows] if sheet_key == "sondajlar" else []
                    rebuild_sheet(sheet_key, new_rows, source_nos)
                    imported += len(new_rows)
            validate_workbook(show_status=True)
            self.set_status(f"Excel'den workbook'a {imported} satır aktarıldı.", level="success")

        tk.Button(top, text="Excel Al", command=import_workbook, bg="#2E86C1", fg="white", font=FONT_BOLD).pack(side="left", padx=3)
        tk.Button(top, text="Excel Aktar", command=export_workbook, bg="#1E8449", fg="white", font=FONT_BOLD).pack(side="left", padx=3)
        tk.Button(top, text="Panodan Yapistir", command=paste_active_sheet, bg="#8E44AD", fg="white", font=FONT_BOLD).pack(side="left", padx=3)
        tk.Button(top, text="Kopyala", command=copy_selection, bg="#5D6D7E", fg="white", font=FONT_BOLD).pack(side="left", padx=3)
        tk.Button(top, text="+ Satır", command=active_add_row, bg=COLOR_ACCENT, fg="white", font=FONT_BOLD).pack(side="left", padx=3)
        tk.Button(top, text="Satır Ekle", command=active_insert_row, bg="#5499C7", fg="white", font=FONT_BOLD).pack(side="left", padx=3)
        tk.Button(top, text="Satır Sil", command=active_delete_row, bg=COLOR_DANGER, fg="white", font=FONT_BOLD).pack(side="left", padx=3)
        tk.Button(top, text="Temizle", command=active_clear_row, bg="#7F8C8D", fg="white", font=FONT_BOLD).pack(side="left", padx=3)
        tk.Button(top, text="Sütun Temizle", command=active_clear_column, bg="#7F8C8D", fg="white", font=FONT_BOLD).pack(side="left", padx=3)
        tk.Button(top, text="Doldur", command=fill_down_selection, bg="#AF7AC5", fg="white", font=FONT_BOLD).pack(side="left", padx=3)
        tk.Button(top, text="Kol -", command=lambda: active_resize_col(-2), bg="#D5DBDB", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(top, text="Kol +", command=lambda: active_resize_col(2), bg="#D5DBDB", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(top, text="Geri", command=undo_workbook, bg="#ECF0F1", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(top, text="İleri", command=redo_workbook, bg="#ECF0F1", font=FONT_BOLD).pack(side="left", padx=2)
        tk.Button(top, text="Kontrol", command=lambda: validate_workbook(show_status=True), bg=COLOR_WARNING, fg="white", font=FONT_BOLD).pack(side="left", padx=3)
        tk.Button(top, text="Uygula", command=lambda: apply_workbook(False), bg=COLOR_SUCCESS, fg="white", font=FONT_BOLD).pack(side="right", padx=4)
        tk.Button(top, text="Uygula ve Kapat", command=lambda: apply_workbook(True), bg=COLOR_PRIMARY, fg="white", font=FONT_BOLD).pack(side="right", padx=4)


