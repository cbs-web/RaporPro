# Dosya: RaporPro/arayuz.py
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, Canvas, Scrollbar, Listbox, Toplevel, Frame
import datetime
import json
import os
import threading

from harita_cikti import eski_paylasimli_temp_harita_yolu_mu
from sabitler import *
from yardimcilar import *
from performans import ERROR_LOG_PATH, PERF_LOG_PATH, log_exception, perf_timer, perf_tracked
from proje_motoru import (
    format_hesap_ozeti,
    hesap_ozeti,
    proje_saglik_ozeti,
    rapor_onizleme_metni,
)
from proje_arsiv import (
    arsiv_kaydi_ekle,
    arsiv_kaydi_sil,
    arsiv_kayitlari_yukle,
    biten_isler_kml_yaz,
    proje_merkez_koordinati,
)
from raporlama import raporla
from workbook_motoru import (
    WORKBOOK_SHEET_DEFS,
    build_initial_rows as wb_build_initial_rows,
    yeni_sondaj_sablonu as wb_yeni_sondaj_sablonu,
)
from widgets import UndoRedoEntry

from harita_motoru import TopluHarita
from kalite_kontrol import (
    analyze_word_template,
    backup_project_file,
    build_preflight_report,
    format_preflight_report,
    format_template_analysis,
    get_supported_tags,
)
from ui_cikti import CiktiMerkeziMixin
from ui_haritalar import HaritalarSekmesiMixin
from ui_jeofizik import JeofizikMixin
from ui_karot_tcr import KarotTCRMixin
from ui_kesit import KesitCizimMixin
from ui_kontrol import KontrolPaneliMixin
from ui_rapor import RaporSekmesiMixin
from ui_spt_okuma import SPTOkumaMixin
from ui_sondaj import SondajMixin
from ui_workbook import WorkbookMixin
from arayuz_temel import ArayuzTemelMixin

APP_DIR = os.path.dirname(os.path.abspath(__file__))
AUTOSAVE_DIR = os.path.join(APP_DIR, "autosave")
AUTOSAVE_PATH = os.path.join(AUTOSAVE_DIR, "raporpro_autosave.json")
RECENT_PROJECTS_PATH = os.path.join(APP_DIR, "recent_projects.json")
# ============================================================================
# ÖZEL SPT VERİ GİRİŞ PENCERESİ (OTOMATİK HESAPLAMA VE DERİNLİK ARTIŞI)
class RaporRobotuArayuz(ArayuzTemelMixin, RaporSekmesiMixin, HaritalarSekmesiMixin, CiktiMerkeziMixin, KontrolPaneliMixin, KesitCizimMixin, WorkbookMixin, SPTOkumaMixin, KarotTCRMixin, SondajMixin, JeofizikMixin):
    @perf_tracked("ui.__init__")
    def __init__(self, root):
        self.root = root
        self.root.report_callback_exception = self._tk_exception_handler
        self.setup_styles()
        self.root.title("Zemin Rapor Pro v49.0 - Hızlı MT Veri Girişi")
        self.pencere_ekrana_sigdir(self.root, "1450x950", (1024, 680), width_ratio=0.96, height_ratio=0.92)
        self.root.after_idle(self.ana_pencere_tam_ekran_yap)
        self.root.configure(bg=COLOR_BG)
        
        self.aktif_dosya_yolu = None 
        self.kml_path = None
        
        self.veri = self.veri_yukle()
        self.word_path = None
        self.img_yer = None; self.img_tkgm = None; self.img_pga = None; self.img_mjh = None
        
        self.word_img_sondaj = None
        self.word_img_jeofizik = None
        
        self.lab_excel_path = None
        self.jeo_excel_path = None 
        self.last_focused = None
        self.last_preflight_report = None
        self.autosave_after_id = None
        self._closing = False
        self._kilitli_kayda_izin_ver = False
        self.last_save_time = None
        self.autosave_status_var = tk.StringVar(value="Kayıt durumu: bekleniyor")
        self.recent_projects = self.recent_projects_yukle()
        
        self.root.bind_all("<Button-1>", self.track_focus, add="+")
        self.sondaj_ui_rows = [] 
        self.kur_arayuz()
        self.kur_kisayollar()
        self.doldur_arayuz()
        
        if self.aktif_dosya_yolu:
            self.root.title(f"Zemin Rapor Pro - {os.path.basename(self.aktif_dosya_yolu)}")
            self.set_status(f"Yüklendi: {self.aktif_dosya_yolu}")
        else:
            self.set_status("Yeni Proje (Kaydedilmemiş)")
        if getattr(self, "bootstrap_theme_active", False):
            self.set_status(f"Modern tema etkin: {self.bootstrap_theme_name}", level="success")
        if getattr(self, "_startup_load_error", None):
            self.set_status(f"Varsayılan proje okunamadı: {self._startup_load_error}", level="warning")
        self.kurtarma_durumu_bildir()
        self.root.protocol("WM_DELETE_WINDOW", self.uygulamayi_kapat)
        self.start_autosave()

    def _tk_exception_handler(self, exc_type, exc_value, exc_tb):
        log_exception("tk.callback", exc_type, exc_value, exc_tb)
        try:
            self.set_status(f"Hata gunlugune yazildi: {exc_value}", level="error")
        except Exception:
            pass

    def track_focus(self, event):
        w = event.widget
        if isinstance(w, (ttk.Entry, tk.Entry, UndoRedoEntry)): self.last_focused = w

    def start_autosave(self):
        if getattr(self, "_closing", False):
            return
        if self.autosave_after_id:
            try:
                self.root.after_cancel(self.autosave_after_id)
            except Exception:
                pass
            self.autosave_after_id = None
        try:
            if not self.root.winfo_exists():
                return
        except Exception:
            return
        self.autosave_after_id = self.root.after(90000, self.autosave_tick)

    def autosave_tick(self):
        self.autosave_after_id = None
        if getattr(self, "_closing", False):
            return
        try:
            if not self.root.winfo_exists():
                return
        except Exception:
            return
        self.otomatik_kaydet()
        self.start_autosave()

    def autosave_zamanlayici_iptal(self):
        after_id = getattr(self, "autosave_after_id", None)
        self.autosave_after_id = None
        if after_id:
            try:
                self.root.after_cancel(after_id)
            except Exception:
                pass

    def uygulamayi_kapat(self):
        self._closing = True
        self.autosave_zamanlayici_iptal()
        try:
            self.root.destroy()
        except tk.TclError:
            pass

    def otomatik_kaydet(self):
        if getattr(self, "_closing", False):
            return
        try:
            if self.proje_kilitli_mi():
                self.set_save_indicator("Proje kilitli: otomatik kayıt yok", "warning")
                return
            if hasattr(self, "e_kunye"):
                self.guncelle_veri_objesi(silent=True)
            os.makedirs(AUTOSAVE_DIR, exist_ok=True)
            payload = {
                "saved_at": datetime.datetime.now().isoformat(timespec="seconds"),
                "active_path": self.aktif_dosya_yolu,
                "veri": self.veri,
            }
            atomic_json_dump(payload, AUTOSAVE_PATH, indent=2, ensure_ascii=False)
            self.set_save_indicator(f"Otomatik kayıt: {datetime.datetime.now().strftime('%H:%M')}", "info")
        except Exception as exc:
            log_exception("autosave.write", exc_value=exc)
            self.set_save_indicator("Otomatik kayıt hatası", "error")

    def kurtarma_durumu_bildir(self):
        try:
            if not os.path.exists(AUTOSAVE_PATH):
                return
            autosave_time = os.path.getmtime(AUTOSAVE_PATH)
            project_time = os.path.getmtime(self.aktif_dosya_yolu) if self.aktif_dosya_yolu and os.path.exists(self.aktif_dosya_yolu) else 0
            if autosave_time > project_time:
                self.set_status("Kurtarma dosyasi bulundu. Ust menuden 'Kurtar' ile yukleyebilirsiniz.", level="warning")
        except Exception as exc:
            log_exception("autosave.check", exc_value=exc)

    @perf_tracked("project.restore_autosave")
    def otomatik_kayit_yukle(self):
        try:
            with open(AUTOSAVE_PATH, "r", encoding="utf-8") as f:
                payload = json.load(f)
            veri = payload.get("veri", {})
            self.veri_eksikleri_tamamla(veri, self.varsayilan_veri_olustur())
            self.veri = veri
            self.aktif_dosya_yolu = payload.get("active_path")
            self.doldur_arayuz()
            self.set_status(f"Otomatik kayıt yüklendi: {payload.get('saved_at', '-')}", level="success")
        except Exception as exc:
            log_exception("autosave.restore", exc_value=exc)
            messagebox.showerror("Kurtarma", f"Otomatik kayıt yüklenemedi:\n{exc}")

    def recent_projects_yukle(self):
        try:
            if not os.path.exists(RECENT_PROJECTS_PATH):
                return []
            with open(RECENT_PROJECTS_PATH, "r", encoding="utf-8") as f:
                payload = json.load(f)
            items = payload.get("projects", payload) if isinstance(payload, dict) else payload
            if not isinstance(items, list):
                return []
            recent = []
            seen = set()
            for item in items:
                path = item.get("path") if isinstance(item, dict) else item
                if not path:
                    continue
                path = os.path.abspath(str(path))
                key = os.path.normcase(path)
                if key in seen or not os.path.exists(path):
                    continue
                seen.add(key)
                recent.append(path)
            return recent[:8]
        except Exception as exc:
            log_exception("recent_projects.load", exc_value=exc)
            return []

    def recent_projects_kaydet(self):
        try:
            atomic_json_dump({"projects": self.recent_projects[:8]}, RECENT_PROJECTS_PATH, indent=2, ensure_ascii=False)
        except Exception as exc:
            log_exception("recent_projects.save", exc_value=exc)

    def recent_project_ekle(self, path):
        if not path:
            return
        path = os.path.abspath(str(path))
        current = getattr(self, "recent_projects", [])
        key = os.path.normcase(path)
        cleaned = []
        seen = {key}
        for item in current:
            item_path = os.path.abspath(str(item))
            item_key = os.path.normcase(item_path)
            if item_key in seen or not os.path.exists(item_path):
                continue
            seen.add(item_key)
            cleaned.append(item_path)
        self.recent_projects = [path] + cleaned
        self.recent_projects_kaydet()

    def proje_dosyasi_yukle(self, dosya_yolu):
        with perf_timer("project.open_read_apply"):
            with open(dosya_yolu, 'r', encoding='utf-8') as f:
                yuklenen_veri = json.load(f)
            varsayilan = self.varsayilan_veri_olustur()
            self.veri_eksikleri_tamamla(yuklenen_veri, varsayilan)
            self.veri = yuklenen_veri
            self.aktif_dosya_yolu = dosya_yolu
            self.doldur_arayuz()
        self.proje_baslik_guncelle()
        self.recent_project_ekle(dosya_yolu)
        self.set_status(f"Proje açıldı: {dosya_yolu}", level="success")
        self.proje_kilit_durumunu_goster()

    def son_projeler_penceresi(self):
        self.recent_projects = self.recent_projects_yukle()
        win = Toplevel(self.root)
        self.pencere_hazirla(win, "Son Projeler", "720x360", (620, 300), modal=True)

        ttk.Label(win, text="Son açılan projeler", font=FONT_HEADER).pack(anchor="w", padx=14, pady=(12, 6))
        list_frame = ttk.Frame(win)
        list_frame.pack(fill="both", expand=True, padx=14, pady=6)
        lb = Listbox(list_frame, height=9, font=("Consolas", 9))
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=lb.yview)
        lb.configure(yscrollcommand=scrollbar.set)
        lb.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        def refresh():
            lb.delete(0, tk.END)
            if not self.recent_projects:
                lb.insert(tk.END, "Son proje yok.")
                return
            for path in self.recent_projects:
                status = "" if os.path.exists(path) else "  [bulunamadı]"
                lb.insert(tk.END, f"{os.path.basename(path):<32} {path}{status}")

        def selected_path():
            if not self.recent_projects:
                return None
            sel = lb.curselection()
            if not sel:
                return None
            idx = sel[0]
            return self.recent_projects[idx] if idx < len(self.recent_projects) else None

        def open_selected(event=None):
            path = selected_path()
            if not path:
                return "break"
            if not os.path.exists(path):
                messagebox.showwarning("Son Projeler", "Bu proje dosyası artık bulunamıyor. Listeden kaldırıldı.")
                self.recent_projects = [p for p in self.recent_projects if os.path.normcase(os.path.abspath(p)) != os.path.normcase(os.path.abspath(path))]
                self.recent_projects_kaydet()
                refresh()
                return "break"
            try:
                self.proje_dosyasi_yukle(path)
                win.destroy()
            except Exception as exc:
                messagebox.showerror("Hata", f"Dosya açılamadı:\n{exc}")
            return "break"

        def remove_selected():
            path = selected_path()
            if not path:
                return
            self.recent_projects = [p for p in self.recent_projects if os.path.normcase(os.path.abspath(p)) != os.path.normcase(os.path.abspath(path))]
            self.recent_projects_kaydet()
            refresh()

        lb.bind("<Double-Button-1>", open_selected)
        lb.bind("<Return>", open_selected)
        btns = ttk.Frame(win)
        btns.pack(fill="x", padx=14, pady=(4, 12))
        tk.Button(btns, text="Aç", command=open_selected, bg=COLOR_PRIMARY, fg="white", font=FONT_BOLD).pack(side="left", padx=(0, 6))
        tk.Button(btns, text="Listeden Kaldır", command=remove_selected, bg="#ECF0F1").pack(side="left", padx=6)
        tk.Button(btns, text="Kapat", command=win.destroy, bg="#ECF0F1").pack(side="right")
        refresh()
        if self.recent_projects:
            lb.selection_set(0)

    def kisayol_calistir(self, command):
        try:
            command()
        except Exception as exc:
            log_exception("shortcut.run", exc_value=exc)
            self.set_status(f"Kısayol hatası: {exc}", level="error")
        return "break"

    def kur_kisayollar(self):
        bindings = {
            "<Control-s>": self.veri_kaydet,
            "<Control-o>": self.proje_ac,
            "<Control-n>": self.yeni_proje,
            "<Control-Shift-S>": self.proje_farkli_kaydet,
            "<Control-w>": self.veri_giris_workbook_tksheet_ac,
            "<Control-r>": self.rapor_on_kontrol,
            "<Control-k>": self.kesit_secim_penceresi,
            "<Control-h>": self.son_projeler_penceresi,
            "<F5>": self.ozet_yenile,
            "<F9>": self.final_kontrol_penceresi,
        }
        for sequence, command in bindings.items():
            try:
                self.root.bind_all(sequence, lambda event, cmd=command: self.kisayol_calistir(cmd), add="+")
            except Exception as exc:
                log_exception("shortcut.bind", exc_value=exc)

    def varsayilan_veri_olustur(self):
        default = {
            "kunye": {"sahibi":"", "il":"", "ilce":"", "mah":"", "mev":"", "paf":"", "ada":"", "par":""},
            "bina": {"kul":"", "sinif":"", "onem":"", "malz":"", "bod":"", "kat":"", "plan":"", "yukseklik":"", "yukseklik_sinif":"", "temel_alan":"", "ins":"", "der":"", "gqe_min":"", "gqe_max":"", "gqe_ort":"", "comb_min":"", "comb_max":"", "comb_ort":"", "ysinif":"", "tem":"", "coklu_blok": False, "bloklar": []},
            "arazi": {"kot":"", "yon":"", "egim":"", "min":"", "max":"", "ort":"", "imar_alani":"", "imar_durumu":"", "zemin":"", "kategori": "", "pga":"", "alan_y": "", "alan_x": ""},
            "sondaj": [],
            "jeofizik": {"tarih": "", "ss_list": [], "mt_list": []},
            "harita_cizimleri": {"vaziyet": {}, "jeoloji": {}, "yerbuldurur": {}},
            "kesit_ayarlari": {},
            "ek_icerikleri": {"normal": {}, "arazi_deneyli": {}},
            "proje_durumu": {"tamamlandi": False, "kilitli": False, "tamamlanma_tarihi": "", "arsiv_notu": ""},
            "ayarlar": {
                "firma_adi": "UB ZEMIN MUHENDISLIK",
                "log_baslik": "SONDAJ LOGU",
                "sorumlu_muhendis_unvan": "Sorumlu Jeoloji Muhendisi",
                "sorumlu_muhendis": "GOKALP DOGAN",
                "sondor_belge_baslik": "Sondor Belge No",
                "sondor_belge": "Murat Ercelik 3629",
                "makine_metodu": "Rotary / Burgusuz",
                "spt_sahmerdan": "Otomatik",
                "delgi_capi": "76mm",
                "varsayilan_word_path": "",
                "varsayilan_cikti_klasor": "",
                "log_export_klasor": "",
                "log_export_format": "JPG",
                "log_export_dpi": "300",
                "log_export_prefix": "Log",
                "cikti_merkezi_klasor": "",
                "cikti_merkezi_format": "JPG",
                "cikti_merkezi_dpi": "300",
                "taahhut_excel_sablon_path": "",
                "taahhut_ilgili_idare": "",
                "taahhut_tarih": "",
                "ek_tutanak_path": "",
                "ek_arazi_deneyli_path": "",
                "tutanak_sablon_path": "",
                "tutanak_sondaj_firma": "Kale Detay Sondaj",
                "tutanak_uygulama_sekli": "Burgusuz/Sulu",
                "tutanak_sondaj_makinesi": "SMK-500",
                "tutanak_jeofizik_cihaz": "GEODE",
                "tutanak_jeofon": "3,0m - 4,5 Hz",
                "tutanak_offset": "3,0m",
                "tutanak_kanal_sayisi": "12",
                "tutanak_kaynak": "Balyoz",
                "taahhut_jeoloji_ad": "Gökalp DOĞAN",
                "taahhut_jeoloji_sicil": "7400",
                "taahhut_jeoloji_unvan": "JEOLOJİ MÜHENDİSİ",
                "taahhut_jeoloji_imza_unvan": "Jeoloji Mühendisi",
                "taahhut_jeoloji_adres": "İsmetpaşa Mh. Hasan Mevsuf Sk. No :4 Da:5",
                "taahhut_jeoloji_telefon": "0 545 639 90 62",
                "taahhut_jeofizik_ad": "Suat ERGİN",
                "taahhut_jeofizik_sicil": "1982",
                "taahhut_jeofizik_unvan": "JEOFİZİK MÜHENDİSİ",
                "taahhut_jeofizik_imza_unvan": "Jeofizik Mühendisi",
                "taahhut_jeofizik_adres": "İsmetpaşa Mh. Hasan Mevsuf Sk. No :4 Da:5",
                "taahhut_jeofizik_telefon": "0 532 281 12 95",
                "yedek_sayisi": "10",
                "spt_guven_esigi": "90",
                "spt_auto_pro": "1"
            },
            "dosyalar": {"kml_path": None, "word_path": None, "lab_excel_path": None, "jeo_excel_path": None, "img_yer": None, "img_tkgm": None, "img_pga": None, "img_mjh": None, "word_img_sondaj": None, "word_img_jeofizik": None}
        }
        return default

    def veri_eksikleri_tamamla(self, hedef, varsayilan):
        for key, value in varsayilan.items():
            if key not in hedef:
                hedef[key] = value
            elif isinstance(value, dict) and isinstance(hedef[key], dict):
                self.veri_eksikleri_tamamla(hedef[key], value)
        return hedef

    def _dosya_yolu_coz(self, path):
        if path is None:
            return None
        path = str(path).strip().strip('"')
        if path in {"", "-", "None", "none", "null"}:
            return None
        path = os.path.expandvars(os.path.expanduser(path))
        if os.path.exists(path):
            return path
        if self.aktif_dosya_yolu:
            proje_klasoru = os.path.dirname(os.path.abspath(self.aktif_dosya_yolu))
            if not os.path.isabs(path):
                aday = os.path.join(proje_klasoru, path)
                if os.path.exists(aday):
                    return aday
            basename = os.path.basename(path)
            if basename:
                aday = os.path.join(proje_klasoru, basename)
                if os.path.exists(aday):
                    return aday
        return path

    def _dosya_yolu_al(self, dosyalar, key, *legacy_keys):
        kaynaklar = [dosyalar if isinstance(dosyalar, dict) else {}, self.veri if isinstance(self.veri, dict) else {}]
        for kaynak in kaynaklar:
            for aday_key in (key, *legacy_keys):
                value = kaynak.get(aday_key)
                resolved = self._dosya_yolu_coz(value)
                if resolved:
                    return resolved
        return None

    def _harita_gorsel_yolu_al(self, dosyalar, key, *legacy_keys):
        path = self._dosya_yolu_al(dosyalar, key, *legacy_keys)
        if eski_paylasimli_temp_harita_yolu_mu(path):
            return None
        return path

    def _harita_gorsel_yolu_kaydet(self, path):
        return None if eski_paylasimli_temp_harita_yolu_mu(path) else path

    def kml_etiket_guncelle(self):
        path = getattr(self, "kml_path", None)
        secili = bool(path)
        mevcut = bool(secili and os.path.exists(path))
        if mevcut:
            text = f"KML Sınır: {os.path.basename(path)}"
            color = "#27AE60"
        elif secili:
            text = f"KML Sınır: {os.path.basename(path)} (bulunamadı)"
            color = COLOR_WARNING
        else:
            text = "KML Sınır: Seçilmedi"
            color = "#333"
        if hasattr(self, 'lbl_kml_top'):
            self.lbl_kml_top.config(text=text, fg=color)
        if hasattr(self, "harita_durum_yenile"):
            self.harita_durum_yenile()

    @perf_tracked("project.load_default")
    def veri_yukle(self):
        default = self.varsayilan_veri_olustur()
        return default

    def get_yedek_sayisi(self):
        try:
            keep = int(str(self.veri.get("ayarlar", {}).get("yedek_sayisi", "10")).strip())
            return max(1, keep)
        except Exception:
            return 10

    def proje_durumu(self):
        return self.veri.setdefault("proje_durumu", {"tamamlandi": False, "kilitli": False, "tamamlanma_tarihi": "", "arsiv_notu": ""})

    def proje_kilitli_mi(self):
        return bool(self.proje_durumu().get("kilitli"))

    def proje_baslik_guncelle(self):
        if self.aktif_dosya_yolu:
            title = f"Zemin Rapor Pro - {os.path.basename(self.aktif_dosya_yolu)}"
        else:
            title = "Zemin Rapor Pro - Yeni Proje"
        if self.proje_kilitli_mi():
            title += " [KİLİTLİ]"
        self.root.title(title)

    def proje_kilit_durumunu_goster(self):
        durum = self.proje_durumu()
        if durum.get("kilitli"):
            tarih = durum.get("tamamlanma_tarihi") or "-"
            self.set_save_indicator(f"Proje kilitli: {tarih}", "warning")
        else:
            self.set_save_indicator("Proje açıldı", "info")

    def proje_tamamlandi_kilitle(self):
        self.guncelle_veri_objesi()
        if not self.aktif_dosya_yolu:
            if not messagebox.askyesno("Proje Kilitle", "Projeyi kilitlemeden önce kaydetmek gerekir. Şimdi kaydedilsin mi?"):
                return
            self.proje_farkli_kaydet()
            if not self.aktif_dosya_yolu:
                return
        if self.proje_kilitli_mi():
            messagebox.showinfo("Proje Kilitli", "Bu proje zaten tamamlandı olarak kilitli.")
            return
        lat, lon = proje_merkez_koordinati(self.veri)
        tarih = datetime.datetime.now().isoformat(timespec="seconds")
        self.proje_durumu().update({
            "tamamlandi": True,
            "kilitli": True,
            "tamamlanma_tarihi": tarih,
        })
        try:
            arsiv_kaydi_ekle(self.veri, self.aktif_dosya_yolu, kml_path=getattr(self, "kml_path", None))
        except Exception as exc:
            log_exception("project.archive.add", exc_value=exc)
            self.set_status(f"Arşiv kaydı oluşturulamadı: {exc}", level="warning")
        try:
            self._kilitli_kayda_izin_ver = True
            self.veri_kaydet()
        finally:
            self._kilitli_kayda_izin_ver = False
        self.proje_baslik_guncelle()
        msg = "Proje tamamlandı olarak kilitlendi ve arşiv listesine eklendi."
        if not lat or not lon:
            msg += "\n\nNot: KML haritasında görünmesi için Arazi sekmesinde proje merkezi veya sondaj koordinatı gerekir."
        messagebox.showinfo("Proje Kilitlendi", msg)

    def proje_kilidini_kaldir(self):
        if not self.proje_kilitli_mi():
            messagebox.showinfo("Proje Kilidi", "Bu proje kilitli değil.")
            return
        if not messagebox.askyesno("Kilidi Kaldır", "Proje kilidi kaldırılsın mı? Bundan sonra proje tekrar düzenlenip kaydedilebilir."):
            return
        self.proje_durumu().update({"tamamlandi": False, "kilitli": False})
        try:
            arsiv_kaydi_sil(self.aktif_dosya_yolu)
        except Exception as exc:
            log_exception("project.archive.remove", exc_value=exc)
        try:
            self._kilitli_kayda_izin_ver = True
            self.veri_kaydet()
        finally:
            self._kilitli_kayda_izin_ver = False
        self.proje_baslik_guncelle()
        self.set_save_indicator("Kilit kaldırıldı", "info")
        messagebox.showinfo("Kilidi Kaldır", "Proje kilidi kaldırıldı.")

    def biten_isler_kml_olustur(self):
        records = arsiv_kayitlari_yukle()
        if not records:
            messagebox.showinfo("Biten İşler KML", "Henüz tamamlandı olarak kilitlenmiş proje yok.")
            return
        path = filedialog.asksaveasfilename(
            title="Biten işler KML kaydet",
            defaultextension=".kml",
            initialfile="RaporPro_Biten_Isler.kml",
            filetypes=[("KML Dosyası", "*.kml"), ("Tüm Dosyalar", "*.*")],
        )
        if not path:
            return
        try:
            info = biten_isler_kml_yaz(records, path)
            self.set_status(f"Biten işler KML oluşturuldu: {os.path.basename(path)}", level="success")
            messagebox.showinfo(
                "Biten İşler KML",
                f"KML oluşturuldu:\n{path}\n\nHaritaya eklenen iş: {info['written']}\nKoordinatı olmadığı için atlanan: {info['skipped']}",
            )
        except Exception as exc:
            log_exception("project.archive.kml", exc_value=exc)
            messagebox.showerror("Biten İşler KML", f"KML oluşturulamadı:\n{exc}")

    @perf_tracked("project.save")
    def veri_kaydet(self):
        self.guncelle_veri_objesi()
        if self.proje_kilitli_mi() and not getattr(self, "_kilitli_kayda_izin_ver", False):
            messagebox.showwarning("Proje Kilitli", "Bu proje tamamlandı olarak kilitli. Kaydetmek için önce Proje > Proje Kilidini Kaldır komutunu kullanın.")
            self.set_save_indicator("Kilitli: kaydedilmedi", "warning")
            return
        if self.aktif_dosya_yolu:
            try:
                backup_path, backup_error = backup_project_file(self.aktif_dosya_yolu, keep=self.get_yedek_sayisi())
                if backup_error:
                    self.set_status(f"Yedekleme uyarısı: {backup_error}", level="warning")
                atomic_json_dump(self.veri, self.aktif_dosya_yolu, indent=4, ensure_ascii=False)
                self.set_status(f"Kaydedildi: {os.path.basename(self.aktif_dosya_yolu)}", level="success")
                self.last_save_time = datetime.datetime.now()
                self.set_save_indicator(f"Son kayıt: {self.last_save_time.strftime('%H:%M')}", "success")
                self.recent_project_ekle(self.aktif_dosya_yolu)
                if backup_path:
                    self.set_status(f"Yedek oluşturuldu: {os.path.basename(backup_path)}", level="info")
            except Exception as e:
                self.set_status(f"Kayıt Hatası: {str(e)}", level="error")
                self.set_save_indicator("Kayıt hatası", "error")
        else:
            self.proje_farkli_kaydet()

    @perf_tracked("project.save_as")
    def proje_farkli_kaydet(self):
        self.guncelle_veri_objesi()
        if self.proje_kilitli_mi() and not getattr(self, "_kilitli_kayda_izin_ver", False):
            messagebox.showwarning("Proje Kilitli", "Bu proje kilitli. Farklı kaydetmek için önce kilidi kaldırın.")
            self.set_save_indicator("Kilitli: farklı kaydedilmedi", "warning")
            return
        proje_adi = self.veri["kunye"].get("sahibi", "Yeni_Proje")
        if not proje_adi: proje_adi = "Zemin_Etud_Projesi"
        varsayilan_isim = f"{proje_adi}.json"

        dosya_yolu = filedialog.asksaveasfilename(
            defaultextension=".json",
            initialfile=varsayilan_isim,
            filetypes=[("RaporPro Projesi", "*.json"), ("Tüm Dosyalar", "*.*")],
            title="Projeyi Farklı Kaydet"
        )
        if dosya_yolu:
            try:
                backup_path, backup_error = backup_project_file(dosya_yolu, keep=self.get_yedek_sayisi())
                if backup_error:
                    self.set_status(f"Yedekleme uyarısı: {backup_error}", level="warning")
                atomic_json_dump(self.veri, dosya_yolu, indent=4, ensure_ascii=False)
                self.aktif_dosya_yolu = dosya_yolu
                self.root.title(f"Zemin Rapor Pro - {os.path.basename(dosya_yolu)}")
                self.set_status(f"Yeni proje olarak kaydedildi: {dosya_yolu}", level="success")
                self.last_save_time = datetime.datetime.now()
                self.set_save_indicator(f"Son kayıt: {self.last_save_time.strftime('%H:%M')}", "success")
                self.recent_project_ekle(dosya_yolu)
                if backup_path:
                    self.set_status(f"Yedek oluşturuldu: {os.path.basename(backup_path)}", level="info")
            except Exception as e:
                messagebox.showerror("Hata", f"Dosya kaydedilemedi:\n{str(e)}")
                self.set_save_indicator("Kayıt hatası", "error")

    @perf_tracked("project.open")
    def proje_ac(self):
        dosya_yolu = filedialog.askopenfilename(
            filetypes=[("RaporPro Projesi", "*.json"), ("Tüm Dosyalar", "*.*")],
            title="Proje Aç"
        )
        if dosya_yolu:
            try:
                self.proje_dosyasi_yukle(dosya_yolu)
            except Exception as e:
                messagebox.showerror("Hata", f"Dosya açılamadı:\n{str(e)}")

    @perf_tracked("project.new")
    def yeni_proje(self):
        if messagebox.askyesno("Yeni Proje", "Mevcut çalışma kaydedilmemişse kaybolacaktır. Yeni proje oluşturmak istiyor musunuz?"):
            self.yeni_proje_sihirbazi()

    def reset_dosya_baglantilari(self):
        self.aktif_dosya_yolu = None
        self.kml_path = None
        self.word_path = None
        self.lab_excel_path = None
        self.jeo_excel_path = None
        self.img_yer = None
        self.img_tkgm = None
        self.img_pga = None
        self.img_mjh = None
        self.word_img_sondaj = None
        self.word_img_jeofizik = None
        self.ek_tutanak_path = None
        self.ek_arazi_deneyli_path = None

    def yeni_proje_sihirbazi(self):
        mevcut_ayarlar = self.veri.get("ayarlar", {}).copy()
        win = Toplevel(self.root)
        self.pencere_hazirla(win, "Yeni Proje Sihirbazı", "560x430", (520, 390), modal=True)
        body = ttk.Frame(win, padding=16)
        body.pack(fill="both", expand=True)

        fields = {}
        defaults = [
            ("Proje adi", "sahibi", ""),
            ("Il", "il", ""),
            ("Ilce", "ilce", ""),
            ("Mahalle", "mah", ""),
            ("Sondaj sayisi", "count", "3"),
            ("Varsayilan derinlik (m)", "depth", "15.0"),
        ]
        for row, (label, key, value) in enumerate(defaults):
            ttk.Label(body, text=label).grid(row=row, column=0, sticky="e", padx=6, pady=6)
            entry = ttk.Entry(body, width=36)
            entry.grid(row=row, column=1, sticky="ew", padx=6, pady=6)
            entry.insert(0, value)
            fields[key] = entry

        kesit_var = tk.StringVar(value="line_projection")
        ttk.Label(body, text="Kesit mantigi").grid(row=len(defaults), column=0, sticky="e", padx=6, pady=6)
        ttk.Combobox(body, textvariable=kesit_var, values=("line_projection", "true_distance", "schematic"), state="readonly", width=34).grid(row=len(defaults), column=1, sticky="ew", padx=6, pady=6)

        word_var = tk.StringVar(value=mevcut_ayarlar.get("varsayilan_word_path", ""))
        ttk.Label(body, text="Word sablonu").grid(row=len(defaults) + 1, column=0, sticky="e", padx=6, pady=6)
        word_entry = ttk.Entry(body, textvariable=word_var, width=36)
        word_entry.grid(row=len(defaults) + 1, column=1, sticky="ew", padx=6, pady=6)
        tk.Button(body, text="Seç", command=lambda: self._ayar_dosya_sec(word_entry, [("Word", "*.docx")]), bg="#ECF0F1").grid(row=len(defaults) + 1, column=2, padx=4, pady=6)

        demo_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(body, text="Örnek litoloji ve SPT satırlarıyla başlat", variable=demo_var).grid(row=len(defaults) + 2, column=0, columnspan=2, sticky="w", pady=(8, 2))
        body.columnconfigure(1, weight=1)

        def olustur():
            count = max(0, int(safe_float(fields["count"].get()) or 0))
            depth = safe_float(fields["depth"].get()) or 15.0
            self.veri = self.varsayilan_veri_olustur()
            self.veri["ayarlar"].update(mevcut_ayarlar)
            self.veri["kunye"].update({
                "sahibi": fields["sahibi"].get().strip(),
                "il": fields["il"].get().strip(),
                "ilce": fields["ilce"].get().strip(),
                "mah": fields["mah"].get().strip(),
            })
            self.veri["sondaj"] = []
            for idx in range(count):
                sondaj = wb_yeni_sondaj_sablonu(idx)
                sondaj["der"] = f"{depth:.1f}"
                if demo_var.get():
                    split = max(1.0, depth * 0.45)
                    sondaj["k"] = f"{100 - idx * 1.2:.2f}"
                    sondaj["litoloji"] = [[0, f"{split:.2f}", "Killi kum"], [f"{split:.2f}", f"{depth:.2f}", "Siltli kil"]]
                    d = 1.5
                    while d <= depth + 0.01:
                        sondaj["spt"].append([f"{d:.2f}", "2", "4", "5", "9"])
                        d += 1.5
                self.veri["sondaj"].append(sondaj)
            self.veri["kesit_ayarlari"] = {"mode": kesit_var.get(), "selected_sondajlar": [s.get("no", "") for s in self.veri["sondaj"]]}
            if word_var.get().strip():
                self.veri["dosyalar"]["word_path"] = word_var.get().strip()
            self.reset_dosya_baglantilari()
            self.word_path = word_var.get().strip() or None
            self.doldur_arayuz()
            self.root.title("Zemin Rapor Pro - Yeni Proje")
            self.set_save_indicator("Yeni proje: kaydedilmedi", "warning")
            self.set_status("Yeni proje sihirbazla oluşturuldu.", level="success")
            win.destroy()

        btns = ttk.Frame(body)
        btns.grid(row=len(defaults) + 3, column=0, columnspan=3, sticky="ew", pady=(18, 0))
        tk.Button(btns, text="Oluştur", command=olustur, bg=COLOR_SUCCESS, fg="white", font=FONT_BOLD).pack(side="right", padx=4)
        tk.Button(btns, text="Vazgeç", command=win.destroy, bg="#ECF0F1").pack(side="right", padx=4)

    def ornek_proje_yukle(self):
        if not messagebox.askyesno("Örnek Proje", "Mevcut çalışma kaydedilmemişse kaybolabilir. Örnek proje yüklensin mi?"):
            return
        mevcut_ayarlar = self.veri.get("ayarlar", {}).copy()
        self.veri = self.varsayilan_veri_olustur()
        self.veri["ayarlar"].update(mevcut_ayarlar)
        self.veri["kunye"].update({"sahibi": "Örnek Zemin Etüdü", "il": "İstanbul", "ilce": "Kadıköy", "mah": "Örnek Mahalle", "ada": "123", "par": "45"})
        self.veri["bina"].update({"kul": "Konut", "kat": "5", "bod": "1", "temel_alan": "450", "der": "3.0"})
        self.veri["arazi"].update({"yon": "Guney", "egim": "5", "min": "98", "max": "104", "ort": "101", "pga": "0.40"})
        coords = [("40.990100", "29.030100"), ("40.990240", "29.030420"), ("40.990420", "29.030760")]
        lithologies = [
            [[0, 1.2, "Bitkisel toprak"], [1.2, 6.0, "Killi kum"], [6.0, 15.0, "Siltli kil"]],
            [[0, 1.0, "Dolgu"], [1.0, 5.5, "Kumlu kil"], [5.5, 15.0, "Cakilli kum"]],
            [[0, 1.4, "Bitkisel toprak"], [1.4, 7.0, "Siltli kil"], [7.0, 15.0, "Kum"]],
        ]
        self.veri["sondaj"] = []
        for idx in range(3):
            sondaj = wb_yeni_sondaj_sablonu(idx)
            sondaj.update({"der": "15.0", "k": f"{102 - idx * 1.4:.2f}", "y": coords[idx][0], "x": coords[idx][1]})
            sondaj["litoloji"] = lithologies[idx]
            sondaj["spt"] = [["1.50", "2", "3", "4", "7"], ["3.00", "3", "5", "6", "11"], ["4.50", "4", "6", "8", "14"], ["6.00", "5", "8", "10", "18"], ["7.50", "6", "10", "12", "22"], ["9.00", "8", "12", "15", "27"], ["10.50", "10", "15", "18", "33"], ["12.00", "12", "18", "20", "38"], ["13.50", "15", "20", "25", "45"]]
            self.veri["sondaj"].append(sondaj)
        self.veri["jeofizik"]["ss_list"] = [{"ad": "SS-1", "coords": ["40.990100", "29.030100", "40.990240", "29.030420", "40.990420", "29.030760"], "layers": [{"h": "6", "vp": "650", "vs": "220"}, {"h": "", "vp": "1200", "vs": "420"}]}]
        self.veri["jeofizik"]["mt_list"] = [{"no": "MT-1", "y": "40.990240", "x": "29.030420", "freq": "2.5", "to": "0.40", "ta": "0.28", "tb": "0.62", "hv": "3.2", "sure": "30"}]
        self.veri["kesit_ayarlari"] = {"mode": "line_projection", "selected_sondajlar": ["SK-1", "SK-2", "SK-3"], "line_start_no": "SK-1", "line_start_y": coords[0][0], "line_start_x": coords[0][1], "line_end_no": "SK-3", "line_end_y": coords[2][0], "line_end_x": coords[2][1]}
        self.reset_dosya_baglantilari()
        self.doldur_arayuz()
        self.root.title("Zemin Rapor Pro - Örnek Proje")
        self.set_save_indicator("Örnek proje: kaydedilmedi", "warning")
        self.set_status("Örnek proje yüklendi.", level="success")

    def proje_sablon_penceresi(self):
        win = Toplevel(self.root)
        self.pencere_hazirla(win, "Proje Şablonları", "520x360", (480, 320), modal=True)
        ttk.Label(win, text="Yeni proje şablonu seçin", font=("Segoe UI", 13, "bold")).pack(anchor="w", padx=16, pady=(16, 8))
        templates = [
            ("Standart 3x15", 3, 15.0, "3 sondaj, 15 m"),
            ("Orta 5x20", 5, 20.0, "5 sondaj, 20 m"),
            ("Detay 8x25", 8, 25.0, "8 sondaj, 25 m"),
            ("Sadece Rapor", 0, 0.0, "Sondajsız rapor dosya bağlantıları"),
            ("Jeofizikli Proje", 3, 15.0, "3 sondaj + jeofizik hazır alan"),
        ]
        selected = tk.StringVar(value=templates[0][0])
        for label, count, depth, desc in templates:
            ttk.Radiobutton(win, text=f"{label} - {desc}", variable=selected, value=label).pack(anchor="w", padx=22, pady=5)

        def apply_template():
            template = next(item for item in templates if item[0] == selected.get())
            self.proje_sablonu_uygula(template)
            win.destroy()

        tk.Button(win, text="Şablonu Uygula", command=apply_template, bg=COLOR_PRIMARY, fg="white", font=FONT_BOLD).pack(fill="x", padx=18, pady=18)

    @perf_tracked("project.template_apply")
    def proje_sablonu_uygula(self, template):
        label, count, depth, desc = template
        mevcut_ayarlar = self.veri.get("ayarlar", {}).copy()
        self.veri = self.varsayilan_veri_olustur()
        self.veri["ayarlar"].update(mevcut_ayarlar)
        self.veri["sondaj"] = []
        for idx in range(count):
            sondaj = wb_yeni_sondaj_sablonu(idx)
            sondaj["der"] = f"{depth:.1f}"
            self.veri["sondaj"].append(sondaj)
        if label == "Jeofizikli Proje":
            self.veri["jeofizik"]["ss_list"] = [{"ad": "SS-1", "coords": [""] * 6, "layers": []}]
            self.veri["jeofizik"]["mt_list"] = [{"no": "MT-1", "y": "", "x": ""}]
        self.aktif_dosya_yolu = None
        self.doldur_arayuz()
        self.root.title(f"Zemin Rapor Pro - {label}")
        self.set_status(f"Proje şablonu uygulandı: {label}", level="success")

    def _dosya_map(self):
        return {
            "word_path": self.word_path,
            "lab_excel_path": self.lab_excel_path,
            "jeo_excel_path": self.jeo_excel_path,
            "kml_path": self.kml_path,
            "img_yer": self.img_yer,
            "img_tkgm": self.img_tkgm,
            "img_pga": self.img_pga,
            "img_mjh": self.img_mjh,
            "word_img_sondaj": self.word_img_sondaj,
            "word_img_jeofizik": self.word_img_jeofizik,
            "ek_tutanak_path": getattr(self, "ek_tutanak_path", None) or self.veri.get("ayarlar", {}).get("ek_tutanak_path"),
            "ek_arazi_deneyli_path": getattr(self, "ek_arazi_deneyli_path", None) or self.veri.get("ayarlar", {}).get("ek_arazi_deneyli_path"),
        }

    @perf_tracked("project.export_data")
    def veri_disari_aktar(self):
        self.guncelle_veri_objesi(silent=True)
        path = filedialog.asksaveasfilename(title="Proje Verisini Excel'e Aktar", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter
            wb = Workbook()
            wb.remove(wb.active)
            rows, _ = wb_build_initial_rows(self.veri, WORKBOOK_SHEET_DEFS)
            for sheet_key, spec in WORKBOOK_SHEET_DEFS.items():
                ws = wb.create_sheet(spec["title"])
                ws.append([label for label, _ in spec["columns"]])
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="D9EAF7")
                for row in rows.get(sheet_key, []):
                    ws.append(row)
                for idx, width in enumerate(spec["widths"], start=1):
                    ws.column_dimensions[get_column_letter(idx)].width = max(10, width / 7)
            health = proje_saglik_ozeti(self.veri, self._dosya_map())
            ws = wb.create_sheet("Saglik")
            ws.append(["Durum", health["state"], health["score"]])
            ws.append(["Başlık", "Sonuç", "Detay"])
            for item in health["items"]:
                ws.append([item["label"], "OK" if item["ok"] else "EKSIK", item["detail"]])
            ws = wb.create_sheet("Hesap Ozeti")
            for line in format_hesap_ozeti(hesap_ozeti(self.veri)).splitlines():
                ws.append([line])
            wb.save(path)
            self.set_status(f"Proje verisi aktarıldı: {os.path.basename(path)}", level="success")
        except Exception as exc:
            log_exception("project.export_data", exc_value=exc)
            messagebox.showerror("Dışa Aktar", f"Veri aktarılamadı:\n{exc}")

    def gunluk_penceresi(self):
        win = Toplevel(self.root)
        self.pencere_hazirla(win, "RaporPro Günlükleri", "900x620", (760, 480), modal=True)
        notebook = ttk.Notebook(win)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        def read_tail(path, max_lines=500):
            if not os.path.exists(path):
                return "Henüz kayıt yok."
            try:
                with open(path, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                return "".join(lines[-max_lines:]) or "Henüz kayıt yok."
            except Exception as exc:
                return f"Günlük okunamadı:\n{exc}"

        for title, path in (("Performans", PERF_LOG_PATH), ("Hatalar", ERROR_LOG_PATH)):
            frame = ttk.Frame(notebook)
            notebook.add(frame, text=title)
            txt = tk.Text(frame, wrap="none", font=("Consolas", 9))
            txt.pack(fill="both", expand=True, padx=8, pady=8)
            txt.insert("1.0", read_tail(path))
            txt.config(state="disabled")

        tk.Button(win, text="Kapat", command=win.destroy, bg=COLOR_PRIMARY, fg="white").pack(pady=(0, 10))

    @perf_tracked("ui.build")
    def kur_arayuz(self):
        toolbar = tk.Frame(self.root, bg="#E9EEF2", height=44)
        toolbar.pack(fill="x", side="top")

        self.toolbar_menu(toolbar, "Proje", [
            ("Yeni Proje", self.yeni_proje),
            ("Proje Aç", self.proje_ac),
            ("Son Projeler", self.son_projeler_penceresi),
            None,
            ("Kaydet", self.veri_kaydet),
            ("Farklı Kaydet", self.proje_farkli_kaydet),
            ("Kurtarma Kaydını Aç", self.otomatik_kayit_yukle),
            None,
            ("Tamamlandı Olarak Kilitle", self.proje_tamamlandi_kilitle),
            ("Proje Kilidini Kaldır", self.proje_kilidini_kaldir),
            ("Biten İşler KML Oluştur", self.biten_isler_kml_olustur),
            None,
            ("Demo Proje", self.ornek_proje_yukle),
        ], bg="#F4F6F7", tooltip="Proje açma, kaydetme ve kurtarma işlemleri", role="primary")
        self.toolbar_menu(toolbar, "Veri", [
            ("Workbook", self.veri_giris_workbook_tksheet_ac),
            ("Sondaj Hızlı Tablo", self.sondaj_hizli_tablo_ac),
            ("Akıllı Tamamla", self.sondaj_akilli_tamamla),
            ("Şablonlar", self.proje_sablon_penceresi),
        ], bg="#D6EAF8", tooltip="Veri girişini hızlandıran araçlar", role="accent")
        self.toolbar_menu(toolbar, "Kontrol", [
            ("Final Kontrol", self.final_kontrol_penceresi),
            ("Rapor Ön Kontrol", self.rapor_on_kontrol),
            ("Proje Özeti", lambda: self._workflow_git("ozet")),
            ("Günlükler", self.gunluk_penceresi),
        ], bg="#FADBD8", tooltip="Eksikleri ve rapor hazırlığını kontrol eder", role="warning")
        self.toolbar_menu(toolbar, "Çıktı", [
            ("Çıktı Merkezi", self.cikti_merkezi_penceresi),
            ("Dışa Aktar", self.veri_disari_aktar),
            ("Toplu Log Kaydet", self.toplu_log_kaydet),
            ("Sadece Grafikleri Çıkar", self.grafikleri_kaydet),
        ], bg="#D5F5E3", tooltip="Rapor, log, kesit ve görsel çıktılarını toplar", role="success")
        self.toolbar_menu(toolbar, "Araçlar", [
            ("KML Sınır Seç", self.kml_sec),
            ("Tüm Koordinatları Seç", self.toplu_harita_ac),
            None,
            ("Ayarlar", self.ayarlar_penceresi),
            ("Etiketler", self.etiket_yoneticisi),
            None,
            ("Geri", self.global_undo),
            ("İleri", self.global_redo),
        ], bg="#E8DAEF", tooltip="Harita, ayar ve geri alma araçları", role="secondary")

        tk.Frame(toolbar, width=12, bg="#E9EEF2").pack(side="left")
        self.lbl_kml_top = tk.Label(toolbar, text="KML Sınır: Seçilmedi", bg="#E9EEF2", fg="#333", font=("Arial", 8))
        self.lbl_kml_top.pack(side="left", padx=10)
        self.tooltip_ekle(self.lbl_kml_top, "Seçili KML sınır dosyasının durumunu gösterir")

        self.autosave_status_label = tk.Label(toolbar, textvariable=self.autosave_status_var, bg="#E9EEF2", fg="#333333", font=("Arial", 8, "bold"))
        self.autosave_status_label.pack(side="right", padx=10)
        self.tooltip_ekle(self.autosave_status_label, "Otomatik kayıt ve proje kayıt durumunu gösterir")
        
        main_splitter = tk.PanedWindow(self.root, orient=tk.VERTICAL, sashwidth=4, bg=COLOR_BG)
        main_splitter.pack(fill="both", expand=True)
        top_frame = ttk.Frame(main_splitter)
        nb = ttk.Notebook(top_frame)
        self.nb = nb
        nb.pack(fill="both", expand=True)
        main_splitter.add(top_frame, height=750)
        
        self.tab_ozet=ttk.Frame(nb); nb.add(self.tab_ozet, text="0. Özet"); self.p_ozet(self.tab_ozet)
        self.tab_kunye=ttk.Frame(nb); nb.add(self.tab_kunye, text="1. Künye"); self.p_kunye(self.tab_kunye)
        self.tab_bina=ttk.Frame(nb); nb.add(self.tab_bina, text="2. Bina"); self.p_bina(self.tab_bina)
        self.tab_arazi=ttk.Frame(nb); nb.add(self.tab_arazi, text="3. Arazi"); self.p_arazi(self.tab_arazi)
        self.tab_sondaj=ttk.Frame(nb); nb.add(self.tab_sondaj, text="4. Sondaj"); self.p_sondaj(self.tab_sondaj)
        self.tab_jeofizik=ttk.Frame(nb); nb.add(self.tab_jeofizik, text="5. Jeofizik"); self.p_jeofizik(self.tab_jeofizik)
        self.tab_haritalar=ttk.Frame(nb); nb.add(self.tab_haritalar, text="6. Haritalar"); self.p_haritalar(self.tab_haritalar)
        self.tab_rapor=ttk.Frame(nb); nb.add(self.tab_rapor, text="7. Rapor"); self.p_rapor(self.tab_rapor)
        nb.bind("<<NotebookTabChanged>>", self.notebook_tab_changed)
        
        log_frame = tk.Frame(main_splitter, bg=COLOR_LOG_BG)
        self.log_text = tk.Text(log_frame, height=8, bg=COLOR_LOG_BG, fg=COLOR_LOG_TEXT, font=FONT_LOG, state="disabled")
        self.log_text.pack(fill="both", expand=True)
        main_splitter.add(log_frame, height=200)
        self.log_text.tag_config("error", foreground=COLOR_DANGER)
        self.log_text.tag_config("success", foreground="#00FF00")
        self.log_text.tag_config("warning", foreground=COLOR_WARNING)
        self.log_text.tag_config("normal", foreground=COLOR_LOG_TEXT)

    def kml_sec(self):
        f = filedialog.askopenfilename(filetypes=[("KML Dosyası", "*.kml")])
        if f: 
            self.kml_path = f
            self.veri.setdefault("dosyalar", {})["kml_path"] = f
            self.kml_etiket_guncelle()
            self.set_status("KML Altlığı Yüklendi.", level="success")

    @perf_tracked("map.bulk_open")
    def toplu_harita_ac(self):
        self.guncelle_veri_objesi()
        def coord_pair(y, x):
            yv, xv = safe_float(y), safe_float(x)
            return (yv, xv) if yv != 0 and xv != 0 else None

        initial = {"alan": None, "sondaj": {}, "ss": {}, "mt": {}}
        initial["alan"] = coord_pair(self.veri.get("arazi", {}).get("alan_y"), self.veri.get("arazi", {}).get("alan_x"))
        for idx, sondaj in enumerate(self.veri.get("sondaj", [])):
            coords = coord_pair(sondaj.get("y"), sondaj.get("x"))
            if coords:
                initial["sondaj"][idx] = coords
        for idx, ss in enumerate(self.veri.get("jeofizik", {}).get("ss_list", [])):
            coords = ss.get("coords", [])
            if len(coords) >= 6:
                parsed = [safe_float(value) for value in coords[:6]]
                if parsed[0] and parsed[1] and parsed[4] and parsed[5]:
                    if not (parsed[2] and parsed[3]):
                        parsed[2] = (parsed[0] + parsed[4]) / 2
                        parsed[3] = (parsed[1] + parsed[5]) / 2
                    initial["ss"][idx] = parsed
        for idx, mt in enumerate(self.veri.get("jeofizik", {}).get("mt_list", [])):
            coords = coord_pair(mt.get("y"), mt.get("x"))
            if coords:
                initial["mt"][idx] = coords

        map_data = {
            "sondaj": [s.get("no", f"SK-{i+1}") for i, s in enumerate(self.veri["sondaj"])],
            "ss": [s.get("ad", f"SS-{i+1}") for i, s in enumerate(self.veri["jeofizik"]["ss_list"])],
            "mt": [m.get("no", f"MT-{i+1}") for i, m in enumerate(self.veri["jeofizik"]["mt_list"])],
            "initial": initial,
        }
        TopluHarita(self.root, kml_path=self.kml_path, map_data=map_data, callback=self.toplu_koordinat_kaydet)

    def toplu_koordinat_kaydet(self, results):
        if results.get("alan"):
            self.veri["arazi"]["alan_y"] = f"{results['alan'][0]:.6f}"
            self.veri["arazi"]["alan_x"] = f"{results['alan'][1]:.6f}"
            
        for idx, coords in results.get("sondaj", {}).items():
            self.veri["sondaj"][idx]["y"] = f"{coords[0]:.6f}"
            self.veri["sondaj"][idx]["x"] = f"{coords[1]:.6f}"
            
        for idx, coords in results.get("ss", {}).items():
            str_coords = [f"{c:.6f}" for c in coords]
            self.veri["jeofizik"]["ss_list"][idx]["coords"] = str_coords
            
        for idx, coords in results.get("mt", {}).items():
            self.veri["jeofizik"]["mt_list"][idx]["y"] = f"{coords[0]:.6f}"
            self.veri["jeofizik"]["mt_list"][idx]["x"] = f"{coords[1]:.6f}"
            
        self.doldur_arayuz()
        self.set_status("Tüm harita koordinatları arayüze aktarıldı!", level="success")

    def global_undo(self):
        widget = self.root.focus_get()
        if not callable(getattr(widget, "undo", None)) and self.last_focused:
            widget = self.last_focused
        if callable(getattr(widget, "undo", None)):
            widget.undo()
    def global_redo(self):
        widget = self.root.focus_get()
        if not callable(getattr(widget, "redo", None)) and self.last_focused:
            widget = self.last_focused
        if callable(getattr(widget, "redo", None)):
            widget.redo()

    def ayarlar_penceresi(self):
        self.veri_eksikleri_tamamla(self.veri, self.varsayilan_veri_olustur())
        ayarlar = self.veri.setdefault("ayarlar", {})
        win = Toplevel(self.root)
        self.pencere_hazirla(win, "Ayarlar", "760x560", (680, 480), modal=True)

        nb = ttk.Notebook(win)
        nb.pack(fill="both", expand=True, padx=12, pady=12)
        form = ttk.Frame(nb, padding=15)
        taahhut_form = ttk.Frame(nb, padding=15)
        nb.add(form, text="Genel")
        nb.add(taahhut_form, text="Taahhütname")

        fields = [
            ("Firma adı", "firma_adi"),
            ("Log başlığı", "log_baslik"),
            ("Sorumlu unvanı", "sorumlu_muhendis_unvan"),
            ("Sorumlu mühendis", "sorumlu_muhendis"),
            ("Sondör başlığı", "sondor_belge_baslik"),
            ("Sondor / belge", "sondor_belge"),
            ("Makine metodu", "makine_metodu"),
            ("SPT şahmerdan tipi", "spt_sahmerdan"),
            ("Delgi çapı", "delgi_capi"),
            ("Yedek sayısı", "yedek_sayisi"),
        ]

        entries = {}
        for row_idx, (label, key) in enumerate(fields):
            ttk.Label(form, text=label).grid(row=row_idx, column=0, sticky="e", padx=6, pady=5)
            entry = ttk.Entry(form, width=48)
            entry.grid(row=row_idx, column=1, sticky="ew", padx=6, pady=5)
            entry.insert(0, ayarlar.get(key, ""))
            entries[key] = entry

        start_row = len(fields)
        ttk.Label(form, text="Varsayılan Word şablonu").grid(row=start_row, column=0, sticky="e", padx=6, pady=5)
        word_entry = ttk.Entry(form, width=48)
        word_entry.grid(row=start_row, column=1, sticky="ew", padx=6, pady=5)
        word_entry.insert(0, ayarlar.get("varsayilan_word_path", ""))
        entries["varsayilan_word_path"] = word_entry
        tk.Button(form, text="Seç", command=lambda: self._ayar_dosya_sec(word_entry, [("Word", "*.docx")]), bg="#ECF0F1").grid(row=start_row, column=2, padx=6, pady=5)

        out_row = start_row + 1
        ttk.Label(form, text="Varsayılan çıktı klasörü").grid(row=out_row, column=0, sticky="e", padx=6, pady=5)
        out_entry = ttk.Entry(form, width=48)
        out_entry.grid(row=out_row, column=1, sticky="ew", padx=6, pady=5)
        out_entry.insert(0, ayarlar.get("varsayilan_cikti_klasor", ""))
        entries["varsayilan_cikti_klasor"] = out_entry
        tk.Button(form, text="Seç", command=lambda: self._ayar_klasor_sec(out_entry), bg="#ECF0F1").grid(row=out_row, column=2, padx=6, pady=5)

        form.columnconfigure(1, weight=1)

        ttk.Label(taahhut_form, text="İlgili idare").grid(row=0, column=0, sticky="e", padx=6, pady=5)
        idare_entry = ttk.Entry(taahhut_form, width=48)
        idare_entry.grid(row=0, column=1, columnspan=3, sticky="ew", padx=6, pady=5)
        idare_entry.insert(0, ayarlar.get("taahhut_ilgili_idare", ""))
        entries["taahhut_ilgili_idare"] = idare_entry

        ttk.Label(taahhut_form, text="Tarih").grid(row=1, column=0, sticky="e", padx=6, pady=5)
        tarih_entry = ttk.Entry(taahhut_form, width=18)
        tarih_entry.grid(row=1, column=1, sticky="w", padx=6, pady=5)
        tarih_entry.insert(0, ayarlar.get("taahhut_tarih", ""))
        entries["taahhut_tarih"] = tarih_entry
        ttk.Label(taahhut_form, text="Boşsa bugünün tarihi kullanılır. Yapı adresi ve yapı sahibinin adresi: Mahalle / İlçe / İl.").grid(row=1, column=2, columnspan=2, sticky="w", padx=6, pady=5)

        ttk.Label(taahhut_form, text="Excel şablonu").grid(row=2, column=0, sticky="e", padx=6, pady=5)
        taahhut_template_entry = ttk.Entry(taahhut_form, width=48)
        taahhut_template_entry.grid(row=2, column=1, columnspan=2, sticky="ew", padx=6, pady=5)
        taahhut_template_entry.insert(0, ayarlar.get("taahhut_excel_sablon_path", ""))
        entries["taahhut_excel_sablon_path"] = taahhut_template_entry
        self.modern_button(
            taahhut_form,
            text="Seç",
            command=lambda: self._ayar_dosya_sec(taahhut_template_entry, [("Excel", "*.xlsx")]),
            role="neutral",
            outline=True,
        ).grid(row=2, column=3, padx=6, pady=5)

        def taahhut_profile_frame(parent, title, prefix, col):
            frame = ttk.LabelFrame(parent, text=title, padding=10)
            frame.grid(row=3, column=col, columnspan=2, sticky="nsew", padx=6, pady=(12, 5))
            specs = [
                ("Ad Soyad", "ad"),
                ("Oda sicil no", "sicil"),
                ("Unvan", "unvan"),
                ("İmza unvanı", "imza_unvan"),
                ("Adres", "adres"),
                ("Telefon", "telefon"),
            ]
            for row_idx, (label, suffix) in enumerate(specs):
                key = f"{prefix}_{suffix}"
                ttk.Label(frame, text=label).grid(row=row_idx, column=0, sticky="e", padx=5, pady=4)
                entry = ttk.Entry(frame, width=28)
                entry.grid(row=row_idx, column=1, sticky="ew", padx=5, pady=4)
                entry.insert(0, ayarlar.get(key, ""))
                entries[key] = entry
            frame.columnconfigure(1, weight=1)

        taahhut_profile_frame(taahhut_form, "Jeoloji Mühendisi", "taahhut_jeoloji", 0)
        taahhut_profile_frame(taahhut_form, "Jeofizik Mühendisi", "taahhut_jeofizik", 2)
        for col in range(4):
            taahhut_form.columnconfigure(col, weight=1)

        def kaydet():
            for key, entry in entries.items():
                ayarlar[key] = entry.get().strip()
            try:
                keep = int(ayarlar.get("yedek_sayisi", "10"))
                ayarlar["yedek_sayisi"] = str(max(1, keep))
            except Exception:
                ayarlar["yedek_sayisi"] = "10"
                messagebox.showwarning("Ayarlar", "Yedek sayısı geçersizdi; 10 olarak ayarlandı.")

            self.ayarlari_uygula()
            self.set_status("Ayarlar güncellendi.", level="success")
            if self.aktif_dosya_yolu:
                self.veri_kaydet()
            win.destroy()

        btns = ttk.Frame(win, padding=(12, 0, 12, 12))
        btns.pack(fill="x")
        tk.Button(btns, text="Kaydet", command=kaydet, bg=COLOR_SUCCESS, fg="white", font=FONT_BOLD).pack(side="right", padx=5)
        tk.Button(btns, text="Vazgeç", command=win.destroy, bg="#ECF0F1").pack(side="right", padx=5)

    def _ayar_dosya_sec(self, entry, filetypes):
        path = filedialog.askopenfilename(filetypes=filetypes)
        if path:
            entry.delete(0, tk.END)
            entry.insert(0, path)

    def _ayar_klasor_sec(self, entry):
        path = filedialog.askdirectory()
        if path:
            entry.delete(0, tk.END)
            entry.insert(0, path)

    def ayarlari_uygula(self):
        ayarlar = self.veri.get("ayarlar", {})
        default_word = ayarlar.get("varsayilan_word_path")
        if not self.word_path and default_word and os.path.exists(default_word):
            self.word_path = default_word
            if hasattr(self, 'lbl_sab'):
                self.lbl_sab.config(text=os.path.basename(self.word_path), foreground=COLOR_SUCCESS)

    def etiket_yoneticisi(self):
        win = Toplevel(self.root)
        self.pencere_hazirla(win, "Word Etiket Yöneticisi", "980x640", (820, 520), modal=True)

        notebook = ttk.Notebook(win)
        notebook.pack(fill="both", expand=True, padx=10, pady=10)

        tab_supported = ttk.Frame(notebook)
        tab_template = ttk.Frame(notebook)
        notebook.add(tab_supported, text="Desteklenen Etiketler")
        notebook.add(tab_template, text="Şablon Analizi")

        supported = get_supported_tags()
        paned = tk.PanedWindow(tab_supported, orient=tk.HORIZONTAL, sashwidth=4, bg=COLOR_BG)
        paned.pack(fill="both", expand=True)

        left = ttk.Frame(paned, padding=8)
        right = ttk.Frame(paned, padding=8)
        paned.add(left, width=360)
        paned.add(right, width=580)

        search_var = tk.StringVar()
        ttk.Label(left, text="Ara").pack(anchor="w")
        search_entry = ttk.Entry(left, textvariable=search_var)
        search_entry.pack(fill="x", pady=(0, 6))

        tag_list = tk.Listbox(left, height=24, font=("Consolas", 10))
        tag_list.pack(fill="both", expand=True)

        detail = tk.Text(right, wrap="word", font=("Consolas", 10))
        detail.pack(fill="both", expand=True)

        filtered_items = []

        def refresh_tag_list(*_):
            filtered_items.clear()
            tag_list.delete(0, tk.END)
            query = search_var.get().strip().lower()
            for item in supported:
                haystack = f"{item['tag']} {item['category']} {item['description']}".lower()
                if query and query not in haystack:
                    continue
                filtered_items.append(item)
                tag_list.insert(tk.END, f"{item['tag']}  [{item['category']}]")
            if filtered_items:
                tag_list.selection_set(0)
                show_tag_detail()

        def show_tag_detail(event=None):
            sel = tag_list.curselection()
            if not sel:
                return
            item = filtered_items[sel[0]]
            detail.config(state="normal")
            detail.delete("1.0", tk.END)
            detail.insert(tk.END, f"Etiket: {item['tag']}\n")
            detail.insert(tk.END, f"Tur: {item['category']}\n\n")
            detail.insert(tk.END, item["description"])
            detail.config(state="disabled")

        search_var.trace_add("write", refresh_tag_list)
        tag_list.bind("<<ListboxSelect>>", show_tag_detail)
        refresh_tag_list()

        top = ttk.Frame(tab_template, padding=8)
        top.pack(fill="x")
        selected_path = tk.StringVar(value=self.word_path or "")
        ttk.Label(top, text="Şablon").pack(side="left", padx=(0, 6))
        path_entry = ttk.Entry(top, textvariable=selected_path)
        path_entry.pack(side="left", fill="x", expand=True, padx=(0, 6))

        analysis_text = tk.Text(tab_template, wrap="word", font=("Consolas", 10))
        analysis_text.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        def show_analysis(path):
            analysis = analyze_word_template(path)
            analysis_text.config(state="normal")
            analysis_text.delete("1.0", tk.END)
            analysis_text.insert("1.0", format_template_analysis(analysis))
            analysis_text.config(state="disabled")
            if analysis.get("error"):
                self.set_status("Etiket analizi tamamlanamadi.", level="error")
            elif analysis.get("unknown"):
                self.set_status(f"Etiket analizi {len(analysis['unknown'])} bilinmeyen etiket buldu.", level="warning")
            else:
                self.set_status("Etiket analizi tamamlandi.", level="success")

        def choose_and_scan():
            path = filedialog.askopenfilename(filetypes=[("Word", "*.docx")])
            if not path:
                return
            selected_path.set(path)
            self.word_path = path
            if hasattr(self, 'lbl_sab'):
                self.lbl_sab.config(text=os.path.basename(path), foreground=COLOR_SUCCESS)
            show_analysis(path)

        def set_default_template():
            path = selected_path.get().strip()
            if not path or not os.path.exists(path):
                messagebox.showwarning("Word", "Varsayılan yapmak için önce geçerli bir Word şablonu seçin.")
                return
            self.word_path = path
            self.veri.setdefault("ayarlar", {})["varsayilan_word_path"] = path
            if hasattr(self, 'lbl_sab'):
                self.lbl_sab.config(text=os.path.basename(path), foreground=COLOR_SUCCESS)
            self.set_status(f"Varsayılan Word şablonu ayarlandı: {os.path.basename(path)}", level="success")

        tk.Button(top, text="Tara", command=lambda: show_analysis(selected_path.get()), bg=COLOR_PRIMARY, fg="white").pack(side="left", padx=3)
        tk.Button(top, text="Word Seç", command=choose_and_scan, bg="#ECF0F1").pack(side="left", padx=3)
        tk.Button(top, text="Varsayılan Yap", command=set_default_template, bg="#D6EAF8").pack(side="left", padx=3)

        if selected_path.get():
            show_analysis(selected_path.get())
        else:
            analysis_text.insert("1.0", "Şablon analizi için önce Word dosyası seçin.")
            analysis_text.config(state="disabled")

    def notebook_tab_changed(self, event):
        if hasattr(self, "tab_ozet") and event.widget.select() == str(self.tab_ozet):
            self.ozet_yenile()
        elif hasattr(self, "tab_haritalar") and event.widget.select() == str(self.tab_haritalar):
            self.kml_etiket_guncelle()

    def p_ozet(self, p):
        outer = ttk.Frame(p, padding=14)
        outer.pack(fill="both", expand=True)

        top = ttk.Frame(outer)
        top.pack(fill="x", pady=(0, 6))
        ttk.Label(top, text="Proje Özeti", font=("Segoe UI", 14, "bold")).pack(side="left")
        tk.Button(top, text="Yenile", command=self.ozet_yenile, bg="#ECF0F1", fg="#111", relief="flat").pack(side="right")

        hero = ttk.Frame(outer)
        hero.pack(fill="x", pady=(0, 10))
        hero.columnconfigure(0, weight=2)
        hero.columnconfigure(1, weight=1)

        dashboard = tk.Frame(hero, bg="#FFFFFF", bd=1, relief="solid", padx=14, pady=12)
        dashboard.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        dashboard.columnconfigure(0, weight=1)
        tk.Label(
            dashboard,
            text="Bugünkü Durum",
            bg="#FFFFFF",
            fg=COLOR_PRIMARY,
            font=("Segoe UI", 10, "bold"),
            anchor="w",
        ).grid(row=0, column=0, sticky="ew")
        self.final_dashboard_status_label = tk.Label(
            dashboard,
            text="Proje durumu hazırlanıyor...",
            bg="#FFFFFF",
            fg="#333333",
            font=("Segoe UI", 17, "bold"),
            anchor="w",
        )
        self.final_dashboard_status_label.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        self.final_dashboard_detail_label = tk.Label(
            dashboard,
            text="Final kontrol, veri sağlığı ve ön kontrol sonuçları burada özetlenir.",
            bg="#FFFFFF",
            fg="#555555",
            font=("Segoe UI", 9),
            anchor="w",
            justify="left",
            wraplength=760,
        )
        self.final_dashboard_detail_label.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        self.ozet_missing_labels = []
        missing_frame = ttk.Frame(dashboard)
        missing_frame.grid(row=3, column=0, sticky="ew", pady=(10, 0))
        for idx in range(3):
            lbl = tk.Label(missing_frame, text="-", bg="#FFFFFF", fg="#555555", anchor="w", justify="left")
            lbl.pack(fill="x", pady=1)
            self.ozet_missing_labels.append(lbl)
        dashboard.bind(
            "<Configure>",
            lambda event: self.final_dashboard_detail_label.config(wraplength=max(260, event.width - 40)),
        )

        next_card = tk.Frame(hero, bg="#FFFFFF", bd=1, relief="solid", padx=14, pady=12)
        next_card.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
        next_card.columnconfigure(0, weight=1)
        tk.Label(next_card, text="Sıradaki İş", bg="#FFFFFF", fg=COLOR_PRIMARY, font=("Segoe UI", 10, "bold"), anchor="w").grid(row=0, column=0, sticky="ew")
        self.ozet_next_action_label = tk.Label(
            next_card,
            text="Proje durumu hesaplanıyor...",
            bg="#FFFFFF",
            fg="#333333",
            font=("Segoe UI", 11, "bold"),
            anchor="nw",
            justify="left",
            wraplength=300,
        )
        self.ozet_next_action_label.grid(row=1, column=0, sticky="nsew", pady=(10, 10))
        self.ozet_next_action_button = tk.Button(
            next_card,
            text="Final Kontrol",
            command=self.final_kontrol_penceresi,
            bg=COLOR_WARNING,
            fg="white",
            relief="flat",
            font=FONT_BOLD,
        )
        self.ozet_next_action_button.grid(row=2, column=0, sticky="ew")
        self.workflow_widgets = {}

        hero_layout_state = {"mode": None}

        def layout_hero(event=None):
            width = hero.winfo_width()
            if width <= 1 and event is not None:
                width = event.width
            mode = "stack" if width and width < 900 else "split"
            if hero_layout_state["mode"] == mode:
                return
            hero_layout_state["mode"] = mode
            dashboard.grid_forget()
            next_card.grid_forget()
            if mode == "stack":
                hero.columnconfigure(0, weight=1)
                hero.columnconfigure(1, weight=0)
                dashboard.grid(row=0, column=0, sticky="ew", padx=0, pady=(0, 8))
                next_card.grid(row=1, column=0, sticky="ew", padx=0, pady=0)
            else:
                hero.columnconfigure(0, weight=2)
                hero.columnconfigure(1, weight=1)
                dashboard.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=0)
                next_card.grid(row=0, column=1, sticky="nsew", padx=(8, 0), pady=0)

        hero.bind("<Configure>", layout_hero)
        self.root.after_idle(layout_hero)

        quick = ttk.LabelFrame(outer, text="Kısa Yollar", padding=10)
        quick.pack(fill="x", pady=(0, 10))
        quick_buttons = ttk.Frame(quick)
        quick_buttons.pack(fill="x")
        self.responsive_button_row(quick_buttons, [
            ("Workbook", self.veri_giris_workbook_tksheet_ac, "#D6EAF8"),
            ("SPT Merkezi", self.spt_okuma_merkezi_ac, "#A3E4D7"),
            ("Kesit", self.kesit_secim_penceresi, "#E8DAEF"),
            ("Haritalar", lambda: self._workflow_git("haritalar"), "#D6EAF8"),
            ("Final Kontrol", self.final_kontrol_penceresi, "#F5B7B1"),
            ("Rapor Oluştur", self.raporla, COLOR_SUCCESS),
        ], min_width=155, max_cols=6)

        body = ttk.Frame(outer)
        body.pack(fill="both", expand=True)

        left = ttk.Frame(body)
        right = ttk.Frame(body)
        body_layout_state = {"mode": None}

        def layout_summary_body(event=None):
            width = body.winfo_width()
            if width <= 1 and event is not None:
                width = event.width
            mode = "stack" if width and width < 980 else "split"
            if body_layout_state["mode"] == mode:
                return
            body_layout_state["mode"] = mode
            for child in (left, right):
                child.grid_forget()
            if mode == "stack":
                body.columnconfigure(0, weight=1)
                body.columnconfigure(1, weight=0)
                left.grid(row=0, column=0, sticky="nsew", pady=(0, 8))
                right.grid(row=1, column=0, sticky="nsew")
            else:
                body.columnconfigure(0, weight=1)
                body.columnconfigure(1, weight=0)
                left.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
                right.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
            body.rowconfigure(0, weight=1)
            body.rowconfigure(1, weight=1 if mode == "stack" else 0)

        body.bind("<Configure>", layout_summary_body)
        self.root.after_idle(layout_summary_body)

        left_top = ttk.Frame(left)
        left_top.pack(fill="x", pady=(0, 10))
        left_top.columnconfigure(0, weight=3)
        left_top.columnconfigure(1, weight=2)

        self.ozet_metric_labels = {}
        metrics_frame = ttk.LabelFrame(left_top, text="Veri Durumu", padding=12)
        metrics_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        metrics = [
            ("proje", "Proje"),
            ("konum", "Konum"),
            ("sondaj", "Sondaj"),
            ("litoloji", "Litoloji"),
            ("deney", "Arazi deneyleri"),
            ("jeofizik", "Jeofizik"),
            ("harita", "Harita/görsel"),
        ]
        for row, (key, label) in enumerate(metrics):
            ttk.Label(metrics_frame, text=label).grid(row=row, column=0, sticky="w", padx=5, pady=4)
            value = tk.Label(metrics_frame, text="-", bg=COLOR_BG, fg="#333333", anchor="w", justify="left")
            value.grid(row=row, column=1, sticky="ew", padx=5, pady=4)
            self.ozet_metric_labels[key] = value
        metrics_frame.columnconfigure(1, weight=1)

        health_frame = ttk.LabelFrame(left_top, text="Proje Sağlığı", padding=8)
        health_frame.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        self.health_status_label = tk.Label(health_frame, text="-", bg=COLOR_BG, fg="#333333", font=("Segoe UI", 10, "bold"), anchor="w")
        self.health_status_label.pack(fill="x")
        health_text_wrap = ttk.Frame(health_frame)
        health_text_wrap.pack(fill="both", expand=True, pady=(5, 0))
        self.health_detail_text = tk.Text(health_text_wrap, height=7, wrap="word", font=("Consolas", 8), bg="#FAFAFA")
        health_scroll = ttk.Scrollbar(health_text_wrap, orient="vertical", command=self.health_detail_text.yview)
        self.health_detail_text.configure(yscrollcommand=health_scroll.set)
        self.health_detail_text.pack(side="left", fill="both", expand=True)
        health_scroll.pack(side="right", fill="y")
        self.health_detail_text.config(state="disabled")
        self.health_tag_actions = {}
        self.health_detail_text.bind("<Button-1>", self._health_detail_click)
        self.health_detail_text.bind("<Motion>", self._health_detail_motion)

        self.ozet_file_labels = {}
        files_frame = ttk.LabelFrame(left, text="Dosya Bağlantıları", padding=6)
        files_frame.pack(fill="x", expand=False)
        files = [
            ("word", "Word şablonu"),
            ("lab", "Lab Excel"),
            ("jeo", "Jeofizik Excel"),
            ("kml", "KML sınır"),
            ("yer", "Yerbuldurur"),
            ("tkgm", "TKGM"),
            ("pga", "PGA"),
            ("mjh", "MJH"),
            ("sondaj_img", "Sondaj haritası"),
            ("jeo_img", "Jeofizik haritası"),
        ]
        for idx, (key, label) in enumerate(files):
            row = idx // 4
            base_col = (idx % 4) * 2
            ttk.Label(files_frame, text=label).grid(row=row, column=base_col, sticky="w", padx=(4, 3), pady=2)
            value = tk.Label(files_frame, text="-", bg=COLOR_BG, fg="#333333", anchor="w", justify="left", font=("Segoe UI", 8))
            value.grid(row=row, column=base_col + 1, sticky="ew", padx=(0, 8), pady=2)
            self.ozet_file_labels[key] = value
        files_frame.columnconfigure(1, weight=1)
        files_frame.columnconfigure(3, weight=1)
        files_frame.columnconfigure(5, weight=1)
        files_frame.columnconfigure(7, weight=1)

        preflight_frame = ttk.LabelFrame(right, text="Son Ön Kontrol", padding=8)
        preflight_frame.pack(fill="both", expand=True)
        self.ozet_preflight_text = tk.Text(preflight_frame, wrap="word", font=("Consolas", 8), height=8, width=46, bg="#FAFAFA")
        preflight_scroll = ttk.Scrollbar(preflight_frame, orient="vertical", command=self.ozet_preflight_text.yview)
        self.ozet_preflight_text.configure(yscrollcommand=preflight_scroll.set)
        self.ozet_preflight_text.pack(side="left", fill="both", expand=True)
        preflight_scroll.pack(side="right", fill="y")
        self.ozet_preflight_text.insert("1.0", "Ön kontrol henüz çalıştırılmadı.")
        self.ozet_preflight_text.config(state="disabled")

    def ozet_rapora_git(self):
        if hasattr(self, "nb") and hasattr(self, "tab_rapor"):
            self.nb.select(self.tab_rapor)

    def _workflow_git(self, target):
        tab_map = {
            "ozet": "tab_ozet",
            "kunye": "tab_kunye",
            "bina": "tab_bina",
            "arazi": "tab_arazi",
            "sondaj": "tab_sondaj",
            "jeofizik": "tab_jeofizik",
            "rapor": "tab_rapor",
            "haritalar": "tab_haritalar",
        }
        tab_attr = tab_map.get(target)
        if tab_attr and hasattr(self, tab_attr):
            self.nb.select(getattr(self, tab_attr))
            self.set_status(f"{target} sekmesine gidildi.", level="info")

    def _workflow_set(self, key, text, level):
        widget = getattr(self, "workflow_widgets", {}).get(key, {}).get("status")
        if not widget:
            return
        colors = {"ok": COLOR_SUCCESS, "warn": COLOR_WARNING, "bad": COLOR_DANGER, "info": "#333333"}
        widget.config(text=text, fg=colors.get(level, "#333333"))

    def _workflow_paneli_guncelle(self, health):
        if not hasattr(self, "workflow_widgets"):
            return
        items = {item.get("label"): item for item in health.get("items", [])}
        def ok(label):
            return bool(items.get(label, {}).get("ok"))

        project_ok = ok("Proje bilgisi")
        data_ok = ok("Sondaj kaydı") and ok("Litoloji") and ok("Arazi deneyleri")
        control_ok = health.get("score", 0) >= 85 and not (self.last_preflight_report or {}).get("errors")
        control_warn = health.get("score", 0) >= 60
        visual_ok = ok("Sondaj koordinatları") and bool(self.veri.get("kesit_ayarlari", {}).get("selected_sondajlar") or self.veri.get("sondaj"))
        report_ok = ok("Word şablonu") and control_warn

        self._workflow_set("project", "TAMAM" if project_ok else "EKSİK", "ok" if project_ok else "bad")
        self._workflow_set("data", "TAMAM" if data_ok else "VERİ GEREKLİ", "ok" if data_ok else "warn")
        self._workflow_set("control", "TAMAM" if control_ok else ("UYARI VAR" if control_warn else "KONTROL GEREKLİ"), "ok" if control_ok else "warn")
        self._workflow_set("visual", "HAZIR" if visual_ok else "KOORDİNAT/KESİT GEREKLİ", "ok" if visual_ok else "warn")
        self._workflow_set("report", "RAPORA HAZIR" if report_ok else "ŞABLON/KONTROL GEREKLİ", "ok" if report_ok else "warn")

    def _final_dashboard_guncelle(self, health):
        if not hasattr(self, "final_dashboard_status_label"):
            return
        score = health.get("score", 0)
        preflight = self.last_preflight_report or {}
        error_count = len(preflight.get("errors", []) or [])
        warning_count = len(preflight.get("warnings", []) or [])
        missing = [item.get("label", "") for item in health.get("items", []) if not item.get("ok")]
        if error_count:
            title = f"Rapor ön kontrolünde {error_count} hata var"
            color = COLOR_DANGER
            action = "Eksikleri Göster ile hata satırlarına gidebilirsiniz."
        elif score >= 85 and warning_count == 0:
            title = "Proje rapor almaya hazır görünüyor"
            color = COLOR_SUCCESS
            action = "Raporu Oluştur veya Çıktı Merkezi ile son çıktıları alabilirsiniz."
        elif score >= 60:
            title = "Proje iyi durumda, son kontroller gerekiyor"
            color = COLOR_WARNING
            action = "Final Kontrol ile kalan uyarıları temizlemek iyi olur."
        else:
            title = "Veri girişi tamamlandıkça proje hazır hale gelecek"
            color = COLOR_DANGER
            action = "İş Akışı kartları sıradaki eksik alana götürür."

        details = [f"Proje sağlığı: %{score}"]
        if warning_count:
            details.append(f"Ön kontrol uyarısı: {warning_count}")
        if missing:
            details.append("Eksik görünenler: " + ", ".join(missing[:4]))
            if len(missing) > 4:
                details[-1] += f" ve {len(missing) - 4} kalem daha"
        details.append(action)
        self.final_dashboard_status_label.config(text=title, fg=color)
        self.final_dashboard_detail_label.config(text=" | ".join(details))

        if hasattr(self, "ozet_missing_labels"):
            missing_items = [item for item in health.get("items", []) if not item.get("ok")]
            for idx, label in enumerate(self.ozet_missing_labels):
                if idx < len(missing_items):
                    item = missing_items[idx]
                    text = f"- {item.get('label')}: {item.get('detail')}"
                    label.config(text=text, fg=COLOR_DANGER)
                elif idx == 0:
                    label.config(text="- Kritik eksik görünmüyor.", fg=COLOR_SUCCESS)
                else:
                    label.config(text="", fg="#555555")

        if hasattr(self, "ozet_next_action_label") and hasattr(self, "ozet_next_action_button"):
            missing_items = [item for item in health.get("items", []) if not item.get("ok")]
            if error_count:
                next_text = f"Ön kontrolde {error_count} hata var. Önce hatalı maddeleri temizleyelim."
                btn_text = "Hataları Aç"
                btn_color = COLOR_DANGER
                btn_command = self.final_kontrol_penceresi
            elif missing_items:
                first = missing_items[0]
                suggestion = first.get("suggestion") or first.get("detail") or "Eksik bilgiyi tamamlayın."
                next_text = f"{first.get('label')}: {suggestion}"
                btn_text = "İlgili Sekmeye Git"
                btn_color = COLOR_WARNING
                btn_command = lambda target=first.get("target", "ozet"): self._workflow_git(target)
            elif warning_count:
                next_text = f"Ön kontrolde {warning_count} uyarı var. Son kontrolü açıp karar verelim."
                btn_text = "Final Kontrol"
                btn_color = COLOR_WARNING
                btn_command = self.final_kontrol_penceresi
            else:
                next_text = "Eksik görünmüyor. Raporu oluşturabilir veya çıktı merkezinden son dosyaları toplayabilirsiniz."
                btn_text = "Raporu Oluştur"
                btn_color = COLOR_SUCCESS
                btn_command = self.raporla
            self.ozet_next_action_label.config(text=next_text, fg=btn_color)
            self.ozet_next_action_button.config(text=btn_text, bg=btn_color, fg="white", command=btn_command)

    def p_kunye(self, p):
        f = ttk.LabelFrame(p, text="Proje Künyesi", padding="20")
        f.pack(fill="x", padx=10, pady=10)
        self.e_kunye = {}
        fields = [("Proje Sahibi:", "sahibi"), ("İl:", "il"), ("İlçe:", "ilce"), ("Mahalle:", "mah"), ("Mevkii:", "mev"), ("Pafta:", "paf"), ("Ada:", "ada"), ("Parsel:", "par")]
        for i, (lbl, key) in enumerate(fields):
            r = i // 2; c = (i % 2) * 2
            ttk.Label(f, text=lbl).grid(row=r, column=c, sticky="e", padx=5, pady=8)
            e = UndoRedoEntry(f, width=30); e.grid(row=r, column=c+1, sticky="w", padx=5, pady=8)
            self.e_kunye[key] = e

    def bina_blok_kolonlari(self):
        return [
            ("Blok", "blok_adi", 12),
            ("Kullanım", "kul", 18),
            ("Sınıf", "sinif", 14),
            ("Önem", "onem", 10),
            ("Malzeme", "malz", 14),
            ("Bodrum", "bod", 8),
            ("Kat", "kat", 8),
            ("Plan", "plan", 14),
            ("Hn", "yukseklik", 9),
            ("Yük. Sınıfı", "yukseklik_sinif", 12),
            ("Temel Alanı", "temel_alan", 11),
            ("İnşaat Alanı", "ins", 12),
            ("Kazı Der.", "der", 9),
            ("Temel Tipi", "tem", 14),
            ("GQE Min", "gqe_min", 9),
            ("GQE Maks", "gqe_max", 9),
            ("GQE Ort", "gqe_ort", 9),
            ("Comb Min", "comb_min", 9),
            ("Comb Maks", "comb_max", 9),
            ("Comb Ort", "comb_ort", 9),
        ]

    def bina_sonraki_blok_adi(self):
        idx = len(getattr(self, "bina_blok_rows", []))
        if idx < 26:
            return f"{chr(65 + idx)} Blok"
        return f"Blok {idx + 1}"

    def bina_genel_bilgilerinden_blok(self):
        values = {key: entry.get().strip() for key, entry in getattr(self, "e_bina", {}).items()}
        values["blok_adi"] = self.bina_sonraki_blok_adi()
        return values

    def bina_blok_satir_sec(self, row_idx):
        self.bina_blok_secili_idx = row_idx
        for idx, row in enumerate(getattr(self, "bina_blok_rows", [])):
            bg = "#D6EAF8" if idx == row_idx else COLOR_BG
            try:
                row["label"].configure(background=bg)
            except Exception:
                pass

    def bina_blok_scroll_guncelle(self):
        if hasattr(self, "bina_blok_canvas"):
            self.bina_blok_canvas.update_idletasks()
            self.bina_blok_canvas.configure(scrollregion=self.bina_blok_canvas.bbox("all"))

    def bina_blok_modu_guncelle(self):
        enabled = bool(getattr(self, "bina_coklu_blok_var", tk.BooleanVar(value=False)).get())
        state = "normal" if enabled else "disabled"
        for btn in getattr(self, "bina_blok_buttons", []):
            try:
                btn.configure(state=state)
            except Exception:
                pass
        for row in getattr(self, "bina_blok_rows", []):
            for entry in row.get("entries", {}).values():
                try:
                    entry.configure(state=state)
                except Exception:
                    pass
        if enabled and not getattr(self, "bina_blok_rows", []):
            self.bina_blok_satir_ekle(self.bina_genel_bilgilerinden_blok())

    def bina_blok_satir_ekle(self, values=None):
        if not hasattr(self, "bina_blok_frame"):
            return
        values = values or {"blok_adi": self.bina_sonraki_blok_adi()}
        row_idx = len(self.bina_blok_rows)
        grid_row = row_idx + 1
        row_label = ttk.Label(self.bina_blok_frame, text=str(row_idx + 1), width=4, anchor="center")
        row_label.grid(row=grid_row, column=0, padx=1, pady=2, sticky="nsew")
        entries = {}
        widgets = [row_label]
        for col_idx, (_, key, width) in enumerate(self.bina_blok_kolonlari(), start=1):
            entry = UndoRedoEntry(self.bina_blok_frame, width=width)
            entry.insert(0, str(values.get(key, "") or ""))
            entry.grid(row=grid_row, column=col_idx, padx=1, pady=2, sticky="nsew")
            entry.bind("<FocusIn>", lambda event, idx=row_idx: self.bina_blok_satir_sec(idx), add="+")
            entry.bind("<Button-1>", lambda event, idx=row_idx: self.bina_blok_satir_sec(idx), add="+")
            entries[key] = entry
            widgets.append(entry)
        self.bina_blok_rows.append({"entries": entries, "label": row_label, "widgets": widgets})
        self.bina_blok_satir_sec(row_idx)
        self.bina_blok_modu_guncelle()
        self.bina_blok_scroll_guncelle()

    def bina_bloklari_temizle(self):
        for row in getattr(self, "bina_blok_rows", []):
            for widget in row.get("widgets", []):
                try:
                    widget.destroy()
                except Exception:
                    pass
        self.bina_blok_rows = []
        self.bina_blok_secili_idx = None
        self.bina_blok_scroll_guncelle()

    def bina_bloklari_yukle(self, bloklar):
        self.bina_bloklari_temizle()
        for blok in bloklar or []:
            if isinstance(blok, dict):
                self.bina_blok_satir_ekle(blok)
        self.bina_blok_modu_guncelle()

    def bina_bloklari_topla(self):
        bloklar = []
        for idx, row in enumerate(getattr(self, "bina_blok_rows", [])):
            values = {key: entry.get().strip() for key, entry in row.get("entries", {}).items()}
            if not values.get("blok_adi"):
                values["blok_adi"] = f"Blok {idx + 1}"
            if any(str(value).strip() for value in values.values()):
                bloklar.append(values)
        return bloklar

    def bina_blok_secili_satir(self):
        rows = getattr(self, "bina_blok_rows", [])
        if not rows:
            return None, None
        idx = getattr(self, "bina_blok_secili_idx", None)
        if idx is None or idx < 0 or idx >= len(rows):
            idx = len(rows) - 1
        return idx, rows[idx]

    def bina_blok_ekle(self):
        self.bina_blok_satir_ekle({"blok_adi": self.bina_sonraki_blok_adi()})

    def bina_blok_genelden_ekle(self):
        self.bina_blok_satir_ekle(self.bina_genel_bilgilerinden_blok())

    def bina_blok_cogalt(self):
        idx, row = self.bina_blok_secili_satir()
        if row is None:
            self.bina_blok_ekle()
            return
        values = {key: entry.get().strip() for key, entry in row["entries"].items()}
        values["blok_adi"] = self.bina_sonraki_blok_adi()
        self.bina_blok_satir_ekle(values)

    def bina_blok_sil(self):
        idx, row = self.bina_blok_secili_satir()
        if row is None:
            messagebox.showinfo("Blok Sil", "Silinecek blok satırı yok.")
            return
        if not messagebox.askyesno("Blok Sil", "Seçili blok satırı silinsin mi?"):
            return
        bloklar = self.bina_bloklari_topla()
        if idx < len(bloklar):
            del bloklar[idx]
        self.bina_bloklari_yukle(bloklar)

    def p_bina(self, p):
        main_p = ttk.Frame(p); main_p.pack(fill="both", expand=True, padx=10, pady=10)
        self.e_bina = {}
        top_p = ttk.Frame(main_p); top_p.pack(fill="x")
        left_f = ttk.LabelFrame(top_p, text="Yapı Genel Bilgileri", padding="15"); left_f.pack(side="left", fill="both", expand=True, padx=(0,5))
        fields_l = [("Bina Kullanım Amacı", "kul"), ("Bina Kullanım Sınıfı", "sinif"), ("Bina Önem Katsayısı", "onem"), ("Yapı Malzemesi", "malz"), ("Bodrum Kat Adedi", "bod"), ("Toplam Kat Adedi", "kat"), ("Plan Boyutları", "plan"), ("Yapı Yüksekliği (Hn)", "yukseklik"), ("Bina Yükseklik Sınıfı", "yukseklik_sinif")]
        for i, (l, k) in enumerate(fields_l):
            ttk.Label(left_f, text=l).grid(row=i, column=0, sticky="e", padx=5, pady=5)
            e = UndoRedoEntry(left_f, width=25); e.grid(row=i, column=1, sticky="ew", padx=5, pady=5)
            self.e_bina[k] = e
        right_f = ttk.LabelFrame(top_p, text="Alanlar ve Yükler", padding="15"); right_f.pack(side="right", fill="both", expand=True, padx=(5,0))
        fields_r = [
            ("Temel Alanı (m2)", "temel_alan"),
            ("Toplam İnşaat Alanı (m2)", "ins"),
            ("Olası Kazı Derinliği (m)", "der"),
            ("Temel Tipi", "tem"),
            ("Yerel Zemin Sınıfı", "ysinif"),
        ]
        for i, (l, k) in enumerate(fields_r):
            ttk.Label(right_f, text=l).grid(row=i, column=0, sticky="e", padx=5, pady=5)
            e = UndoRedoEntry(right_f, width=25); e.grid(row=i, column=1, sticky="ew", padx=5, pady=5)
            self.e_bina[k] = e
        load_start_row = len(fields_r)
        ttk.Separator(right_f, orient='horizontal').grid(row=load_start_row, column=0, columnspan=2, sticky="ew", pady=15)
        ttk.Label(right_f, text="Binadan Temel Zeminine Aktarılan En Yükler (t/m2)", font=("Arial", 10, "bold")).grid(row=load_start_row + 1, column=0, columnspan=2, pady=(0,10))
        load_frame = ttk.Frame(right_f); load_frame.grid(row=load_start_row + 2, column=0, columnspan=2)
        ttk.Label(load_frame, text="Yük Tipi").grid(row=0, column=0, padx=5); ttk.Label(load_frame, text="Mim").grid(row=0, column=1, padx=5); ttk.Label(load_frame, text="Maks").grid(row=0, column=2, padx=5); ttk.Label(load_frame, text="Ort.").grid(row=0, column=3, padx=5)
        ttk.Label(load_frame, text="(G+Q+E)").grid(row=1, column=0, padx=5, pady=5)
        self.e_bina["gqe_min"] = UndoRedoEntry(load_frame, width=8); self.e_bina["gqe_min"].grid(row=1, column=1, padx=5)
        self.e_bina["gqe_max"] = UndoRedoEntry(load_frame, width=8); self.e_bina["gqe_max"].grid(row=1, column=2, padx=5)
        self.e_bina["gqe_ort"] = UndoRedoEntry(load_frame, width=8); self.e_bina["gqe_ort"].grid(row=1, column=3, padx=5)
        ttk.Label(load_frame, text="1.4G+1.6Q").grid(row=2, column=0, padx=5, pady=5)
        self.e_bina["comb_min"] = UndoRedoEntry(load_frame, width=8); self.e_bina["comb_min"].grid(row=2, column=1, padx=5)
        self.e_bina["comb_max"] = UndoRedoEntry(load_frame, width=8); self.e_bina["comb_max"].grid(row=2, column=2, padx=5)
        self.e_bina["comb_ort"] = UndoRedoEntry(load_frame, width=8); self.e_bina["comb_ort"].grid(row=2, column=3, padx=5)

        blok_f = ttk.LabelFrame(main_p, text="Çoklu Blok Bilgileri", padding="10")
        blok_f.pack(fill="both", expand=True, pady=(10, 0))
        self.bina_coklu_blok_var = tk.BooleanVar(value=False)
        self.bina_blok_rows = []
        self.bina_blok_secili_idx = None
        self.bina_blok_buttons = []

        toolbar = ttk.Frame(blok_f)
        toolbar.pack(fill="x", pady=(0, 6))
        ttk.Checkbutton(
            toolbar,
            text="Projede birden fazla blok var",
            variable=self.bina_coklu_blok_var,
            command=self.bina_blok_modu_guncelle,
        ).pack(side="left", padx=(0, 12))
        for text, command, color in [
            ("+ Blok", self.bina_blok_ekle, COLOR_ACCENT),
            ("Genelden A Blok", self.bina_blok_genelden_ekle, "#7DCEA0"),
            ("Çoğalt", self.bina_blok_cogalt, "#85C1E9"),
            ("Sil", self.bina_blok_sil, COLOR_DANGER),
        ]:
            btn = tk.Button(toolbar, text=text, command=command, bg=color, fg="white" if color in (COLOR_ACCENT, COLOR_DANGER) else "#111", font=FONT_BOLD)
            btn.pack(side="left", padx=3)
            self.bina_blok_buttons.append(btn)
        ttk.Label(toolbar, text="Kapalıysa raporda tek bina bilgileri kullanılır.").pack(side="left", padx=10)

        table_wrap = ttk.Frame(blok_f)
        table_wrap.pack(fill="both", expand=True)
        self.bina_blok_canvas = Canvas(table_wrap, bg=COLOR_BG, height=185, highlightthickness=0)
        y_scroll = ttk.Scrollbar(table_wrap, orient="vertical", command=self.bina_blok_canvas.yview)
        x_scroll = ttk.Scrollbar(table_wrap, orient="horizontal", command=self.bina_blok_canvas.xview)
        self.bina_blok_frame = ttk.Frame(self.bina_blok_canvas)
        self.bina_blok_canvas.create_window((0, 0), window=self.bina_blok_frame, anchor="nw")
        self.bina_blok_canvas.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        self.bina_blok_canvas.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        table_wrap.rowconfigure(0, weight=1)
        table_wrap.columnconfigure(0, weight=1)
        self.bina_blok_frame.bind("<Configure>", lambda e: self.bina_blok_scroll_guncelle())

        ttk.Label(self.bina_blok_frame, text="#", width=4, anchor="center", font=FONT_BOLD).grid(row=0, column=0, padx=1, pady=2, sticky="nsew")
        for col_idx, (label, _, width) in enumerate(self.bina_blok_kolonlari(), start=1):
            ttk.Label(self.bina_blok_frame, text=label, width=width, anchor="center", font=FONT_BOLD).grid(row=0, column=col_idx, padx=1, pady=2, sticky="nsew")
        self.bina_blok_modu_guncelle()

    def p_arazi(self, p):
        f = ttk.LabelFrame(p, text="Arazi Bilgileri", padding="20")
        f.pack(fill="both", expand=True, padx=20, pady=20)
        
        self.e_arazi = {}
        keys = ["yon","egim","min","max","ort","zemin","pga"]
        labels = ["Eğim Yönü", "Eğim Derecesi", "Min Kot", "Max Kot", "Ortalama Kot", "Zemin Grubu (ZA-ZF)", "PGA (g)"]
        row_idx = 0
        for l, k in zip(labels, keys):
            ttk.Label(f, text=l).grid(row=row_idx, column=0, sticky="e", padx=10, pady=8)
            e = UndoRedoEntry(f, width=40); e.grid(row=row_idx, column=1, sticky="w", padx=10, pady=8)
            self.e_arazi[k] = e
            row_idx += 1
            
        ttk.Label(f, text="Proje Alanı Merkezi Enlem (Y)").grid(row=row_idx, column=0, sticky="e", padx=10, pady=8)
        self.e_arazi["alan_y"] = UndoRedoEntry(f, width=40); self.e_arazi["alan_y"].grid(row=row_idx, column=1, sticky="w", padx=10, pady=8)
        row_idx += 1
        ttk.Label(f, text="Proje Alanı Merkezi Boylam (X)").grid(row=row_idx, column=0, sticky="e", padx=10, pady=8)
        self.e_arazi["alan_x"] = UndoRedoEntry(f, width=40); self.e_arazi["alan_x"].grid(row=row_idx, column=1, sticky="w", padx=10, pady=8)
        row_idx += 1

        ttk.Label(f, text="Zemin Etüt Kategorisi").grid(row=row_idx, column=0, sticky="e", padx=10, pady=8)
        self.e_arazi["kategori"] = ttk.Combobox(f, values=["Kategori 1", "Kategori 2", "Kategori 3"], width=38, state="readonly")
        self.e_arazi["kategori"].grid(row=row_idx, column=1, sticky="w", padx=10, pady=8)
        self.e_arazi["kategori"].set("Kategori 2")
        
        frame_imar = ttk.LabelFrame(p, text="İmar Durumu ve Plan Notları", padding=10)
        frame_imar.pack(fill="x", padx=20, pady=10)
        ttk.Label(frame_imar, text="İmar Alanı (Raporda parantez içinde yazılır):").grid(row=0, column=0, sticky="e", padx=5, pady=5)
        self.e_arazi["imar_alani"] = UndoRedoEntry(frame_imar, width=50)
        self.e_arazi["imar_alani"].grid(row=0, column=1, sticky="w", padx=5, pady=5)
        self.e_arazi["imar_alani"].insert(0, "Konut Alanı") 
        ttk.Label(frame_imar, text="İmar Durumu (Önlemli Alanlar):").grid(row=1, column=0, sticky="e", padx=5, pady=5)
        imar_secenekleri = ["Önlemli Alan 1.1 (ÖA-1.1) : Sıvılaşma Tehlikesi Açısından Önlemli Alanlar", "Önlemli Alan 2.1 (Ö.A-2.1) : Önlem Alınabilecek Nitelikte Stabilite Sorunlu Alanlar", "Önlemli Alan 2.2 (Ö.A-2.2) : Önlem Alınabilecek Nitelikte Kaya Düşmesi Sorunlu Alanlar", "Önlemli Alan 2.3 (Ö.A-2.3) : Önlem Alınabilecek Nitelikte Heyelan ve Kaya Düşmesi", "Önlemli Alan 5.1 (ÖA-5.1) : Önlem Alınabilecek Nitelikte Şişme, Oturma Açısından Sorunlu Alanlar"]
        self.e_arazi["imar_durumu"] = ttk.Combobox(frame_imar, values=imar_secenekleri, width=80, state="readonly")
        self.e_arazi["imar_durumu"].grid(row=1, column=1, sticky="w", padx=5, pady=5)
        self.e_arazi["imar_durumu"].current(0) 

    def guncelle_veri_objesi(self, silent=False):
        for k,e in self.e_kunye.items(): self.veri["kunye"][k]=e.get()
        for k,e in self.e_bina.items(): self.veri["bina"][k]=e.get()
        if hasattr(self, "bina_coklu_blok_var"):
            self.veri["bina"]["coklu_blok"] = bool(self.bina_coklu_blok_var.get())
            self.veri["bina"]["bloklar"] = self.bina_bloklari_topla()
        for k,e in self.e_arazi.items():
            if k == "imar_durumu" or k == "kategori": self.veri["arazi"][k] = e.get()
            else: self.veri["arazi"][k] = e.get()
        if "dosyalar" not in self.veri: self.veri["dosyalar"] = {}
        self.veri["dosyalar"]["kml_path"] = getattr(self, 'kml_path', None)
        self.veri["dosyalar"]["word_path"] = getattr(self, 'word_path', None)
        self.veri["dosyalar"]["lab_excel_path"] = getattr(self, 'lab_excel_path', None)
        self.veri["dosyalar"]["jeo_excel_path"] = getattr(self, 'jeo_excel_path', None)
        self.veri["dosyalar"]["img_yer"] = getattr(self, 'img_yer', None)
        self.veri["dosyalar"]["img_tkgm"] = getattr(self, 'img_tkgm', None)
        self.veri["dosyalar"]["img_pga"] = getattr(self, 'img_pga', None)
        self.veri["dosyalar"]["img_mjh"] = self._harita_gorsel_yolu_kaydet(getattr(self, 'img_mjh', None))
        self.veri["dosyalar"]["word_img_sondaj"] = self._harita_gorsel_yolu_kaydet(getattr(self, 'word_img_sondaj', None))
        self.veri["dosyalar"]["word_img_jeofizik"] = self._harita_gorsel_yolu_kaydet(getattr(self, 'word_img_jeofizik', None))
        if hasattr(self, "e_jeo_tar"):
            self.veri.setdefault("jeofizik", {})["tarih"] = self.e_jeo_tar.get().strip()
        self.veri.setdefault("ayarlar", {})["ek_tutanak_path"] = getattr(self, 'ek_tutanak_path', None) or self.veri.get("ayarlar", {}).get("ek_tutanak_path", "")
        self.veri.setdefault("ayarlar", {})["ek_arazi_deneyli_path"] = getattr(self, 'ek_arazi_deneyli_path', None) or self.veri.get("ayarlar", {}).get("ek_arazi_deneyli_path", "")
        self.sondaj_verilerini_kaydet(silent=silent)
    @perf_tracked("ui.fill")
    def doldur_arayuz(self):
        for k,e in self.e_kunye.items(): e.delete(0,'end'); e.insert(0, self.veri["kunye"].get(k,""))
        for k,e in self.e_bina.items(): e.delete(0,'end'); e.insert(0, self.veri["bina"].get(k,""))
        if hasattr(self, "bina_coklu_blok_var"):
            bina = self.veri.setdefault("bina", {})
            self.bina_coklu_blok_var.set(bool(bina.get("coklu_blok", False)))
            self.bina_bloklari_yukle(bina.get("bloklar", []))
        for k,e in self.e_arazi.items():
            if k == "imar_durumu" or k == "kategori": e.set(self.veri["arazi"].get(k,""))
            else: e.delete(0,'end'); e.insert(0, self.veri["arazi"].get(k,""))
            
        dosyalar = self.veri.get("dosyalar", {})
        self.kml_path = self._dosya_yolu_al(dosyalar, "kml_path", "kml", "kml_file", "kml_dosya", "kml_sinir_path", "kml_siniri_path", "kml_sınır_path")
        self.word_path = self._dosya_yolu_al(dosyalar, "word_path", "word", "word_file", "word_dosya")
        self.lab_excel_path = self._dosya_yolu_al(dosyalar, "lab_excel_path", "lab_excel", "laboratuvar_excel", "lab_dosya")
        self.jeo_excel_path = self._dosya_yolu_al(dosyalar, "jeo_excel_path", "jeofizik_excel", "jeo_excel")
        self.img_yer = self._dosya_yolu_al(dosyalar, "img_yer", "yerbuldurur_img", "yerbuldurur")
        self.img_tkgm = self._dosya_yolu_al(dosyalar, "img_tkgm", "tkgm_img", "tkgm")
        self.img_pga = self._dosya_yolu_al(dosyalar, "img_pga", "pga_img", "pga")
        self.img_mjh = self._harita_gorsel_yolu_al(dosyalar, "img_mjh", "mjh_img", "mjh")
        self.word_img_sondaj = self._harita_gorsel_yolu_al(dosyalar, "word_img_sondaj", "img_sondaj", "sondaj_haritasi")
        self.word_img_jeofizik = self._harita_gorsel_yolu_al(dosyalar, "word_img_jeofizik", "img_jeofizik", "jeofizik_haritasi")
        self.ek_tutanak_path = self.veri.get("ayarlar", {}).get("ek_tutanak_path")
        self.ek_arazi_deneyli_path = self.veri.get("ayarlar", {}).get("ek_arazi_deneyli_path")
        self.ayarlari_uygula()
        if hasattr(self, "e_jeo_tar"):
            self.e_jeo_tar.delete(0, "end")
            self.e_jeo_tar.insert(0, self.veri.get("jeofizik", {}).get("tarih", ""))
        
        self.kml_etiket_guncelle()
        if hasattr(self, 'lbl_sab'):
            self.lbl_sab.config(text=os.path.basename(self.word_path) if self.word_path else "Henüz seçilmedi...", foreground=COLOR_SUCCESS if self.word_path else "red")
        if hasattr(self, 'lbl_lab'):
            self.lbl_lab.config(text=os.path.basename(self.lab_excel_path) if self.lab_excel_path else "Henüz laboratuvar dosyası seçilmedi", foreground=COLOR_SUCCESS if self.lab_excel_path else "red")
        if hasattr(self, 'lbl_jeo_excel'):
            self.lbl_jeo_excel.config(text=os.path.basename(self.jeo_excel_path) if self.jeo_excel_path else "Henüz jeofizik dosyası seçilmedi", foreground=COLOR_SUCCESS if self.jeo_excel_path else "red")
        if hasattr(self, 'lbl_yer'):
            self.lbl_yer.config(text=os.path.basename(self.img_yer) if self.img_yer else "-", foreground=COLOR_SUCCESS if self.img_yer else "#333")
        if hasattr(self, 'lbl_tkgm'):
            self.lbl_tkgm.config(text=os.path.basename(self.img_tkgm) if self.img_tkgm else "-", foreground=COLOR_SUCCESS if self.img_tkgm else "#333")
        if hasattr(self, 'lbl_pga'):
            self.lbl_pga.config(text=os.path.basename(self.img_pga) if self.img_pga else "-", foreground=COLOR_SUCCESS if self.img_pga else "#333")
        if hasattr(self, 'lbl_mjh'):
            self.lbl_mjh.config(text=os.path.basename(self.img_mjh) if self.img_mjh else "-", foreground=COLOR_SUCCESS if self.img_mjh else "#333")
            
        self.sondaj_tablosunu_ciz(); self.jeo_yenile(); self.mt_yenile()
        self.ozet_yenile(collect=False)

if __name__ == "__main__":
    root = tk.Tk()
    app = RaporRobotuArayuz(root)
    root.mainloop()


