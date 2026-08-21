# Dosya: RaporPro/ui_lab_sheet.py
import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from excel_guvenligi import excel_satiri_guvenli_yap
from laboratuvar_motoru import laboratuvar_baslik_bilgisi
from performans import perf_tracked
from sabitler import COLOR_ACCENT, COLOR_DANGER, COLOR_SUCCESS, COLOR_WARNING, FONT_BOLD


LAB_SHEET_DEFAULT_ROWS = 80
LAB_SHEET_DEFAULT_COLS = 35


def lab_sheet_rows_temizle(rows):
    clean_rows = []
    for row in rows or []:
        cells = ["" if cell is None else str(cell).strip() for cell in (row or [])]
        while cells and cells[-1] == "":
            cells.pop()
        clean_rows.append(cells)
    while clean_rows and not any(str(cell).strip() for cell in clean_rows[-1]):
        clean_rows.pop()
    return clean_rows


def lab_sheet_grid_hazirla(rows, min_rows=LAB_SHEET_DEFAULT_ROWS, min_cols=LAB_SHEET_DEFAULT_COLS):
    clean_rows = lab_sheet_rows_temizle(rows)
    col_count = max(min_cols, max((len(row) for row in clean_rows), default=0))
    row_count = max(min_rows, len(clean_rows))
    grid = []
    for idx in range(row_count):
        source = clean_rows[idx] if idx < len(clean_rows) else []
        grid.append(source + [""] * max(0, col_count - len(source)))
    return grid


def lab_sheet_var_mi(veri):
    rows = (veri or {}).get("lab_sheet", {}).get("rows", [])
    return any(any(str(cell).strip() for cell in row) for row in rows or [])


def lab_excel_satirlari_oku(path):
    """Bağlı LAB Excel dosyasının etkin sayfasını satır listesi olarak oku."""
    source = os.fspath(path) if path else ""
    if not source or not os.path.isfile(source):
        return []
    extension = os.path.splitext(source)[1].lower()
    if extension in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(source, data_only=True, read_only=True)
        try:
            worksheet = workbook.active
            return [
                ["" if cell is None else cell for cell in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
        finally:
            workbook.close()

    import pandas as pd

    frame = pd.read_excel(source, header=None)
    return [
        ["" if pd.isna(cell) else cell for cell in row]
        for row in frame.itertuples(index=False, name=None)
    ]


def lab_kaynak_satirlari(veri, excel_path=None):
    """Raporla aynı öncelikle LAB Sheet'i, gerekirse bağlı Excel'i kullan."""
    rows = lab_sheet_rows_temizle(
        ((veri or {}).get("lab_sheet", {}) or {}).get("rows", [])
    )
    if any(any(str(cell).strip() for cell in row) for row in rows):
        return rows, "LAB Sheet"
    source = os.fspath(excel_path) if excel_path else ""
    if source and os.path.isfile(source):
        return lab_sheet_rows_temizle(lab_excel_satirlari_oku(source)), os.path.basename(source)
    return [], ""


class LabSheetMixin:
    def _lab_label_guncelle(self):
        if not hasattr(self, "lbl_lab"):
            return
        has_sheet = lab_sheet_var_mi(getattr(self, "veri", {}))
        path = getattr(self, "lab_excel_path", None)
        if has_sheet:
            row_count = len(lab_sheet_rows_temizle(self.veri.get("lab_sheet", {}).get("rows", [])))
            suffix = f" + {os.path.basename(path)}" if path else ""
            self.lbl_lab.config(text=f"LAB Sheet hazir ({row_count} satir){suffix}", foreground=COLOR_SUCCESS)
        elif path and os.path.isfile(path):
            self.lbl_lab.config(text=os.path.basename(path), foreground=COLOR_SUCCESS)
        elif path:
            self.lbl_lab.config(text=f"{os.path.basename(path)} (bulunamadı)", foreground=COLOR_WARNING)
        else:
            self.lbl_lab.config(text="Laboratuvar Excel seçilmedi", foreground="red")
        if hasattr(self, "rapor_durum_guncelle"):
            self.rapor_durum_guncelle()

    @perf_tracked("lab_sheet.open")
    def lab_sheet_ac(self):
        try:
            from tksheet import Sheet
        except Exception as exc:
            messagebox.showerror(
                "LAB Sheet",
                f"tksheet yuklenemedi:\n{exc}\n\nCozum: Bu Python icin `pip install tksheet` calistirin.",
            )
            return

        self.guncelle_veri_objesi(silent=True)
        win = tk.Toplevel(self.root)
        self.pencere_hazirla(win, "LAB Sheet - Laboratuvar Verisi", "1320x760", (980, 620), modal=False)

        top = ttk.Frame(win, padding=(10, 8))
        top.pack(fill="x")
        ttk.Label(top, text="LAB Sheet", font=("Segoe UI", 13, "bold")).pack(side="left", padx=(0, 10))
        info_var = tk.StringVar(value="Laboratuvar Excel tablonuzu buraya yapistirin.")
        ttk.Label(top, textvariable=info_var, foreground="#1F618D").pack(side="left", fill="x", expand=True)

        frame = ttk.Frame(win, padding=(10, 0, 10, 8))
        frame.pack(fill="both", expand=True)
        headers = [f"{idx + 1}" for idx in range(LAB_SHEET_DEFAULT_COLS)]
        sheet = Sheet(
            frame,
            headers=headers,
            data=lab_sheet_grid_hazirla(self.veri.get("lab_sheet", {}).get("rows", [])),
            show_row_index=True,
            show_header=True,
            theme="light blue",
            paste_can_expand_y=True,
            paste_can_expand_x=True,
            edit_cell_return="down",
            edit_cell_tab="right",
            default_column_width=92,
            default_row_index_width=52,
            column_drag_and_drop_perform=False,
            row_drag_and_drop_perform=True,
        )
        sheet.pack(fill="both", expand=True)
        sheet.enable_bindings("all")

        def rows_from_sheet():
            return lab_sheet_rows_temizle(sheet.get_sheet_data())

        def refresh_info():
            rows = rows_from_sheet()
            filled_rows = sum(1 for row in rows if any(str(cell).strip() for cell in row))
            filled_cols = max((len(row) for row in rows), default=0)
            info_var.set(f"{filled_rows} satir / {filled_cols} sutun kayda hazir.")

        def save_sheet(close=False):
            rows = rows_from_sheet()
            self.veri.setdefault("lab_sheet", {})["rows"] = rows
            self.ozet_yenile(collect=False)
            if hasattr(self, "lbl_lab"):
                self._lab_label_guncelle()
            if hasattr(self, "otomatik_kaydet"):
                self.otomatik_kaydet()
            self.set_status(f"LAB Sheet kaydedildi: {len(rows)} satir.", level="success")
            refresh_info()
            if close:
                self.pencere_kapat(win)

        def clear_sheet():
            if not messagebox.askyesno("LAB Sheet", "LAB Sheet verisi temizlensin mi?"):
                return
            sheet.set_sheet_data(lab_sheet_grid_hazirla([]), reset_col_positions=False, reset_row_positions=True)
            sheet.refresh()
            refresh_info()

        def add_rows(count=20):
            col_count = max(sheet.get_total_columns(), LAB_SHEET_DEFAULT_COLS)
            sheet.insert_rows([[""] * col_count for _ in range(count)], idx="end", undo=True)
            refresh_info()

        def import_excel():
            path = filedialog.askopenfilename(
                title="LAB Excel'den Al",
                filetypes=[("Excel", "*.xlsx;*.xlsm;*.xls")],
            )
            if not path:
                return
            try:
                rows = lab_excel_satirlari_oku(path)
            except Exception as exc:
                messagebox.showerror("LAB Sheet", f"Excel okunamadi:\n{exc}")
                return
            sheet.set_sheet_data(lab_sheet_grid_hazirla(rows), reset_col_positions=False, reset_row_positions=True)
            sheet.refresh()
            refresh_info()
            self.set_status(f"LAB Excel sheet'e alindi: {os.path.basename(path)}", level="success")

        def export_excel():
            try:
                from openpyxl import Workbook
            except Exception as exc:
                messagebox.showerror("LAB Sheet", f"openpyxl yuklenemedi:\n{exc}")
                return
            path = filedialog.asksaveasfilename(title="LAB Sheet Excel'e Aktar", defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
            if not path:
                return
            rows = rows_from_sheet()
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "LAB"
                for row in rows:
                    ws.append(excel_satiri_guvenli_yap(row))
                wb.save(path)
            except Exception as exc:
                messagebox.showerror("LAB Sheet", f"Excel kaydedilemedi:\n{exc}")
                return
            self.set_status(f"LAB Sheet Excel'e aktarildi: {os.path.basename(path)}", level="success")

        def show_context_menu(event=None):
            menu = tk.Menu(win, tearoff=0)
            menu.add_command(label="20 satir ekle", command=lambda: add_rows(20))
            menu.add_command(label="Temizle", command=clear_sheet)
            menu.add_separator()
            menu.add_command(label="Kaydet", command=lambda: save_sheet(False))
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()
            return "break"

        def safe_bind(sequence, func):
            try:
                sheet.bind(sequence, func)
            except tk.TclError:
                pass

        safe_bind("<Button-3>", show_context_menu)
        safe_bind("<Button-2>", show_context_menu)
        sheet.extra_bindings("all_modified_events", lambda event=None: refresh_info())

        buttons = ttk.Frame(win, padding=(10, 0, 10, 10))
        buttons.pack(fill="x")
        tk.Button(buttons, text="Excel'den Al", command=import_excel, bg="#D6EAF8", font=FONT_BOLD).pack(side="left", padx=(0, 4))
        tk.Button(buttons, text="Excel'e Aktar", command=export_excel, bg="#D5F5E3", font=FONT_BOLD).pack(side="left", padx=4)
        self.modern_button(buttons, text="+20 Satır", command=lambda: add_rows(20), role="primary", outline=True).pack(side="left", padx=4)
        self.modern_button(buttons, text="Temizle", command=clear_sheet, role="danger", outline=True).pack(side="left", padx=4)
        self.modern_button(buttons, text="Kaydet", command=lambda: save_sheet(False), role="success", outline=True).pack(side="right", padx=(4, 0))
        self.modern_button(buttons, text="Kaydet ve Kapat", command=lambda: save_sheet(True), role="success").pack(side="right", padx=4)

        refresh_info()
