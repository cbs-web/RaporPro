import datetime
import json
import os
from pathlib import Path
import re
import shutil
import time
import unicodedata
from dataclasses import dataclass, field

from excel_guvenligi import excel_satiri_guvenli_yap
from gizli_depo import gizli_deger_coz, gizli_deger_mi, gizli_deger_sakla
from performans import log_exception, perf_log
from spt_gorsel import dosya_parmak_izi, gorsel_api_payload_hazirla
from spt_saglayicilar import (
    http_post_with_retry as _http_post_with_retry,
    spt_ai_metin_iste,
)
from yardimcilar import atomic_json_dump, safe_float, temizle_baslik
from uygulama_yollari import SOURCE_DIR, kullanici_veri_dizini, kullanici_yolu


HEDEF_DERINLIKLER = [
    1.50, 3.00, 4.50, 6.00, 7.50, 9.00, 10.50, 12.00, 13.50, 15.00,
    16.50, 18.00, 19.50, 21.00, 22.50, 24.00, 25.50, 27.00, 28.50, 30.00,
]

HEDEF_DERINLIK_ARALIKLARI = [
    "1.50-1.95", "3.0-3.45", "4.5-4.95", "6.0-6.45", "7.5-7.95",
    "9.0-9.45", "10.5-10.95", "12.0-12.45", "13.5-13.95", "15.0-15.45",
    "16.5-16.95", "18.0-18.45", "19.5-19.95", "21.0-21.45", "22.5-22.95",
    "24.0-24.45", "25.5-25.95", "27.0-27.45", "28.5-28.95", "30.0-30.45",
]

SPT_OKUMA_KLASORU = Path.home() / "Desktop" / "SPT Okuma"
LEGACY_SPT_AYARLAR_PATH = SPT_OKUMA_KLASORU / "ayarlar.json"
RAPORPRO_CONFIG_DIR = kullanici_veri_dizini()
SPT_AYARLAR_PATH = Path(
    kullanici_yolu(
        "ayarlar.json",
        legacy=SOURCE_DIR / "ayarlar.json",
    )
)
SPT_LOG_DIR = Path(kullanici_yolu("logs"))
SPT_GECMIS_PATH = Path(
    kullanici_yolu(
        "logs",
        "spt_okuma_gecmisi.jsonl",
        legacy=SOURCE_DIR / "logs" / "spt_okuma_gecmisi.jsonl",
    )
)
SPT_OGRENME_DIR = Path(
    kullanici_yolu("spt_ogrenme_verisi", legacy=SOURCE_DIR / "spt_ogrenme_verisi")
)
SPT_CROP_DIR = Path(kullanici_yolu("logs", "spt_kirpilanlar"))
DEFAULT_SPT_GEMINI_MODEL = "gemini-3.6-flash"
DEFAULT_SPT_OPENAI_MODEL = "gpt-5.6-luna"
DEFAULT_SPT_PRO_OPENAI_MODEL = "gpt-5.6-terra"
DEFAULT_SPT_UST_OPENAI_MODEL = "gpt-5.6-sol"
DEFAULT_REVIZYON_OPENAI_MODEL = "gpt-5.5"
SECRET_SETTING_KEYS = ("openai_api_key", "gemini_api_key")
LEGACY_SECRET_SETTING_KEYS = SECRET_SETTING_KEYS + ("groq_api_key",)
SPT_OKUMA_MOTORLARI = ("gemini", "openai", "openai_pro", "openai_ust")


@dataclass
class SPTKaydi:
    sondaj_no: str = ""
    derinlik: str = ""
    v15: str = ""
    v30: str = ""
    v45: str = ""
    n30: str = ""
    guven: str = ""
    kaynak: str = ""
    kaynak_yolu: str = ""
    uyari: str = ""
    raw: dict = field(default_factory=dict)

    def spt_satiri(self):
        return [self.derinlik, self.v15, self.v30, self.v45, self.n30]

    def to_dict(self):
        return {
            "sondaj_no": self.sondaj_no,
            "derinlik": self.derinlik,
            "v15": self.v15,
            "v30": self.v30,
            "v45": self.v45,
            "n30": self.n30,
            "guven": self.guven,
            "kaynak": self.kaynak,
            "kaynak_yolu": self.kaynak_yolu,
            "uyari": self.uyari,
            "raw": dict(self.raw or {}),
        }


@dataclass
class SPTImportSonucu:
    kayitlar: list = field(default_factory=list)
    uyarilar: list = field(default_factory=list)


def normalize_header(cell):
    text = str(cell if cell is not None else "").strip().lower()
    text = text.replace("\u0131", "i").replace("\u0130", "i")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return temizle_baslik(text)


def temiz_metin(value):
    return "" if value is None else str(value).strip()


def _paylasilabilir_dosya_adi(value):
    text = str(value or "").strip().replace("\\", "/").rstrip("/")
    return text.rsplit("/", 1)[-1] if text else ""


def normalize_derinlik(value):
    text = temiz_metin(value).replace(",", ".")
    text = re.sub(r"^[^\d\-]+", "", text)
    if not text:
        return ""
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return ""
    val = safe_float(match.group(0))
    if val <= 0:
        return ""
    return f"{val:.2f}"


def hedef_derinlige_yuvarla(value, maksimum_sapma=0.76):
    der = safe_float(normalize_derinlik(value))
    if der <= 0:
        return ""
    nearest = min(HEDEF_DERINLIKLER, key=lambda item: abs(item - der))
    if abs(nearest - der) > maksimum_sapma:
        return ""
    return f"{nearest:.2f}"


def _token_temizle(token):
    text = temiz_metin(token).upper().replace(" ", "")
    text = text.replace("REFÜ", "R").replace("REFU", "R").replace("REF", "R")
    text = text.replace(",", ".")
    text = re.sub(r"50\s*/\s*", "50/", text)
    return text


def refu_mu(value):
    text = _token_temizle(value)
    return text in ("R", "REF", "REFU", "REFÜ") or "50/" in text


def sayi_token(value):
    text = _token_temizle(value)
    if refu_mu(text):
        return text if "50/" in text else "R"
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return ""
    number = safe_float(match.group(0))
    return str(int(number)) if float(number).is_integer() else str(number)


def spt_degerlerini_ayikla(value):
    text = temiz_metin(value)
    if not text:
        return "", "", ""
    upper = text.upper().replace(",", ".")
    upper = upper.replace("REFÜ", "R").replace("REFU", "R")
    tokens = re.findall(r"50\s*/\s*\d+|R|\d+(?:\.\d+)?", upper)
    values = []
    for token in tokens:
        clean = sayi_token(token)
        if clean:
            values.append(clean)
        if len(values) >= 3:
            break
    while len(values) < 3:
        values.append("")
    return values[0], values[1], values[2]


def n30_hesapla(v30="", v45="", mevcut=""):
    mevcut = temiz_metin(mevcut)
    if mevcut:
        return "R" if refu_mu(mevcut) else mevcut
    if refu_mu(v30) or refu_mu(v45):
        return "R"
    n30 = safe_float(v30) + safe_float(v45)
    if n30 <= 0:
        return ""
    return str(int(n30)) if float(n30).is_integer() else str(n30)


def normalize_sondaj_no(value, default=""):
    text = temiz_metin(value).upper()
    if not text:
        text = temiz_metin(default).upper()
    if not text:
        return ""
    text = text.replace("İ", "I")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.replace(" ", "").replace(".", "").replace("_", "-").replace("=", "-").replace(":", "-")
    text = re.sub(r"-+", "-", text)
    match = re.search(r"\b(S[CK]|SK|BH|K|KUYU)-?0*(\d+[A-Z]?)\b", text)
    if match:
        prefix = match.group(1)
        if prefix in ("SC", "SCK"):
            prefix = "SK"
        if prefix in ("K", "KUYU"):
            prefix = "SK"
        return f"{prefix}-{match.group(2)}"
    match = re.search(r"\b0*(\d{1,3}[A-Z]?)\b", text)
    if match and temiz_metin(default).upper().startswith("SK"):
        return f"SK-{match.group(1)}"
    return temiz_metin(default) if default else text


def kayit_normalize_et(values, default_sondaj_no=""):
    raw = dict(values or {})
    if raw.get("_motor") and not raw.get("motor"):
        raw["motor"] = raw.get("_motor")
    if raw.get("_model") and not raw.get("model"):
        raw["model"] = raw.get("_model")
    if raw.get("_okuma_suresi") is not None and raw.get("okuma_suresi") is None:
        raw["okuma_suresi"] = raw.get("_okuma_suresi")
    if isinstance(raw.get("_gorsel"), dict):
        raw.setdefault("gorsel", dict(raw["_gorsel"]))
        raw.setdefault("kaynak_hash", raw["_gorsel"].get("kaynak_hash", ""))
    sondaj_no = normalize_sondaj_no(
        raw.get("sondaj_no") or raw.get("sondaj") or raw.get("sondaj_adi") or raw.get("kuyu_no"),
        default_sondaj_no,
    )
    raw_derinlik = raw.get("derinlik") or raw.get("derinlik_araligi") or raw.get("der") or raw.get("metraj") or raw.get("depth")
    okunan_derinlik = normalize_derinlik(raw_derinlik)
    derinlik = hedef_derinlige_yuvarla(raw_derinlik)
    if okunan_derinlik and derinlik:
        raw["okunan_derinlik"] = okunan_derinlik
        raw["hedef_derinlik"] = derinlik
        fark = abs(safe_float(okunan_derinlik) - safe_float(derinlik))
        raw["derinlik_duzeltme_m"] = round(fark, 3)
    v15 = sayi_token(raw.get("v15") or raw.get("15") or raw.get("n15"))
    v30 = sayi_token(raw.get("v30") or raw.get("30") or raw.get("n30_vurus"))
    v45 = sayi_token(raw.get("v45") or raw.get("45") or raw.get("n45"))
    n30 = temiz_metin(raw.get("n30"))

    spt_text = raw.get("spt") or raw.get("spt_degeri") or raw.get("vurus") or raw.get("vurus_sayisi") or raw.get("darbe") or raw.get("darbeler")
    if spt_text and not (v15 or v30 or v45):
        v15, v30, v45 = spt_degerlerini_ayikla(spt_text)

    n30 = n30_hesapla(v30, v45, n30)
    if not n30 and (refu_mu(v15) or refu_mu(spt_text)):
        n30 = "R"
    uyari = []
    if not sondaj_no:
        uyari.append("Sondaj no eksik")
    if not derinlik:
        uyari.append("Derinlik okunamadı")
    if not (v15 or v30 or v45 or n30):
        uyari.append("SPT değeri okunamadı")
    if n30 == "R":
        raw["bilgi"] = ", ".join(filter(None, [temiz_metin(raw.get("bilgi")), "Refü"]))
    if (
        okunan_derinlik
        and derinlik
        and safe_float(raw.get("derinlik_duzeltme_m")) > 0.35
    ):
        uyari.append(f"Derinlik {okunan_derinlik} -> {derinlik} olarak düzeltildi")

    return SPTKaydi(
        sondaj_no=sondaj_no,
        derinlik=derinlik,
        v15=v15,
        v30=v30,
        v45=v45,
        n30=n30,
        guven=temiz_metin(raw.get("guven") or raw.get("confidence")),
        kaynak=temiz_metin(raw.get("kaynak") or raw.get("foto") or raw.get("resim")),
        kaynak_yolu=temiz_metin(raw.get("kaynak_yolu") or raw.get("path") or raw.get("dosya_yolu")),
        uyari=", ".join(uyari),
        raw=raw,
    )


def _header_map(cells):
    aliases = {
        "sondajno": "sondaj_no", "sondaj": "sondaj_no", "sk": "sondaj_no",
        "kuyu": "sondaj_no", "kuyuno": "sondaj_no", "sondajadi": "sondaj_no",
        "derinlik": "derinlik", "der": "derinlik", "metraj": "derinlik",
        "depth": "derinlik", "derinlikm": "derinlik",
        "spt": "spt", "sptdegeri": "spt", "sptvurus": "spt", "vurus": "spt",
        "15": "v15", "n15": "v15", "spt15": "v15",
        "30": "v30", "n30vurus": "v30", "spt30": "v30",
        "45": "v45", "n45": "v45", "spt45": "v45",
        "n30": "n30", "n": "n30",
        "guven": "guven", "confidence": "guven", "oran": "guven",
        "kaynak": "kaynak", "foto": "kaynak", "fotograf": "kaynak", "resim": "kaynak",
        "sondajadi": "sondaj_no", "sondajadiadi": "sondaj_no", "kuyuno": "sondaj_no",
    }
    mapped = [aliases.get(normalize_header(cell)) for cell in cells]
    return mapped if sum(1 for item in mapped if item) >= 2 else None


def _rows_to_records(raw_rows, default_sondaj_no="", kaynak=""):
    sonuc = SPTImportSonucu()
    if not raw_rows:
        return sonuc

    mapping = _header_map(raw_rows[0])
    data_rows = raw_rows[1:] if mapping else raw_rows
    for idx, row in enumerate(data_rows, start=2 if mapping else 1):
        values = {}
        if mapping:
            for col_idx, value in enumerate(row):
                if col_idx < len(mapping) and mapping[col_idx]:
                    values[mapping[col_idx]] = temiz_metin(value)
        else:
            cells = [temiz_metin(cell) for cell in row]
            if len(cells) >= 6:
                values = {
                    "sondaj_no": cells[0], "derinlik": cells[1],
                    "v15": cells[2], "v30": cells[3], "v45": cells[4], "n30": cells[5],
                }
            elif len(cells) >= 5:
                values = {
                    "derinlik": cells[0],
                    "v15": cells[1], "v30": cells[2], "v45": cells[3], "n30": cells[4],
                }
            elif len(cells) >= 2:
                values = {"derinlik": cells[0], "spt": cells[1]}
                if len(cells) >= 3:
                    values["guven"] = cells[2]

        if kaynak and not values.get("kaynak"):
            values["kaynak"] = kaynak
        kayit = kayit_normalize_et(values, default_sondaj_no)
        if not (kayit.derinlik or kayit.v15 or kayit.v30 or kayit.v45 or kayit.n30):
            continue
        if kayit.uyari:
            sonuc.uyarilar.append(f"Satır {idx}: {kayit.uyari}")
        sonuc.kayitlar.append(kayit)

    sonuc.kayitlar.sort(key=lambda item: (item.sondaj_no, safe_float(item.derinlik), item.kaynak))
    return sonuc


def excelden_spt_oku(path, default_sondaj_no=""):
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError(f"openpyxl yüklenemedi: {exc}") from exc

    # Dosya yalnızca okunuyor; read_only büyük saha tablolarını hücre nesneleri
    # olarak belleğe almadan satır satır işler.
    wb = load_workbook(path, data_only=True, read_only=True)
    genel_sonuc = SPTImportSonucu()
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            raw_rows = []
            for row in ws.iter_rows(values_only=True):
                cells = ["" if cell is None else str(cell) for cell in row]
                if any(cell.strip() for cell in cells):
                    raw_rows.append(cells)
            if not raw_rows:
                continue
            sheet_result = _rows_to_records(raw_rows, default_sondaj_no, kaynak=sheet_name)
            for kayit in sheet_result.kayitlar:
                kayit.kaynak_yolu = str(path)
                kayit.raw["kaynak_yolu"] = str(path)
                kayit.kaynak = f"{Path(path).name} / {kayit.kaynak}" if kayit.kaynak else Path(path).name
            genel_sonuc.kayitlar.extend(sheet_result.kayitlar)
            genel_sonuc.uyarilar.extend([f"{sheet_name}: {msg}" for msg in sheet_result.uyarilar])
    finally:
        wb.close()

    genel_sonuc.kayitlar.sort(key=lambda item: (item.sondaj_no, safe_float(item.derinlik), item.kaynak))
    return genel_sonuc


def spt_ayarlarini_yukle(path=None):
    ayarlar = {
        "aktif_motor": "gemini",
        "openai_api_key": "",
        "openai_model": DEFAULT_SPT_OPENAI_MODEL,
        "spt_pro_openai_model": DEFAULT_SPT_PRO_OPENAI_MODEL,
        "spt_ust_openai_model": DEFAULT_SPT_UST_OPENAI_MODEL,
        "revizyon_openai_model": DEFAULT_REVIZYON_OPENAI_MODEL,
        "gemini_api_key": "",
        "spt_gemini_model": DEFAULT_SPT_GEMINI_MODEL,
    }
    varsayilan_yol = path is None
    path = Path(path) if path else SPT_AYARLAR_PATH
    if varsayilan_yol and not path.exists() and LEGACY_SPT_AYARLAR_PATH.exists():
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(LEGACY_SPT_AYARLAR_PATH, path)
        except Exception:
            path = LEGACY_SPT_AYARLAR_PATH
    if path.exists():
        try:
            with path.open("r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                secured_data = dict(data)
                migrated = False
                for key in LEGACY_SECRET_SETTING_KEYS:
                    raw_value = str(data.get(key, "") or "")
                    if not raw_value:
                        continue
                    try:
                        data[key] = gizli_deger_coz(raw_value)
                        if not gizli_deger_mi(raw_value):
                            secured_data[key] = gizli_deger_sakla(raw_value)
                            migrated = True
                    except Exception as exc:
                        data[key] = ""
                        log_exception(f"spt.settings.decrypt.{key}", exc_value=exc)
                if "groq_api_key" in secured_data:
                    secured_data.pop("groq_api_key", None)
                    data.pop("groq_api_key", None)
                    migrated = True
                ayarlar.update({key: str(value) for key, value in data.items() if value is not None})
                normalized_motor = spt_motorunu_normalize_et(ayarlar.get("aktif_motor"))
                if normalized_motor not in ("gemini", "openai"):
                    normalized_motor = "openai" if normalized_motor.startswith("openai") else "gemini"
                if normalized_motor != ayarlar.get("aktif_motor"):
                    ayarlar["aktif_motor"] = normalized_motor
                    secured_data["aktif_motor"] = normalized_motor
                    migrated = True
                if migrated:
                    atomic_json_dump(secured_data, path, ensure_ascii=False, indent=2)
        except Exception:
            pass

    env_map = {
        "openai_api_key": ("RAPORPRO_SPT_OPENAI_API_KEY", "OPENAI_API_KEY"),
        "openai_model": ("RAPORPRO_SPT_OPENAI_MODEL", "RAPORPRO_OPENAI_MODEL", "OPENAI_MODEL"),
        "spt_pro_openai_model": ("RAPORPRO_SPT_PRO_OPENAI_MODEL",),
        "spt_ust_openai_model": ("RAPORPRO_SPT_UST_OPENAI_MODEL",),
        "revizyon_openai_model": ("RAPORPRO_REVIZYON_OPENAI_MODEL", "RAPORPRO_RAPOR_OPENAI_MODEL"),
        "gemini_api_key": ("RAPORPRO_SPT_GEMINI_API_KEY", "GEMINI_API_KEY"),
        "spt_gemini_model": ("RAPORPRO_SPT_GEMINI_MODEL",),
        "aktif_motor": ("RAPORPRO_SPT_MOTOR",),
    }
    for key, names in env_map.items():
        for name in names:
            val = os.environ.get(name)
            if val:
                ayarlar[key] = val.strip()
                break
    aktif_motor = spt_motorunu_normalize_et(ayarlar.get("aktif_motor"))
    if aktif_motor not in ("gemini", "openai"):
        aktif_motor = "openai" if aktif_motor.startswith("openai") else "gemini"
    if aktif_motor == "gemini" and not ayarlar.get("gemini_api_key") and ayarlar.get("openai_api_key"):
        aktif_motor = "openai"
    elif aktif_motor == "openai" and not ayarlar.get("openai_api_key") and ayarlar.get("gemini_api_key"):
        aktif_motor = "gemini"
    ayarlar["aktif_motor"] = aktif_motor
    return ayarlar


def spt_motorunu_normalize_et(value):
    """Eski SPT motor adlarini guncel rol tabanli motorlara donustur."""
    motor = str(value or "gemini").strip().lower()
    aliases = {
        "groq": "gemini",
        "gemini_pro": "openai_pro",
        "openai_luna": "openai",
        "openai_terra": "openai_pro",
        "openai_sol": "openai_ust",
    }
    return aliases.get(motor, motor)


def openai_model_sec(ayarlar=None, amac="spt"):
    """SPT rolleri ve rapor revizyonu icin OpenAI modelini ayarlardan sec."""
    ayarlar = ayarlar or spt_ayarlarini_yukle()
    amac = str(amac or "").strip().lower()
    if amac in ("revizyon", "rapor", "duzeltme", "metin"):
        model = str(ayarlar.get("revizyon_openai_model") or "").strip()
        if model:
            return model
        model = str(ayarlar.get("openai_model") or "").strip()
        return model or DEFAULT_REVIZYON_OPENAI_MODEL
    if amac in ("spt_pro", "pro", "terra"):
        model = str(ayarlar.get("spt_pro_openai_model") or "").strip()
        return model or DEFAULT_SPT_PRO_OPENAI_MODEL
    if amac in ("spt_ust", "ust", "zor", "sol"):
        model = str(ayarlar.get("spt_ust_openai_model") or "").strip()
        return model or DEFAULT_SPT_UST_OPENAI_MODEL
    model = str(ayarlar.get("openai_model") or "").strip()
    return model or DEFAULT_SPT_OPENAI_MODEL


def gemini_model_sec(ayarlar=None):
    """Normal SPT okumasinda kullanilacak Gemini modelini sec."""
    ayarlar = ayarlar or spt_ayarlarini_yukle()
    model = str(ayarlar.get("spt_gemini_model") or "").strip()
    return model or DEFAULT_SPT_GEMINI_MODEL


def spt_ayarlarini_kaydet(ayarlar, path=None):
    path = Path(path) if path else SPT_AYARLAR_PATH
    path.parent.mkdir(parents=True, exist_ok=True)
    mevcut = {}
    if path.exists():
        try:
            with path.open("r", encoding="utf-8") as f:
                loaded = json.load(f)
            if isinstance(loaded, dict):
                mevcut.update(loaded)
        except Exception:
            mevcut = {}
    mevcut.pop("groq_api_key", None)
    for key in (
        "aktif_motor",
        "openai_api_key",
        "openai_model",
        "spt_pro_openai_model",
        "spt_ust_openai_model",
        "revizyon_openai_model",
        "gemini_api_key",
        "spt_gemini_model",
    ):
        if key in ayarlar:
            value = ayarlar.get(key, "")
            if key == "aktif_motor":
                value = spt_motorunu_normalize_et(value)
                if value not in ("gemini", "openai"):
                    value = "openai" if value.startswith("openai") else "gemini"
            mevcut[key] = gizli_deger_sakla(value) if key in SECRET_SETTING_KEYS and value else value
    atomic_json_dump(mevcut, path, ensure_ascii=False, indent=2)
    return path


def _spt_prompt():
    hedefler = ", ".join(HEDEF_DERINLIK_ARALIKLARI)
    return f"""Sen bir geoteknik mühendisisin. Gönderilen fotoğraftaki zemin etüt tabelasından sadece fotoğrafta AÇIKÇA yazan SPT satırını oku.
Her fotoğraf normalde tek bir SPT deneyini gösterir. Bu nedenle en fazla 1 JSON nesnesi döndür.
Fotoğrafta birden fazla SPT satırı tamamen net görünmüyorsa tahmin yapma; en belirgin/merkezdeki satırı döndür.
Fotoğrafta sondaj/kuyu numarası görünüyorsa sondaj_no alanında döndür. SK1, SK 1, S.K.-1 gibi yazımları gördüğünde sondaj_no değerini SK-1 biçiminde ver.
Beyaz tahta/saha etiketi formatlarını tanı: "SK=4", "SK:4", "S.K. 4" sondaj_no=SK-4 demektir.
Beyaz tahtada "D=10.50-10.95" veya "D:10.50-10.95" derinlik aralığıdır.
Derinlik satırının hemen altındaki "D=8-9-10", "D:8-9-10", "8 - 9 - 10" gibi üçlü değerler SPT vuruşlarıdır.
"Abl", "kot", "Parsel/Porsal", "PMT", "P", tarih, proje adı ve zemin notları SPT değildir; bunları yok say.
Dikkat: Geçerli SPT derinlikleri sadece şunlardır: {hedefler}
Derinliği farklı okuyorsan en yakın geçerli derinliği yaz. Örneğin 4.70 okursan 4.50, 14.90 okursan 15.00 yaz.
Listeden eksik satır üretme, bu listenin tamamını asla yazma.
SPT değerini mümkünse 4-5-6 biçiminde ver. Refü varsa spt alanına "R" veya "50/.." biçiminde yaz.
Derinlik veya SPT vuruşları fotoğrafta okunmuyorsa {{"items": []}} döndür.
Kendi okuma güvenini 0 ile 100 arasında guven alanına yaz.
Sadece {{"items": [...]}} biçiminde ham JSON nesnesi döndür. Markdown, açıklama veya kod bloğu yazma.
Örnekler:
{{"items": [{{"sondaj_no": "SK-1", "derinlik": "1.50-1.95", "spt": "4-4-5", "guven": 95}}]}}
{{"items": [{{"sondaj_no": "SK-4", "derinlik": "10.50-10.95", "spt": "8-9-10", "guven": 95}}]}}
Bulamazsan {{"items": []}} döndür."""


def _image_payload(path):
    image_b64, mime_type, _metadata = gorsel_api_payload_hazirla(path)
    return image_b64, mime_type


def _image_payload_with_meta(path):
    return gorsel_api_payload_hazirla(path)


def _json_liste_ayikla(text):
    cleaned = str(text or "").strip()
    cleaned = cleaned.replace("```json", "").replace("```", "").strip()
    try:
        parsed = json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("[")
        end = cleaned.rfind("]")
        if start >= 0 and end > start:
            parsed = json.loads(cleaned[start:end + 1])
        else:
            raise
    if isinstance(parsed, dict):
        items = parsed.get("items")
        if isinstance(items, list):
            parsed = items
        elif any(key in parsed for key in ("sondaj_no", "sondaj", "derinlik", "spt")):
            parsed = [parsed]
        else:
            parsed = []
    if not isinstance(parsed, list):
        return []
    return [item for item in parsed if isinstance(item, dict)]


def _path_unique_key(path):
    try:
        return os.path.normcase(os.path.realpath(os.path.abspath(str(path))))
    except Exception:
        return os.path.normcase(os.path.abspath(str(path)))


def _sondaj_no_from_path(path):
    try:
        parts = Path(str(path)).parts
    except Exception:
        parts = []
    for part in reversed(parts):
        text = temiz_metin(part).upper()
        text = text.replace("İ", "I").replace("S.K", "SK").replace("S K", "SK")
        text = unicodedata.normalize("NFKD", text)
        text = "".join(ch for ch in text if not unicodedata.combining(ch))
        match = re.search(r"(?<![A-Z0-9])(SCK|SC|SK|BH|KUYU|K)[\s._:=/-]*0*(\d{1,3}[A-Z]?)(?![A-Z0-9])", text)
        if match:
            prefix = match.group(1)
            if prefix in ("SC", "SCK", "K", "KUYU"):
                prefix = "SK"
            return f"{prefix}-{match.group(2)}"
    return ""


def _spt_location_key(kayit):
    return (
        normalize_sondaj_no(getattr(kayit, "sondaj_no", "")),
        round(safe_float(getattr(kayit, "derinlik", "")), 2),
    )


def _spt_record_score(kayit):
    score = 0.0
    for field_name in ("v15", "v30", "v45", "n30"):
        if str(getattr(kayit, field_name, "") or "").strip():
            score += 1.0
    score += min(max(safe_float(getattr(kayit, "guven", "")), 0), 100) / 100.0
    if getattr(kayit, "uyari", ""):
        score -= 0.25
    return score


def spt_kayit_puani(kayit):
    """Arayuz ve motorun ayni aday karsilastirma puanini kullanmasini sagla."""
    return _spt_record_score(kayit)


def _candidate_snapshot(kayit):
    return {
        "sondaj_no": kayit.sondaj_no,
        "derinlik": kayit.derinlik,
        "v15": kayit.v15,
        "v30": kayit.v30,
        "v45": kayit.v45,
        "n30": kayit.n30,
        "guven": kayit.guven,
        "kaynak": kayit.kaynak,
        "kaynak_yolu": kayit.kaynak_yolu,
        "uyari": kayit.uyari,
        "raw": dict(kayit.raw or {}),
    }


def _single_photo_candidate(records, expected_depth=None):
    if len(records) <= 1:
        return records[0] if records else None
    if expected_depth is not None:
        chosen = _select_best_for_expected_depth(records, expected_depth)
        reason = f"beklenen derinlik {expected_depth:.2f} m"
    else:
        chosen = max(records, key=_spt_record_score)
        reason = "veri butunlugu ve guven puani"
    alternatives = [_candidate_snapshot(item) for item in records if item is not chosen]
    chosen.raw["alternatif_okumalar"] = alternatives
    chosen.raw["aday_secim_nedeni"] = reason
    chosen.raw["aday_sayisi"] = len(records)
    warning = f"Fotoğrafta {len(records)} SPT adayı bulundu; bir sonuç seçildi"
    chosen.uyari = ", ".join(filter(None, [chosen.uyari, warning]))
    return chosen


def _infer_expected_depth_offset(records_by_path, path_order):
    max_offset = max(1, len(HEDEF_DERINLIKLER) - len(path_order) + 1)
    best_offset = 0
    best_hits = 0
    best_score = float("-inf")
    best_distance = float("inf")
    record_map = {key: records for key, records in records_by_path}

    for offset in range(max_offset):
        score = 0.0
        hits = 0
        distance_total = 0.0
        for idx, key in enumerate(path_order):
            depth_idx = offset + idx
            if depth_idx >= len(HEDEF_DERINLIKLER):
                continue
            records = record_map.get(key) or []
            depths = [safe_float(record.derinlik) for record in records if safe_float(record.derinlik) > 0]
            if not depths:
                continue
            expected = HEDEF_DERINLIKLER[depth_idx]
            best = min(abs(depth - expected) for depth in depths)
            distance_total += best
            if best <= 0.08:
                score += 3.0
                hits += 1
            elif best <= 0.30:
                score += 1.0
                hits += 1
            else:
                score -= min(best, 3.0)
        if (score, hits, -distance_total) > (best_score, best_hits, -best_distance):
            best_offset = offset
            best_hits = hits
            best_score = score
            best_distance = distance_total

    return best_offset, best_hits, best_score


def _select_best_for_expected_depth(records, expected_depth):
    def rank(record):
        distance = abs(safe_float(record.derinlik) - expected_depth)
        return (distance, -_spt_record_score(record))

    return min(records, key=rank)


def _select_spt_records_for_batch(records_by_path, paths, path_order=None):
    if path_order is None:
        path_order = [_path_unique_key(path) for path in paths]
    else:
        path_order = list(path_order)
    path_index = {}
    for index, key in enumerate(path_order):
        path_index.setdefault(key, index)
    has_multirow_photo = any(len(records) > 1 for _, records in records_by_path)
    selected = []
    removed_by_sequence = 0

    if len(path_order) > 1 and has_multirow_photo:
        offset, hits, _ = _infer_expected_depth_offset(records_by_path, path_order)
        min_hits = max(2, min(len(path_order), len(HEDEF_DERINLIKLER)) // 2)
        use_sequence = hits >= min_hits
    else:
        offset = 0
        use_sequence = False

    for key, records in records_by_path:
        if not records:
            continue
        chosen = None
        if use_sequence and len(records) > 1:
            order_index = path_index.get(key)
            expected_idx = offset + order_index if order_index is not None else -1
            if 0 <= expected_idx < len(HEDEF_DERINLIKLER):
                chosen = _single_photo_candidate(records, HEDEF_DERINLIKLER[expected_idx])
        if chosen is None:
            chosen = _single_photo_candidate(records)
        if chosen is not None:
            selected.append(chosen)
            removed_by_sequence += max(0, len(records) - 1)

    deduped = []
    locations = {}
    merged_by_location = 0
    for record in selected:
        loc_key = _spt_location_key(record)
        if loc_key[0] and loc_key[1] > 0:
            existing_idx = locations.get(loc_key)
            if existing_idx is not None:
                existing = deduped[existing_idx]
                if _spt_record_score(record) > _spt_record_score(existing):
                    deduped[existing_idx] = record
                merged_by_location += 1
                continue
            locations[loc_key] = len(deduped)
        deduped.append(record)

    return deduped, removed_by_sequence, merged_by_location


def _api_key_kontrol(aktif, ayarlar):
    if aktif in ("openai", "openai_pro", "openai_ust") and not ayarlar.get("openai_api_key"):
        raise RuntimeError("OpenAI API anahtarı bulunamadı. RaporPro ayarları veya OPENAI_API_KEY kontrol edilmeli.")
    if aktif == "gemini" and not ayarlar.get("gemini_api_key"):
        raise RuntimeError("Gemini API anahtarı bulunamadı. RaporPro ayarları veya GEMINI_API_KEY kontrol edilmeli.")


def yapay_zeka_ile_spt_oku(
    resim_yolu,
    ayarlar=None,
    motor_zorla=None,
    timeout=45,
    stop_event=None,
):
    ayarlar = ayarlar or spt_ayarlarini_yukle()
    aktif = spt_motorunu_normalize_et(motor_zorla or ayarlar.get("aktif_motor") or "gemini")
    if aktif not in SPT_OKUMA_MOTORLARI:
        raise RuntimeError(f"Desteklenmeyen SPT okuma motoru: {aktif}")
    _api_key_kontrol(aktif, ayarlar)

    started_at = time.perf_counter()
    image_b64, mime_type, image_meta = _image_payload_with_meta(resim_yolu)
    prompt = _spt_prompt()
    text_response, model_name = spt_ai_metin_iste(
        aktif=aktif,
        ayarlar=ayarlar,
        prompt=prompt,
        image_b64=image_b64,
        mime_type=mime_type,
        timeout=timeout,
        stop_event=stop_event,
        openai_model=openai_model_sec(
            ayarlar,
            "spt_pro" if aktif == "openai_pro" else "spt_ust" if aktif == "openai_ust" else "spt",
        ),
        gemini_model=gemini_model_sec(ayarlar),
    )
    raw_items = _json_liste_ayikla(text_response)

    elapsed = time.perf_counter() - started_at
    perf_log(
        "spt.ai_read",
        elapsed,
        f"motor={aktif};model={model_name};bytes={image_meta.get('islenmis_bayt', 0)};rows={len(raw_items)}",
    )
    for item in raw_items:
        item.setdefault("_motor", aktif)
        item.setdefault("_model", model_name)
        item.setdefault("_okuma_suresi", round(elapsed, 3))
        item.setdefault("_gorsel", dict(image_meta))
        item.setdefault("kaynak_hash", image_meta.get("kaynak_hash", ""))
    return raw_items


def _raw_items_score(raw_items, default_sondaj_no=""):
    records = [kayit_normalize_et(item, default_sondaj_no) for item in (raw_items or [])]
    valid = [
        item for item in records
        if item.derinlik and (item.v15 or item.v30 or item.v45 or item.n30)
    ]
    if not valid:
        return float("-inf")
    best = max(_spt_record_score(item) for item in valid)
    return best - max(0, len(valid) - 1) * 0.15


def fotograflardan_spt_oku(
    paths,
    default_sondaj_no="",
    ayarlar=None,
    progress_callback=None,
    stop_event=None,
    auto_pro=True,
    guven_esigi=90,
):
    ayarlar = ayarlar or spt_ayarlarini_yukle()
    sonuc = SPTImportSonucu()
    unique_paths = []
    seen_paths = set()
    skipped_duplicate_paths = 0
    for path in list(paths or []):
        key = _path_unique_key(path)
        if key in seen_paths:
            skipped_duplicate_paths += 1
            continue
        seen_paths.add(key)
        unique_paths.append(path)
    paths = unique_paths
    if skipped_duplicate_paths:
        sonuc.uyarilar.append(f"{skipped_duplicate_paths} tekrar fotoğraf yolu atlandı.")
    total = len(paths)
    merged_duplicate_rows = 0
    records_by_path = []
    for idx, path in enumerate(paths, start=1):
        if stop_event is not None and stop_event.is_set():
            break
        name = os.path.basename(str(path))
        if progress_callback:
            progress_callback(idx - 1, total, name, "okunuyor")
        try:
            source_sondaj_no = _sondaj_no_from_path(path)
            learned = spt_ogrenme_eslesmesi_bul(path)
            if learned:
                raw_items = [dict(learned)]
                raw_items[0].update({
                    "_motor": "yerel_ogrenme",
                    "_model": "duzeltilmis_kayit",
                    "guven": raw_items[0].get("guven") or "100",
                    "kaynak_hash": dosya_parmak_izi(path),
                })
            else:
                raw_items = yapay_zeka_ile_spt_oku(
                    path,
                    ayarlar=ayarlar,
                    stop_event=stop_event,
                )

            active_motor = spt_motorunu_normalize_et(ayarlar.get("aktif_motor") or "gemini")
            confidences = [
                safe_float(item.get("guven"))
                for item in raw_items
                if str(item.get("guven", "")).strip()
            ]
            pro_reason = ""
            if not raw_items:
                pro_reason = "ilk motor sonuc bulamadi"
            elif not confidences:
                pro_reason = "guven degeri yok"
            elif min(confidences) < (safe_float(guven_esigi) or 90):
                pro_reason = f"guven %{min(confidences):g} esigin altinda"

            second_opinion_motor = (
                "openai_pro"
                if active_motor in ("openai", "openai_pro", "openai_ust")
                else "openai"
            )
            if (
                not learned
                and auto_pro
                and pro_reason
                and ayarlar.get("openai_api_key")
            ):
                time.sleep(0.3)
                pro_items = yapay_zeka_ile_spt_oku(
                    path,
                    ayarlar=ayarlar,
                    motor_zorla=second_opinion_motor,
                    timeout=60,
                    stop_event=stop_event,
                )
                default_no = source_sondaj_no or default_sondaj_no
                initial_score = _raw_items_score(raw_items, default_no)
                pro_score = _raw_items_score(pro_items, default_no)
                if pro_score > initial_score:
                    for item in pro_items:
                        item["_auto_pro_nedeni"] = pro_reason
                        item["_onceki_motor_skoru"] = initial_score
                    raw_items = pro_items
                else:
                    for item in raw_items:
                        item["_auto_pro_nedeni"] = pro_reason
                        item["_pro_skoru"] = pro_score
                        item["_pro_sonucu_kullanilmadi"] = True
            if not raw_items:
                sonuc.uyarilar.append(f"{name}: SPT verisi bulunamadı")
            file_records = []
            file_locations = {}
            for item in raw_items:
                item = dict(item)
                item["kaynak"] = name
                item["kaynak_yolu"] = str(path)
                kayit = kayit_normalize_et(item, source_sondaj_no or default_sondaj_no)
                if source_sondaj_no:
                    ai_sondaj_no = normalize_sondaj_no(
                        item.get("sondaj_no") or item.get("sondaj") or item.get("kuyu_no")
                    )
                    if ai_sondaj_no and ai_sondaj_no != source_sondaj_no:
                        conflict = f"Fotoğraf {ai_sondaj_no}, dosya yolu {source_sondaj_no} gösteriyor"
                        kayit.raw["sondaj_celiskisi"] = conflict
                        kayit.uyari = ", ".join(filter(None, [kayit.uyari, conflict]))
                    kayit.sondaj_no = source_sondaj_no
                if kayit.uyari:
                    sonuc.uyarilar.append(f"{name}: {kayit.uyari}")
                loc_key = _spt_location_key(kayit)
                if loc_key[0] and loc_key[1] > 0:
                    existing_idx = file_locations.get(loc_key)
                    if existing_idx is not None:
                        existing = file_records[existing_idx]
                        if _spt_record_score(kayit) > _spt_record_score(existing):
                            file_records[existing_idx] = kayit
                        merged_duplicate_rows += 1
                        continue
                    file_locations[loc_key] = len(file_records)
                file_records.append(kayit)
            records_by_path.append((_path_unique_key(path), file_records))
        except Exception as exc:
            sonuc.uyarilar.append(f"{name}: {exc}")
            records_by_path.append((_path_unique_key(path), []))
        if progress_callback:
            progress_callback(idx, total, name, "tamam")

    selected_records, removed_by_sequence, merged_cross_photo_rows = _select_spt_records_for_batch(
        records_by_path,
        paths,
        path_order=(key for key, _records in records_by_path),
    )
    sonuc.kayitlar[:] = selected_records
    if merged_duplicate_rows:
        sonuc.uyarilar.append(f"{merged_duplicate_rows} tekrar SPT satırı aynı fotoğraf/derinlik olduğu için birleştirildi.")
    if removed_by_sequence:
        sonuc.uyarilar.append(f"{removed_by_sequence} ek SPT satırı fotoğraf sırası/derinlik kontrolüyle atlandı.")
    if merged_cross_photo_rows:
        sonuc.uyarilar.append(f"{merged_cross_photo_rows} tekrar SPT satırı aynı kuyu/derinlik olduğu için birleştirildi.")

    sonuc.kayitlar.sort(key=lambda item: (item.sondaj_no, safe_float(item.derinlik), item.kaynak))
    return sonuc


def spt_gecmis_kaydet(islem, kayit=None, detay=None):
    payload = {
        "tarih": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "islem": islem,
        "kayit": kayit.to_dict() if hasattr(kayit, "to_dict") else (kayit or {}),
        "detay": detay or {},
    }
    try:
        SPT_LOG_DIR.mkdir(parents=True, exist_ok=True)
        with SPT_GECMIS_PATH.open("a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=False) + "\n")
    except Exception as exc:
        log_exception("spt.history.write", exc_value=exc)
        payload["log_hatasi"] = str(exc)
    return payload


def spt_gecmisi_oku(limit=300):
    if not SPT_GECMIS_PATH.exists():
        return []
    rows = []
    try:
        with SPT_GECMIS_PATH.open("r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rows.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
    except Exception:
        return []
    return rows[-limit:]


def spt_ogrenme_kaydet(kayit, duzeltilmis=None, not_metni=""):
    duzeltilmis = duzeltilmis or {}
    SPT_OGRENME_DIR.mkdir(parents=True, exist_ok=True)
    image_dir = SPT_OGRENME_DIR / "resimler"
    image_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    copied_path = ""
    source_path = getattr(kayit, "kaynak_yolu", "")
    source_hash = dosya_parmak_izi(source_path) if source_path and os.path.exists(source_path) else ""
    if source_path and os.path.exists(source_path):
        suffix = Path(source_path).suffix or ".jpg"
        safe_name = re.sub(r"[^A-Za-z0-9_.-]+", "_", f"{timestamp}_{Path(source_path).stem}") + suffix
        target_path = image_dir / safe_name
        try:
            shutil.copy2(source_path, target_path)
            copied_path = str(target_path)
        except Exception:
            copied_path = source_path
    label = {
        "tarih": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "orijinal": kayit.to_dict() if hasattr(kayit, "to_dict") else {},
        "duzeltilmis": duzeltilmis,
        "not": not_metni,
        "kopyalanan_resim": copied_path,
        "kaynak_hash": source_hash,
    }
    labels_path = SPT_OGRENME_DIR / "etiketler.jsonl"
    with labels_path.open("a", encoding="utf-8") as f:
        f.write(json.dumps(label, ensure_ascii=False) + "\n")
    spt_gecmis_kaydet("dogrusunu_ogret", kayit, {"duzeltilmis": duzeltilmis, "kopyalanan_resim": copied_path})
    return label


def spt_ogrenme_eslesmesi_bul(source_path):
    """Daha once duzeltilmis ayni fotograf icin yerel dogruyu dondur."""
    if not source_path or not os.path.exists(source_path):
        return None
    source_hash = dosya_parmak_izi(source_path)
    if not source_hash:
        return None
    labels_path = SPT_OGRENME_DIR / "etiketler.jsonl"
    if not labels_path.exists():
        return None
    try:
        lines = labels_path.read_text(encoding="utf-8").splitlines()
    except Exception:
        return None
    for line in reversed(lines[-5000:]):
        try:
            item = json.loads(line)
        except (TypeError, json.JSONDecodeError):
            continue
        item_hash = item.get("kaynak_hash", "")
        if not item_hash and item.get("kopyalanan_resim"):
            item_hash = dosya_parmak_izi(item.get("kopyalanan_resim"))
        if item_hash != source_hash:
            continue
        corrected = item.get("duzeltilmis")
        if isinstance(corrected, dict) and corrected:
            result = dict(corrected)
            result["_ogrenme_tarihi"] = item.get("tarih", "")
            return result
    return None


def spt_kirp_kaydet(source_path, crop_box):
    try:
        from PIL import Image, ImageOps
    except Exception as exc:
        raise RuntimeError(f"Pillow yüklenemedi: {exc}") from exc
    if not source_path or not os.path.exists(source_path):
        raise RuntimeError("Kırpılacak fotoğraf bulunamadı.")
    left, top, right, bottom = [int(round(v)) for v in crop_box]
    if right <= left or bottom <= top:
        raise RuntimeError("Geçerli bir kırpma alanı seçilmedi.")
    SPT_CROP_DIR.mkdir(parents=True, exist_ok=True)
    with Image.open(source_path) as image:
        try:
            image = ImageOps.exif_transpose(image)
        except Exception:
            pass
        left = max(0, min(left, image.width - 1))
        right = max(left + 1, min(right, image.width))
        top = max(0, min(top, image.height - 1))
        bottom = max(top + 1, min(bottom, image.height))
        cropped = image.crop((left, top, right, bottom))
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", Path(source_path).stem)
        out_path = SPT_CROP_DIR / f"{timestamp}_{safe_stem}_crop.jpg"
        cropped.convert("RGB").save(out_path, "JPEG", quality=95)
    return str(out_path)


def spt_kaynak_raporu_kaydet(kayitlar, path):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise RuntimeError(f"openpyxl yüklenemedi: {exc}") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "SPT Kaynak Raporu"
    headers = [
        "Sondaj", "Derinlik", "15", "30", "45", "N30",
        "Güven", "Uyarı", "Okunan Derinlik", "Kullanılan Derinlik",
        "Motor", "Model", "Okuma Süresi (sn)", "Kaynak Hash",
        "Kaynak", "Kaynak Yolu",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for kayit in kayitlar:
        raw = getattr(kayit, "raw", {}) or {}
        ws.append(excel_satiri_guvenli_yap([
            kayit.sondaj_no,
            kayit.derinlik, kayit.v15, kayit.v30, kayit.v45,
            kayit.n30, kayit.guven, kayit.uyari,
            raw.get("okunan_derinlik", ""),
            raw.get("hedef_derinlik", kayit.derinlik),
            raw.get("motor", ""),
            raw.get("model", ""),
            raw.get("okuma_suresi", ""),
            raw.get("kaynak_hash", ""),
            kayit.kaynak,
            _paylasilabilir_dosya_adi(kayit.kaynak_yolu),
        ]))
    widths = [14, 10, 8, 8, 8, 8, 10, 32, 16, 18, 16, 28, 16, 30, 28, 60]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)
    return path
