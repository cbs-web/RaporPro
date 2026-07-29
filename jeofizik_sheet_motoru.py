# Dosya: RaporPro/jeofizik_sheet_motoru.py
import csv
import math
import os
import re
import unicodedata

try:
    from yardimcilar import safe_float
except ImportError:
    def safe_float(value):
        try:
            return float(str(value).replace(",", "."))
        except Exception:
            return 0.0


JEOFIZIK_SHEET_DEFAULT_ROWS = 120
JEOFIZIK_SHEET_DEFAULT_COLS = 8


def _is_nan(value):
    try:
        return isinstance(value, float) and math.isnan(value)
    except Exception:
        return False


def _cell_text(value):
    if value is None or _is_nan(value):
        return ""
    return str(value).strip()


def _is_blank(value):
    text = _cell_text(value)
    return text in {"", "-", "None", "none", "nan", "NaN", "null"}


def _norm(value):
    text = _cell_text(value)
    text = text.translate(str.maketrans({
        "İ": "I", "ı": "i", "Ş": "S", "ş": "s", "Ğ": "G", "ğ": "g",
        "Ü": "U", "ü": "u", "Ö": "O", "ö": "o", "Ç": "C", "ç": "c",
    }))
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def jeofizik_serim_anahtari(value):
    """SS-1, SS 1 ve Serim 1 gibi adları aynı kalıcı anahtara çevir."""
    normalized = _norm(value)
    compact = re.sub(r"[^a-z0-9]+", "", normalized)
    match = re.fullmatch(r"(?:ss|serim|sismikserim)0*(\d+)", compact)
    if match:
        return f"serim:{int(match.group(1))}"
    return f"ad:{normalized}"


def _row_text(row):
    return " ".join(_norm(cell) for cell in row if not _is_blank(cell))


def _row_first_text(row, count=2):
    return " ".join(_norm(cell) for cell in list(row or [])[:count] if not _is_blank(cell))


def _value(row, idx, default="-"):
    if idx >= len(row):
        return default
    value = row[idx]
    return default if _is_blank(value) else value


def _format_serim_name(raw, fallback_idx):
    text = _cell_text(raw)
    if not text:
        return f"Serim {fallback_idx}"
    match = re.search(r"(serim|ss)\s*[-:]?\s*(\d+)", text, flags=re.IGNORECASE)
    if match:
        prefix = "Serim" if match.group(1).lower().startswith("serim") else "SS"
        return f"{prefix} {match.group(2)}" if prefix == "Serim" else f"SS-{match.group(2)}"
    return text


def _serim_adi_bul(row, fallback_idx):
    whole = _row_text(row)
    if "sismik olcu" not in whole and "serim" not in whole and not re.search(r"\bss\s*-?\s*\d+\b", whole):
        return ""
    for cell in row or []:
        norm = _norm(cell)
        if re.search(r"\bserim\s*[-:]?\s*\d+\b", norm) or re.search(r"\bss\s*-?\s*\d+\b", norm):
            return _format_serim_name(cell, fallback_idx)
    return f"Serim {fallback_idx}"


def _parametre_anahtari(row):
    text = _row_first_text(row, 2)
    if not text:
        return ""
    if "vs30" in text:
        return "vs30"
    if "vp/vs" in text or "hiz orani" in text:
        return "ratio"
    if "buyutme" in text:
        return "b"
    if "tabaka kalin" in text:
        return "h"
    if "tabaka yogun" in text or "yogunluk" in text:
        return "rho"
    if "poisson" in text:
        return "nu"
    if "elastisite" in text or "young" in text:
        return "E"
    if "vp" in text or "boyuna dalga" in text:
        return "vp"
    if "vs" in text or "enine dalga" in text:
        return "vs"
    if "kayma" in text or "gmax" in text:
        return "G"
    if "bulk" in text or "sikismaz" in text:
        return "K"
    return ""


def jeofizik_sheet_rows_temizle(rows):
    clean_rows = []
    for row in rows or []:
        cells = [_cell_text(cell) for cell in (row or [])]
        while cells and cells[-1] == "":
            cells.pop()
        clean_rows.append(cells)
    while clean_rows and not any(_cell_text(cell) for cell in clean_rows[-1]):
        clean_rows.pop()
    return clean_rows


def jeofizik_sheet_grid_hazirla(rows, min_rows=JEOFIZIK_SHEET_DEFAULT_ROWS, min_cols=JEOFIZIK_SHEET_DEFAULT_COLS):
    clean_rows = jeofizik_sheet_rows_temizle(rows)
    col_count = max(min_cols, max((len(row) for row in clean_rows), default=0))
    row_count = max(min_rows, len(clean_rows))
    grid = []
    for idx in range(row_count):
        source = clean_rows[idx] if idx < len(clean_rows) else []
        grid.append(source + [""] * max(0, col_count - len(source)))
    return grid


def jeofizik_sheet_var_mi(veri):
    rows = (veri or {}).get("jeofizik_sheet", {}).get("rows", [])
    return any(any(_cell_text(cell) for cell in row) for row in rows or [])


def jeofizik_sheet_rapora_hazir_mi(veri):
    """Sheet'in yalnızca dolu değil, raporda kullanılabilir olduğunu döndür."""
    return bool(
        jeofizik_sheet_var_mi(veri)
        and jeofizik_sheet_ozeti(veri).get("ready", False)
    )


def _hesapla_parametreler(vp, vs, h="", rho=None):
    vp_f, vs_f = safe_float(vp), safe_float(vs)
    if vp_f <= 0 or vs_f <= 0:
        return {"nu": 0, "rho": 0, "G": 0, "E": 0, "K": 0, "ratio": 0}
    rho_f = safe_float(rho)
    if rho_f <= 0:
        rho_f = 0.31 * (vp_f ** 0.25)
    nu = (vp_f ** 2 - 2 * vs_f ** 2) / (2 * (vp_f ** 2 - vs_f ** 2)) if vp_f > vs_f else 0.49
    g_mod = (rho_f * vs_f ** 2) / 100
    e_mod = 2 * g_mod * (1 + nu)
    k_mod = e_mod / (3 * (1 - 2 * nu)) if (1 - 2 * nu) != 0 else 0
    return {
        "nu": round(nu, 2),
        "rho": round(rho_f, 2),
        "G": round(g_mod, 2),
        "E": round(e_mod, 2),
        "K": round(k_mod, 2),
        "ratio": round(vp_f / vs_f if vs_f else 0, 2),
    }


def _vs30_hesapla(layers):
    toplam_h, toplam_t = 0.0, 0.0
    for idx, layer in enumerate(layers or []):
        vs = safe_float(layer.get("vs"))
        if vs <= 0:
            continue
        kalan = 30.0 - toplam_h
        if kalan <= 0:
            break
        h = safe_float(layer.get("h"))
        use_h = kalan if idx == len(layers) - 1 else min(h, kalan)
        if use_h <= 0:
            continue
        toplam_h += use_h
        toplam_t += use_h / vs
    return round(30.0 / toplam_t, 2) if toplam_t > 0 else 0


def _bloklari_ayir(rows):
    blocks = []
    current = None
    for row in rows:
        serim_adi = _serim_adi_bul(row, len(blocks) + 1)
        if serim_adi:
            current = {"ad": serim_adi, "raw": {}}
            blocks.append(current)
            continue
        if current is None:
            continue
        key = _parametre_anahtari(row)
        if key:
            current["raw"][key] = row
    return blocks


def _bloktan_ss_uret(block, warnings):
    raw = block.get("raw", {})
    if not raw.get("vp") and not raw.get("vs"):
        warnings.append(f"{block.get('ad', 'Serim')}: VP/VS satiri bulunamadi.")
        return None

    max_cols = max((len(row) for row in raw.values()), default=0)
    layers = []
    for col_idx in range(2, max_cols):
        vp = _value(raw.get("vp", []), col_idx)
        vs = _value(raw.get("vs", []), col_idx)
        if _is_blank(vp) and _is_blank(vs):
            continue
        layer = {
            "vp": vp,
            "vs": vs,
            "h": _value(raw.get("h", []), col_idx),
            "rho": _value(raw.get("rho", []), col_idx),
            "nu": _value(raw.get("nu", []), col_idx),
            "E": _value(raw.get("E", []), col_idx),
            "G": _value(raw.get("G", []), col_idx),
            "K": _value(raw.get("K", []), col_idx),
            "b": _value(raw.get("b", []), col_idx),
            "ratio": _value(raw.get("ratio", []), col_idx),
            "vs30": "-",
        }
        hesap = _hesapla_parametreler(layer["vp"], layer["vs"], layer["h"], layer["rho"])
        for key in ("rho", "nu", "E", "G", "K", "ratio"):
            if _is_blank(layer.get(key)):
                layer[key] = hesap.get(key, "-")
        layers.append(layer)

    if not layers:
        warnings.append(f"{block.get('ad', 'Serim')}: tabaka verisi bulunamadi.")
        return None

    vs30 = "-"
    raw_vs30 = raw.get("vs30", [])
    for val in raw_vs30[2:]:
        if not _is_blank(val):
            vs30 = val
            break
    if _is_blank(vs30):
        vs30 = _vs30_hesapla(layers) or "-"
    layers[0]["vs30"] = vs30

    return {"ad": block.get("ad") or "Serim", "coords": [""] * 6, "layers": layers}


def jeofizik_sheet_satirlarini_coz(rows):
    clean_rows = jeofizik_sheet_rows_temizle(rows)
    warnings = []
    ss_list = []
    for block in _bloklari_ayir(clean_rows):
        ss = _bloktan_ss_uret(block, warnings)
        if ss:
            ss_list.append(ss)
    if clean_rows and not ss_list:
        warnings.append("Jeofizik Sheet icinde okunabilir Serim/SS parametre blogu bulunamadi.")
    return {"ss_list": ss_list, "warnings": warnings}


def jeofizik_sheet_rows_to_ss_list(rows):
    return jeofizik_sheet_satirlarini_coz(rows).get("ss_list", [])


def jeofizik_sheet_ozeti(veri_or_rows):
    rows = veri_or_rows
    if isinstance(veri_or_rows, dict):
        rows = veri_or_rows.get("jeofizik_sheet", {}).get("rows", [])
    parsed = jeofizik_sheet_satirlarini_coz(rows)
    ss_list = parsed.get("ss_list", [])
    layer_count = sum(len(ss.get("layers", []) or []) for ss in ss_list)
    return {
        "ready": bool(ss_list),
        "serim": len(ss_list),
        "layers": layer_count,
        "warnings": parsed.get("warnings", []),
    }


def jeofizik_ss_koordinatlarini_koru(yeni_ss_list, mevcut_ss_list):
    coords_by_name = {}
    for ss in mevcut_ss_list or []:
        coords = list(ss.get("coords", []) or [])
        if any(_cell_text(v) for v in coords):
            coords_by_name[jeofizik_serim_anahtari(ss.get("ad", ""))] = coords
    for ss in yeni_ss_list or []:
        coords = coords_by_name.get(jeofizik_serim_anahtari(ss.get("ad", "")))
        if coords:
            coords = list(coords)
            while len(coords) < 6:
                coords.append("")
            ss["coords"] = coords[:6]
    return yeni_ss_list


def jeofizik_excel_dosyasi_oku(path):
    ext = os.path.splitext(path or "")[1].lower()
    rows = []
    if ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.reader(handle):
                rows.append(["" if cell is None else str(cell) for cell in row])
        return rows
    if ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True, read_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                rows.append(["" if cell is None else cell for cell in row])
        finally:
            wb.close()
        return rows
    try:
        import pandas as pd
        df = pd.read_excel(path, header=None)
        for _, row in df.iterrows():
            rows.append(["" if pd.isna(cell) else cell for cell in row.tolist()])
        return rows
    except Exception:
        raise
