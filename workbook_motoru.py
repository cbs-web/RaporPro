import datetime
import unicodedata

from karot_motoru import derinlik_baslangic
from performans import perf_tracked
from yardimcilar import litoloji_yazim_uyarilari, safe_float, temizle_baslik


WORKBOOK_SHEET_DEFS = {
    "sondajlar": {
        "title": "Sondajlar",
        "columns": [
            ("SondajNo", "no"), ("Derinlik", "der"),
            ("Enlem", "y"), ("Boylam", "x"), ("Kot", "k"),
            ("Bas.Tarih", "bas_tar"), ("Bit.Tarih", "bit_tar"), ("YASS Ilk", "yass_d1"),
            ("YASS T1", "yass_t1"), ("YASS Son", "yass_d2"), ("YASS T2", "yass_t2")
        ],
        "widths": [110, 85, 130, 130, 80, 110, 110, 85, 110, 85, 110],
    },
    "litoloji": {
        "title": "Litoloji",
        "columns": [("SondajNo", "sondaj_no"), ("Baslangic", "top"), ("Bitis", "bot"), ("Tanim", "tanim")],
        "widths": [110, 90, 90, 360],
    },
    "spt": {
        "title": "SPT",
        "columns": [("SondajNo", "sondaj_no"), ("Derinlik", "der"), ("15", "v15"), ("30", "v30"), ("45", "v45"), ("N30", "n30")],
        "widths": [110, 90, 70, 70, 70, 70],
    },
    "pmt": {
        "title": "PMT",
        "columns": [("SondajNo", "sondaj_no"), ("Derinlik", "der"), ("Em", "em"), ("Pl", "pl")],
        "widths": [110, 90, 110, 110],
    },
    "kaya": {
        "title": "Kaya",
        "columns": [("SondajNo", "sondaj_no"), ("Derinlik", "der"), ("TCR", "tcr"), ("SCR", "scr"), ("RQD", "rqd")],
        "widths": [110, 90, 80, 80, 80],
    },
    "numune": {
        "title": "Numune",
        "columns": [("SondajNo", "sondaj_no"), ("Derinlik/Aralik", "aralik"), ("Turu/No", "tur")],
        "widths": [110, 150, 180],
    },
}


def yeni_sondaj_sablonu(idx):
    bugun = datetime.datetime.now()
    bugun_str = bugun.strftime("%d.%m.%Y")
    t2_str = (bugun + datetime.timedelta(days=10)).strftime("%d.%m.%Y")
    return {
        "no": f"SK-{idx + 1}", "der": "15.0", "y": "", "x": "", "k": "",
        "bas_tar": bugun_str, "bit_tar": bugun_str,
        "yass_d1": "", "yass_t1": bugun_str, "yass_d2": "", "yass_t2": t2_str,
        "litoloji": [], "spt": [], "pmt": [], "kaya": [], "numuneler": []
    }


def sondaj_turu_degeri(sondaj):
    text = str((sondaj or {}).get("sondaj_turu", "")).strip().lower()
    if text in ("kaya", "rock"):
        return "Kaya"
    if text in ("zemin", "soil"):
        return "Zemin"
    return "Kaya" if (sondaj or {}).get("kaya") else "Zemin"


def sondaj_delgi_capi_degeri(sondaj, fallback="76mm"):
    text = str((sondaj or {}).get("delgi_capi") or fallback or "76mm").strip().replace(" ", "")
    if text.lower() in ("76", "76mm"):
        return "76mm"
    if text.lower() in ("89", "89mm"):
        return "89mm"
    return "76mm"


def proje_sondaj_turu_degeri(veri):
    text = str(((veri or {}).get("ayarlar", {}) or {}).get("sondaj_turu") or "Zemin").strip().lower()
    if text in ("kaya", "rock"):
        return "Kaya"
    return "Zemin"


def sondaj_sayfa_degeri(sondaj, col_key, default_delgi_capi="76mm"):
    if col_key == "sondaj_turu":
        return sondaj_turu_degeri(sondaj)
    if col_key == "delgi_capi":
        return sondaj_delgi_capi_degeri(sondaj, default_delgi_capi)
    return (sondaj or {}).get(col_key, "")


def normalize_header(cell):
    text = str(cell).strip().lower()
    text = text.replace("\u0131", "i").replace("\u0130", "i")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return temizle_baslik(text)


def row_has_data(values, ignored=None):
    ignored = ignored or set()
    return any(str(value).strip() for key, value in values.items() if key not in ignored)


def row_values_to_list(sheet_key, values, sheet_defs=None):
    sheet_defs = sheet_defs or WORKBOOK_SHEET_DEFS
    return [values.get(col_key, "") for _, col_key in sheet_defs[sheet_key]["columns"]]


@perf_tracked("workbook.excel_write")
def excel_workbook_yaz(path, sheet_payloads):
    """UI'dan bağımsız Excel yazıcısı; büyük workbook aktarımını arka planda çalıştırır."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    total_rows = 0
    for payload in sheet_payloads:
        ws = wb.create_sheet(str(payload.get("title") or "Sayfa"))
        ws.append(list(payload.get("headers") or []))
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        rows = list(payload.get("rows") or [])
        for row in rows:
            ws.append(list(row))
        total_rows += len(rows)
        for col_idx, width in enumerate(payload.get("widths") or [], start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, float(width) / 7)
    wb.save(path)
    return {"path": str(path), "sheet_count": len(sheet_payloads), "row_count": total_rows}


def normalize_data_rows(sheet_key, data, sheet_defs=None):
    sheet_defs = sheet_defs or WORKBOOK_SHEET_DEFS
    ncols = len(sheet_defs[sheet_key]["columns"])
    normalized = []
    for row in data:
        values = ["" if cell is None else str(cell) for cell in list(row)[:ncols]]
        if len(values) < ncols:
            values.extend([""] * (ncols - len(values)))
        normalized.append(values)
    return normalized


def rows_to_dicts(sheet_key, data, sheet_defs=None):
    sheet_defs = sheet_defs or WORKBOOK_SHEET_DEFS
    normalized = normalize_data_rows(sheet_key, data, sheet_defs)
    keys = [col_key for _, col_key in sheet_defs[sheet_key]["columns"]]
    return [{key: str(row[idx]).strip() for idx, key in enumerate(keys)} for row in normalized]


def ensure_dict_rows(sheet_key, rows, sheet_defs=None):
    sheet_defs = sheet_defs or WORKBOOK_SHEET_DEFS
    if not rows:
        return []
    if all(isinstance(row, dict) for row in rows):
        return [
            {key: "" if value is None else str(value).strip() for key, value in row.items()}
            for row in rows
        ]
    return rows_to_dicts(sheet_key, rows, sheet_defs)


def ensure_workbook_dict_rows(rows_by_sheet, sheet_defs=None):
    sheet_defs = sheet_defs or WORKBOOK_SHEET_DEFS
    return {
        sheet_key: ensure_dict_rows(sheet_key, rows_by_sheet.get(sheet_key, []), sheet_defs)
        for sheet_key in sheet_defs
    }


def default_row_values(sheet_key, rows_by_sheet=None, row_count=0):
    rows_by_sheet = rows_by_sheet or {}
    if sheet_key == "sondajlar":
        return yeni_sondaj_sablonu(row_count)
    values = {}
    data = rows_by_sheet.get(sheet_key, [])
    if data:
        last = data[-1]
        values["sondaj_no"] = last.get("sondaj_no", "")
        if sheet_key == "litoloji":
            values["top"] = last.get("bot", "")
        elif sheet_key == "spt":
            values["der"] = f"{safe_float(last.get('der')) + 1.5:.2f}" if last.get("der") else ""
    else:
        sondajlar = rows_by_sheet.get("sondajlar", [])
        values["sondaj_no"] = next((r.get("no") for r in sondajlar if r.get("no")), "SK-1")
    return values


def build_initial_rows(veri, sheet_defs=None):
    sheet_defs = sheet_defs or WORKBOOK_SHEET_DEFS
    initial = {key: [] for key in sheet_defs}
    source_nos = []
    default_delgi_capi = ((veri or {}).get("ayarlar", {}) or {}).get("delgi_capi", "76mm")
    for sondaj in veri.get("sondaj", []):
        initial["sondajlar"].append([
            sondaj_sayfa_degeri(sondaj, col_key, default_delgi_capi)
            for _, col_key in sheet_defs["sondajlar"]["columns"]
        ])
        source_nos.append(sondaj.get("no", ""))
        no = sondaj.get("no", "")
        for row in sondaj.get("litoloji", []):
            initial["litoloji"].append([no, row[0] if len(row) > 0 else "", row[1] if len(row) > 1 else "", row[2] if len(row) > 2 else ""])
        for row in sondaj.get("spt", []):
            initial["spt"].append([no, row[0] if len(row) > 0 else "", row[1] if len(row) > 1 else "", row[2] if len(row) > 2 else "", row[3] if len(row) > 3 else "", row[4] if len(row) > 4 else ""])
        for row in sondaj.get("pmt", []):
            initial["pmt"].append([no, row[0] if len(row) > 0 else "", row[1] if len(row) > 1 else "", row[2] if len(row) > 2 else ""])
        for row in sondaj.get("kaya", []):
            initial["kaya"].append([no, row[0] if len(row) > 0 else "", row[1] if len(row) > 1 else "", row[2] if len(row) > 2 else "", row[3] if len(row) > 3 else ""])
        for row in sondaj.get("numuneler", []):
            initial["numune"].append([no, row[0] if len(row) > 0 else "", row[1] if len(row) > 1 else ""])
    return initial, source_nos


def header_map(sheet_key, cells, sheet_defs=None):
    sheet_defs = sheet_defs or WORKBOOK_SHEET_DEFS
    aliases = {
        "sondajno": "sondaj_no", "sondaj": "sondaj_no", "sk": "sondaj_no", "kuyuno": "sondaj_no",
        "no": "no", "sondajadi": "no", "derinlik": "der", "der": "der", "derinlikm": "der",
        "enlem": "y", "lat": "y", "latitude": "y", "y": "y", "boylam": "x", "lon": "x", "longitude": "x", "x": "x",
        "tur": "sondaj_turu", "turu": "sondaj_turu", "sondajturu": "sondaj_turu", "zeminkaya": "sondaj_turu",
        "delgicapi": "delgi_capi", "delgicap": "delgi_capi", "cap": "delgi_capi", "capi": "delgi_capi",
        "kot": "k", "bastarih": "bas_tar", "bastarihi": "bas_tar", "baslangictarihi": "bas_tar",
        "bittarih": "bit_tar", "bittarihi": "bit_tar", "bitistarihi": "bit_tar",
        "yassilk": "yass_d1", "yassd1": "yass_d1", "yass1": "yass_d1", "yasst1": "yass_t1", "yassilktarih": "yass_t1",
        "yassson": "yass_d2", "yassd2": "yass_d2", "yass2": "yass_d2", "yasst2": "yass_t2", "yasssontarih": "yass_t2",
        "baslangic": "top", "bas": "top", "ust": "top", "top": "top", "bitis": "bot", "bit": "bot", "alt": "bot", "bot": "bot",
        "tanim": "tanim", "litoloji": "tanim", "birim": "tanim",
        "15": "v15", "n15": "v15", "30": "v30", "n30vurus": "v30", "45": "v45", "n45": "v45", "n30": "n30",
        "em": "em", "pl": "pl", "tcr": "tcr", "scr": "scr", "rqd": "rqd",
        "aralik": "aralik", "derinlikaralik": "aralik", "tur": "tur", "turu": "tur", "turuno": "tur", "numune": "tur",
    }
    allowed = {key for _, key in sheet_defs[sheet_key]["columns"]}
    mapped = []
    for cell in cells:
        normalized = normalize_header(cell)
        key = aliases.get(normalized)
        if sheet_key == "sondajlar" and normalized in ("tur", "turu", "sondajturu", "zeminkaya"):
            key = "sondaj_turu"
        if sheet_key == "sondajlar" and key == "sondaj_no":
            key = "no"
        elif sheet_key != "sondajlar" and key == "no":
            key = "sondaj_no"
        mapped.append(key if key in allowed else None)
    return mapped if sum(1 for item in mapped if item) >= 2 else None


def calc_n30(v30, v45, existing=""):
    if str(existing).strip():
        return str(existing).strip()
    joined = " ".join([str(v30), str(v45)]).lower()
    if "50/" in joined or "-" in joined:
        return "R"
    total = safe_float(v30) + safe_float(v45)
    if not total:
        return ""
    return str(int(total)) if float(total).is_integer() else str(total)


def validate_rows(rows_by_sheet):
    rows_by_sheet = ensure_workbook_dict_rows(rows_by_sheet)
    errors = []
    warnings = []
    sondaj_rows = rows_by_sheet.get("sondajlar", [])
    valid_nos, seen, depth_by_no = set(), {}, {}

    def mark(sheet_key, row_idx, col_key, level="error"):
        item = (sheet_key, row_idx, col_key)
        (errors if level == "error" else warnings).append(item)

    for row_idx, values in enumerate(sondaj_rows):
        if not row_has_data(values):
            continue
        no = values.get("no", "").strip()
        if not no:
            mark("sondajlar", row_idx, "no")
            continue
        if no in seen:
            mark("sondajlar", row_idx, "no")
            mark("sondajlar", seen[no], "no")
        seen[no] = row_idx
        valid_nos.add(no)
        der = safe_float(values.get("der"))
        depth_by_no[no] = der
        if der <= 0:
            mark("sondajlar", row_idx, "der", "warning")
        for coord_key in ("y", "x"):
            val = values.get(coord_key, "").strip()
            if val and safe_float(val) == 0:
                mark("sondajlar", row_idx, coord_key, "warning")

    for sheet_key in ("litoloji", "spt", "pmt", "kaya", "numune"):
        for row_idx, values in enumerate(rows_by_sheet.get(sheet_key, [])):
            if not row_has_data(values, {"sondaj_no"}):
                continue
            no = values.get("sondaj_no", "").strip()
            if not no or no not in valid_nos:
                mark(sheet_key, row_idx, "sondaj_no")

    lit_by_no = {}
    for row_idx, values in enumerate(rows_by_sheet.get("litoloji", [])):
        if not row_has_data(values, {"sondaj_no"}):
            continue
        no = values.get("sondaj_no", "").strip()
        top_d, bot_d = safe_float(values.get("top", "")), safe_float(values.get("bot", ""))
        if bot_d <= top_d:
            mark("litoloji", row_idx, "top")
            mark("litoloji", row_idx, "bot")
        if no in depth_by_no and depth_by_no[no] > 0 and bot_d > depth_by_no[no]:
            mark("litoloji", row_idx, "bot", "warning")
        if litoloji_yazim_uyarilari(values.get("tanim", "")):
            mark("litoloji", row_idx, "tanim", "warning")
        lit_by_no.setdefault(no, []).append((row_idx, top_d, bot_d))

    for no, rows in lit_by_no.items():
        rows = sorted(rows, key=lambda item: item[1])
        if rows and rows[0][1] > 0.05:
            mark("litoloji", rows[0][0], "top", "warning")
        prev_bot = None
        for row_idx, top_d, bot_d in rows:
            if prev_bot is not None:
                if top_d < prev_bot - 0.01:
                    mark("litoloji", row_idx, "top")
                elif top_d > prev_bot + 0.01:
                    mark("litoloji", row_idx, "top", "warning")
            prev_bot = bot_d
        if no in depth_by_no and depth_by_no[no] > 0 and prev_bot is not None and prev_bot < depth_by_no[no] - 0.05:
            mark("litoloji", rows[-1][0], "bot", "warning")

    for sheet_key in ("spt", "pmt", "kaya"):
        for row_idx, values in enumerate(rows_by_sheet.get(sheet_key, [])):
            if not row_has_data(values, {"sondaj_no"}):
                continue
            no = values.get("sondaj_no", "").strip()
            der = derinlik_baslangic(values.get("der")) if sheet_key == "kaya" else safe_float(values.get("der"))
            if der <= 0:
                mark(sheet_key, row_idx, "der", "warning")
            if no in depth_by_no and depth_by_no[no] > 0 and der > depth_by_no[no]:
                mark(sheet_key, row_idx, "der")
            if no in lit_by_no and der > 0:
                intervals = [(top_d, bot_d) for _, top_d, bot_d in lit_by_no[no] if bot_d > top_d]
                if intervals and not any(top_d - 0.05 <= der <= bot_d + 0.05 for top_d, bot_d in intervals):
                    mark(sheet_key, row_idx, "der", "warning")

    return {"errors": errors, "warnings": warnings}


def apply_rows_to_veri(current_veri, rows_by_sheet, source_nos=None):
    rows_by_sheet = ensure_workbook_dict_rows(rows_by_sheet)
    source_nos = source_nos or []
    old_by_no = {s.get("no", ""): s for s in current_veri.get("sondaj", []) if s.get("no")}
    no_alias, new_sondajlar, warnings = {}, [], []
    for idx, values in enumerate(rows_by_sheet.get("sondajlar", [])):
        if not row_has_data(values):
            continue
        old_no = source_nos[idx] if idx < len(source_nos) else ""
        no = values.get("no") or f"SK-{len(new_sondajlar) + 1}"
        source = old_by_no.get(old_no) or old_by_no.get(no) or yeni_sondaj_sablonu(len(new_sondajlar))
        sondaj = source.copy()
        sondaj.update(values)
        sondaj["no"] = no
        sondaj["sondaj_turu"] = proje_sondaj_turu_degeri(current_veri)
        sondaj["delgi_capi"] = sondaj_delgi_capi_degeri(
            {},
            (current_veri.get("ayarlar", {}) or {}).get("delgi_capi", "76mm"),
        )
        if old_no and old_no != no:
            no_alias[old_no] = no
        for key in ("litoloji", "spt", "pmt", "kaya", "numuneler"):
            sondaj[key] = []
        new_sondajlar.append(sondaj)

    by_no = {s.get("no", ""): s for s in new_sondajlar}

    def target_sondaj(values):
        raw_no = values.get("sondaj_no", "")
        no = no_alias.get(raw_no, raw_no)
        return by_no.get(no) if no else None

    for values in rows_by_sheet.get("litoloji", []):
        if not row_has_data(values, {"sondaj_no"}):
            continue
        sondaj = target_sondaj(values)
        if not sondaj:
            warnings.append(f"Litoloji satırı atlandı: {values.get('sondaj_no', '')}")
            continue
        sondaj["litoloji"].append([values.get("top", ""), values.get("bot", ""), values.get("tanim", "")])
    for values in rows_by_sheet.get("spt", []):
        if not row_has_data(values, {"sondaj_no"}):
            continue
        sondaj = target_sondaj(values)
        if not sondaj:
            warnings.append(f"SPT satırı atlandı: {values.get('sondaj_no', '')}")
            continue
        sondaj["spt"].append([values.get("der", ""), values.get("v15", ""), values.get("v30", ""), values.get("v45", ""), values.get("n30", "")])
    for values in rows_by_sheet.get("pmt", []):
        if not row_has_data(values, {"sondaj_no"}):
            continue
        sondaj = target_sondaj(values)
        if not sondaj:
            warnings.append(f"PMT satırı atlandı: {values.get('sondaj_no', '')}")
            continue
        sondaj["pmt"].append([values.get("der", ""), values.get("em", ""), values.get("pl", "")])
    for values in rows_by_sheet.get("kaya", []):
        if not row_has_data(values, {"sondaj_no"}):
            continue
        sondaj = target_sondaj(values)
        if not sondaj:
            warnings.append(f"Kaya satırı atlandı: {values.get('sondaj_no', '')}")
            continue
        sondaj["kaya"].append([values.get("der", ""), values.get("tcr", ""), values.get("scr", ""), values.get("rqd", "")])
    for values in rows_by_sheet.get("numune", []):
        if not row_has_data(values, {"sondaj_no"}):
            continue
        sondaj = target_sondaj(values)
        if not sondaj:
            warnings.append(f"Numune satırı atlandı: {values.get('sondaj_no', '')}")
            continue
        sondaj["numuneler"].append([values.get("aralik", ""), values.get("tur", "")])

    return new_sondajlar, warnings
