# Dosya: RaporPro/pmt_excel_motoru.py
import os
import re
from pathlib import Path

try:
    from yardimcilar import safe_float
except ImportError:
    def safe_float(value):
        try:
            return float(str(value).replace(",", "."))
        except Exception:
            return 0.0


def _is_blank(value):
    return value is None or str(value).strip() in {"", "-", "None", "none", "nan", "NaN", "null"}


def _number(value):
    if _is_blank(value):
        return None
    try:
        return float(str(value).strip().replace(",", "."))
    except Exception:
        return None


def _fmt_depth(value):
    number = _number(value)
    if number is None:
        return ""
    return f"{number:.2f}"


def _fmt_em(value):
    number = _number(value)
    if number is None:
        return ""
    return str(int(round(number)))


def _fmt_pl(value):
    number = _number(value)
    if number is None:
        return ""
    if abs(number - round(number)) < 1e-9:
        return str(int(round(number)))
    return f"{number:.2f}".rstrip("0").rstrip(".")


def _sondaj_no_normalize(value):
    text = str(value or "").strip().upper()
    text = text.replace("_", "-").replace(" ", "")
    match = re.search(r"SK-?(\d+)", text)
    if match:
        return f"SK-{int(match.group(1))}"
    return text


def _dosya_adindan_sondaj_ve_derinlik(path):
    name = Path(path).stem.upper()
    sondaj = ""
    derinlik = ""
    sondaj_match = re.search(r"SK[-_ ]?(\d+)", name)
    if sondaj_match:
        sondaj = f"SK-{int(sondaj_match.group(1))}"
    depth_match = re.search(r"SK[-_ ]?\d+[-_ ]+(\d+(?:[,_]\d+)?)\s*M", name)
    if depth_match:
        derinlik = _fmt_depth(depth_match.group(1).replace("_", ".").replace(",", "."))
    return sondaj, derinlik


def _em_formulunden_hesapla(ws):
    vo = _number(ws["I6"].value)
    p1 = _number(ws["C53"].value)
    p2 = _number(ws["C54"].value)
    v1 = _number(ws["E53"].value)
    v2 = _number(ws["E54"].value)
    if None in (vo, p1, p2, v1, v2):
        return None
    delta_v = v2 - v1
    if abs(delta_v) < 1e-12:
        return None
    return 2.66 * (vo + ((v1 + v2) / 2.0)) * ((p2 - p1) / delta_v)


def pmt_excel_dosyasi_oku(path):
    """Presiyometre Excel dosyasından tek PMT kaydı okur."""
    from openpyxl import load_workbook

    warnings = []
    sondaj_from_name, derinlik_from_name = _dosya_adindan_sondaj_ve_derinlik(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb["Presiyometre"] if "Presiyometre" in wb.sheetnames else wb.active
        sondaj_no = _sondaj_no_normalize(ws["F7"].value) or sondaj_from_name
        derinlik = _fmt_depth(ws["F8"].value) or derinlik_from_name

        em_raw = _number(ws["S16"].value)
        if em_raw is None:
            em_raw = _em_formulunden_hesapla(ws)
        pl_raw = _number(ws["S22"].value)
        net_pl_raw = _number(ws["S19"].value)

        if not sondaj_no:
            warnings.append(f"{os.path.basename(path)}: sondaj no okunamadı.")
        if not derinlik:
            warnings.append(f"{os.path.basename(path)}: deney derinliği okunamadı.")
        if em_raw is None:
            warnings.append(f"{os.path.basename(path)}: Em/Es değeri hesaplanamadı.")
        if pl_raw is None:
            warnings.append(f"{os.path.basename(path)}: Pl değeri okunamadı.")

        return {
            "sondaj_no": sondaj_no,
            "der": derinlik,
            "em": _fmt_em(em_raw),
            "pl": _fmt_pl(pl_raw),
            "net_pl": _fmt_pl(net_pl_raw),
            "source": str(path),
            "warnings": warnings,
        }
    finally:
        wb.close()


def pmt_excel_dosyalarini_oku(paths):
    records = []
    warnings = []
    for path in paths or []:
        try:
            record = pmt_excel_dosyasi_oku(path)
            records.append(record)
            warnings.extend(record.get("warnings", []))
        except Exception as exc:
            warnings.append(f"{os.path.basename(str(path))}: okunamadı ({exc})")
    return {"records": records, "warnings": warnings}


def pmt_kayitlarini_veriye_aktar(veri, records, update_existing=True):
    sondajlar = (veri or {}).setdefault("sondaj", [])
    by_no = {_sondaj_no_normalize(s.get("no")): s for s in sondajlar if s.get("no")}
    imported = 0
    updated = 0
    skipped = 0
    warnings = []

    for record in records or []:
        no = _sondaj_no_normalize(record.get("sondaj_no"))
        sondaj = by_no.get(no)
        if not sondaj:
            skipped += 1
            warnings.append(f"{no or '?'}: projede sondaj bulunamadı.")
            continue
        row = [record.get("der", ""), record.get("em", ""), record.get("pl", "")]
        if not row[0] or not (row[1] or row[2]):
            skipped += 1
            warnings.append(f"{no}: PMT satırı eksik olduğu için atlandı.")
            continue
        pmt_rows = sondaj.setdefault("pmt", [])
        target_depth = round(safe_float(row[0]), 2)
        replaced = False
        if update_existing:
            for idx, existing in enumerate(pmt_rows):
                if existing and round(safe_float(existing[0]), 2) == target_depth:
                    pmt_rows[idx] = row
                    updated += 1
                    replaced = True
                    break
        if not replaced:
            pmt_rows.append(row)
            imported += 1
        pmt_rows.sort(key=lambda item: safe_float(item[0] if item else 0))

    return {"imported": imported, "updated": updated, "skipped": skipped, "warnings": warnings}
