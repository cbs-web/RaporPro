import datetime
import os
import re
import tempfile
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


TAAHHUT_SHEET = "tahhütname"
APP_DIR = os.path.dirname(os.path.abspath(__file__))
BUILTIN_TEMPLATE_PATH = os.path.join(APP_DIR, "sablonlar", "taahhutname_base.xlsx")
TAAHHUT_METNI = (
    " Yukarıdaki bilgilere sahip projenin müellifliğini üstlenmemde 6235 sayılı Türk Mühendis ve "
    "Mimar Odaları Birliği Kanunu, 3194 sayılı İmar Kanunu ve ilgili mevzuat kapsamında süreli veya "
    "süresiz olarak mesleki faaliyet haklarımda herhangi bir kısıtlılık bulunmadığını ve odama "
    "üyeliğimin devam ettiğini taahhüt ederim.\n\u00a0\n"
    "Yukarıdaki bilgilere sahip yapıya ilişkin hazırlanacak tüm projelerde, 3194 sayılı Kanun ve "
    "deprem, yangın,enerji verimliliği, asansör gibi ilgili tüm mevzuat hükümlerini eksiksiz "
    "uygulayacağımı taahhüt ederim"
)
TAAHHUT_ALT_NOT = (
    "Gerçeğe aykırı beyanda bulunduğu tespit edilenlerin işlemleri iptal edilecek ve bu kişiler "
    "hakkında 5237 sayılı Türk  Ceza  Kanununun  ilgili hükümleri gereği Cumhuriyet Savcılığına "
    "suç duyurusunda bulunulacak, ayrıca 6235 sayılı Türk  Mühendis ve  Mimar  Odaları  Birliği  "
    "Kanunu  ve ilgili  mevzuatı  uyarınca işlem yapılmak üzere ilgili Meslek Odasına bilgi verilecektir."
)


def _clean(value, fallback=""):
    text = "" if value is None else str(value).strip()
    return text if text else fallback


def _safe_name(value, fallback="dosya"):
    text = _clean(value, fallback)
    text = re.sub(r"[^\w\-\.]+", "_", text, flags=re.UNICODE).strip("._")
    return text or fallback


def _mahalle_metni(mahalle):
    text = _clean(mahalle)
    if not text:
        return ""
    lowered = text.casefold()
    if "mah" in lowered or "mahalle" in lowered:
        return text
    return f"{text} Mah."


def taahhutname_yapi_adresi(veri):
    kunye = (veri or {}).get("kunye", {})
    parts = [_mahalle_metni(kunye.get("mah")), _clean(kunye.get("ilce")), _clean(kunye.get("il"))]
    return " ".join(part for part in parts if part).strip()


def _ilgili_idare(veri):
    ayarlar = (veri or {}).get("ayarlar", {})
    kunye = (veri or {}).get("kunye", {})
    explicit = _clean(ayarlar.get("taahhut_ilgili_idare"))
    if explicit:
        return explicit
    ilce = _clean(kunye.get("ilce"))
    il = _clean(kunye.get("il"))
    if ilce:
        return f"{ilce} Belediyesi"
    if il:
        return f"{il} Belediyesi"
    return ""


def _taahhut_tarihi(veri):
    ayarlar = (veri or {}).get("ayarlar", {})
    return _clean(ayarlar.get("taahhut_tarih"), datetime.datetime.now().strftime("%d.%m.%Y"))


def _profile(veri, tur):
    ayarlar = (veri or {}).get("ayarlar", {})
    prefix = "taahhut_jeofizik" if tur == "jeofizik" else "taahhut_jeoloji"
    defaults = {
        "jeoloji": {
            "ad": "Gökalp DOĞAN",
            "sicil": "7400",
            "unvan": "JEOLOJİ MÜHENDİSİ",
            "imza_unvan": "Jeoloji Mühendisi",
            "adres": "İsmetpaşa Mh. Hasan Mevsuf Sk. No :4 Da:5",
            "telefon": "0 545 639 90 62",
        },
        "jeofizik": {
            "ad": "Suat ERGİN",
            "sicil": "1982",
            "unvan": "JEOFİZİK MÜHENDİSİ",
            "imza_unvan": "Jeofizik Mühendisi",
            "adres": "İsmetpaşa Mh. Hasan Mevsuf Sk. No :4 Da:5",
            "telefon": "0 532 281 12 95",
        },
    }[tur]
    return {key: _clean(ayarlar.get(f"{prefix}_{key}"), value) for key, value in defaults.items()}


def taahhutname_context(veri, tur):
    tur = "jeofizik" if tur == "jeofizik" else "jeoloji"
    kunye = (veri or {}).get("kunye", {})
    profile = _profile(veri, tur)
    adres = taahhutname_yapi_adresi(veri)
    return {
        "tur": tur,
        "profile": profile,
        "oda_sicil_no": profile["sicil"],
        "unvan": profile["unvan"],
        "muhendis_ad": profile["ad"],
        "muhendis_imza_unvan": profile["imza_unvan"],
        "muhendis_adres": profile["adres"],
        "muhendis_telefon": profile["telefon"],
        "il": _clean(kunye.get("il")),
        "ilce": _clean(kunye.get("ilce")),
        "mahalle": _clean(kunye.get("mah")),
        "mevki": _clean(kunye.get("mev"), "-"),
        "ilgili_idare": _ilgili_idare(veri),
        "pafta": _clean(kunye.get("paf"), "-"),
        "ada": _clean(kunye.get("ada")),
        "parsel": _clean(kunye.get("par")),
        "yapi_adresi": adres,
        "yapi_sahibi": _clean(kunye.get("sahibi")),
        "yapi_sahibi_adresi": adres,
        "proje_turu": "ZEMİN ETÜDÜ RAPORU",
        "tarih": _taahhut_tarihi(veri),
    }


def _set(ws, cell, value):
    ws[cell] = value


def _fill_ana_sayfa(wb, ctx):
    if "ANA SAYFA" not in wb.sheetnames:
        return
    ws = wb["ANA SAYFA"]
    values = {
        "C2": ctx["yapi_sahibi"],
        "C3": ctx["il"],
        "C4": ctx["ilce"],
        "C5": ctx["mahalle"],
        "C6": ctx["mevki"],
        "C7": ctx["pafta"],
        "C8": ctx["ada"],
        "C9": ctx["parsel"],
        "C11": ctx["ilgili_idare"],
        "C12": ctx["yapi_adresi"],
    }
    for cell, value in values.items():
        _set(ws, cell, value)


def _load_builtin_workbook():
    if not os.path.exists(BUILTIN_TEMPLATE_PATH):
        return None
    wb = load_workbook(BUILTIN_TEMPLATE_PATH)
    ws = wb.active
    ws.title = TAAHHUT_SHEET
    ws.sheet_view.showGridLines = False
    return wb


def _merge(ws, cell_range):
    try:
        ws.merge_cells(cell_range)
    except ValueError:
        pass


def _set_merged(ws, cell_range, value, font=None, alignment=None, border=None, fill=None):
    from openpyxl.worksheet.cell_range import CellRange

    _merge(ws, cell_range)
    target = CellRange(cell_range)
    cell = ws.cell(target.min_row, target.min_col)
    cell.value = value
    if font is not None:
        cell.font = font
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if fill is not None:
        cell.fill = fill
    return cell


def _apply_border(ws, cell_range, border):
    from openpyxl.worksheet.cell_range import CellRange

    target = CellRange(cell_range)
    for row in ws.iter_rows(
        min_row=target.min_row,
        max_row=target.max_row,
        min_col=target.min_col,
        max_col=target.max_col,
    ):
        for cell in row:
            cell.border = border


def _apply_outline_border(ws, cell_range, side):
    from openpyxl.worksheet.cell_range import CellRange

    target = CellRange(cell_range)
    for row in ws.iter_rows(
        min_row=target.min_row,
        max_row=target.max_row,
        min_col=target.min_col,
        max_col=target.max_col,
    ):
        for cell in row:
            cell.border = Border(
                left=side if cell.column == target.min_col else cell.border.left,
                right=side if cell.column == target.max_col else cell.border.right,
                top=side if cell.row == target.min_row else cell.border.top,
                bottom=side if cell.row == target.max_row else cell.border.bottom,
            )


def _setup_generated_workbook():
    from openpyxl.styles import fonts as openpyxl_fonts

    default_font = openpyxl_fonts.DEFAULT_FONT
    openpyxl_fonts.DEFAULT_FONT = Font(name="Aptos Narrow", size=11, family=2, charset=162, scheme="minor")
    try:
        wb = Workbook()
    finally:
        openpyxl_fonts.DEFAULT_FONT = default_font
    ws = wb.active
    ws.title = TAAHHUT_SHEET
    ws.sheet_view.showGridLines = False
    ws.sheet_format.defaultRowHeight = 14.4

    widths = {
        "A": 12.44140625, "B": 6.33203125, "C": 13.0, "D": 13.0, "E": 9.44140625,
        "F": 10.0, "G": 13.0, "H": 13.88671875, "I": 8.6640625,
        "J": 12.44140625, "K": 6.33203125, "L": 13.0, "M": 13.0, "N": 9.44140625,
        "O": 10.0, "P": 13.0, "Q": 13.88671875, "R": 8.6640625,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    thin = Side(style="thin", color="000000")
    title_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    signature_border = Border(right=thin)
    title_font = Font(name="Times New Roman", size=11, bold=True)
    section_font = Font(name="Times New Roman", size=10, bold=True)
    label_font = Font(name="Times New Roman", size=10, bold=True)
    value_font = Font(name="Aptos Narrow", size=11)
    center = Alignment(horizontal="center", vertical="center")
    label_align = Alignment(vertical="center")
    value_align = Alignment(horizontal="left")

    for start_col, end_col, label_col, colon_col, value_col in (
        ("A", "I", "A", "B", "C"),
        ("J", "R", "J", "K", "L"),
    ):
        _set_merged(ws, f"{start_col}3:{end_col}3", "TAAHHÜTNAME", title_font, center, title_border)
        _set_merged(ws, f"{start_col}9:{end_col}9", "Müellifliği Üstlenilen Proje", section_font, center, title_border)
        for row, label in (
            (4, "Oda Sicil No"),
            (5, "Unvanı"),
            (6, "Adresi"),
            (7, "Telefonu"),
        ):
            ws[f"{label_col}{row}"] = label
            ws[f"{label_col}{row}"].font = label_font
            ws[f"{label_col}{row}"].alignment = label_align
            ws[f"{colon_col}{row}"] = ":"
            ws[f"{colon_col}{row}"].alignment = center
            ws[f"{value_col}{row}"].font = value_font
            ws[f"{value_col}{row}"].alignment = value_align

    for cell_range in (
        "C4:I4", "C5:I5", "C6:I6", "C7:I7",
        "L4:R4", "L5:R5", "L6:R6", "L7:R7",
        "D11:I11", "D12:I12", "D13:G13", "H13:I13", "D14:I14", "D15:I15", "D16:I16", "D17:I17",
        "M11:R11", "M12:R12", "M13:P13", "Q13:R13", "M14:R14", "M15:R15", "M16:R16", "M17:R17",
    ):
        _merge(ws, cell_range)
    for cell_range in ("A3:I7", "J3:R7", "A9:I18", "J9:R18"):
        _apply_outline_border(ws, cell_range, thin)

    project_labels = [
        (11, "İl / İlçe"),
        (12, "İlgili İdare"),
        (13, "Pafta/Ada/Parsel No"),
        (14, "Yapı Adresi"),
        (15, "Yapı Sahibi"),
        (16, "Yapı Sahibinin Adresi"),
        (17, "Projenin Türü"),
    ]
    for row, label in project_labels:
        for label_col, colon_col in (("A", "C"), ("J", "L")):
            ws[f"{label_col}{row}"] = label
            ws[f"{label_col}{row}"].font = label_font
            ws[f"{label_col}{row}"].alignment = label_align
            ws[f"{colon_col}{row}"] = ":"
            ws[f"{colon_col}{row}"].alignment = center

    body_font = Font(name="Aptos Narrow", size=11)
    footer_font = Font(name="Times New Roman", size=11)
    _set_merged(ws, "A20:I26", "", body_font, Alignment(horizontal="left", wrap_text=True))
    _set_merged(ws, "J20:R26", "", body_font, Alignment(horizontal="left", wrap_text=True))
    _apply_outline_border(ws, "A20:I42", thin)
    _apply_outline_border(ws, "J20:R42", thin)
    _set_merged(ws, "F30:I30", "", Font(name="Times New Roman", size=11), center, signature_border)
    _set_merged(ws, "F31:I31", "", Font(name="Times New Roman", size=11), center, signature_border)
    _set_merged(ws, "F32:I32", "", Font(name="Times New Roman", size=11), center, signature_border)
    _set_merged(ws, "O30:R30", "", Font(name="Times New Roman", size=11), center, signature_border)
    _set_merged(ws, "O31:R31", "", Font(name="Times New Roman", size=11), center, signature_border)
    _set_merged(ws, "O32:R32", "", Font(name="Times New Roman", size=11), center, signature_border)
    _set_merged(ws, "A44:I47", TAAHHUT_ALT_NOT, footer_font, Alignment(horizontal="left", vertical="top", wrap_text=True), title_border)
    _set_merged(ws, "J44:R47", TAAHHUT_ALT_NOT, footer_font, Alignment(horizontal="left", vertical="top", wrap_text=True), title_border)
    _apply_outline_border(ws, "A44:I47", thin)
    _apply_outline_border(ws, "J44:R47", thin)

    ws.page_margins.left = 0.7
    ws.page_margins.right = 0.7
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    return wb


def _fill_profile_cells(ws, veri, tur):
    profile = _profile(veri, tur)
    value_font = Font(name="Aptos Narrow", size=11)
    value_align = Alignment(horizontal="left")
    signature_font = Font(name="Times New Roman", size=11)
    signature_align = Alignment(horizontal="center", vertical="center")

    ws["A4"] = "Oda Sicil No"
    ws["C4"] = profile["sicil"]
    ws["C5"] = profile["unvan"]
    ws["C6"] = profile["adres"]
    ws["C7"] = profile["telefon"]
    ws["F31"] = profile["ad"]
    ws["F32"] = profile["imza_unvan"]
    for cell_ref in ("C4", "C5", "C6", "C7"):
        ws[cell_ref].font = value_font
        ws[cell_ref].alignment = value_align
    for cell_ref in ("F30", "F31", "F32"):
        ws[cell_ref].font = signature_font
        ws[cell_ref].alignment = signature_align


def _fill_project_cells(ws, ctx):
    il_ilce = " / ".join(part for part in (ctx["il"], ctx["ilce"]) if part).strip()
    pafta_ada_parsel = " / ".join(part for part in (ctx["pafta"], ctx["ada"], ctx["parsel"]) if part and part != "-").strip() or "-"
    values = {
        "D11": il_ilce,
        "D12": ctx["ilgili_idare"],
        "D13": pafta_ada_parsel,
        "D14": ctx["yapi_adresi"],
        "D15": ctx["yapi_sahibi"],
        "D16": ctx["yapi_sahibi_adresi"],
        "D17": ctx["proje_turu"],
        "F30": ctx["tarih"],
    }
    for cell, value in values.items():
        _set(ws, cell, value)
        ws[cell].font = Font(name="Aptos Narrow", size=11)
        if cell == "F30":
            ws[cell].alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws[cell].alignment = Alignment(horizontal="left")


def _range_overlaps(a, b):
    return not (a.max_col < b.min_col or a.min_col > b.max_col or a.max_row < b.min_row or a.min_row > b.max_row)


def _merge_body_text(ws, cell_range):
    from openpyxl.worksheet.cell_range import CellRange

    target = CellRange(cell_range)
    exact_merge_exists = any(str(merged) == str(target) for merged in ws.merged_cells.ranges)
    if not exact_merge_exists:
        for merged in list(ws.merged_cells.ranges):
            if _range_overlaps(merged, target):
                ws.unmerge_cells(str(merged))
        ws.merge_cells(cell_range)
    cell = ws.cell(target.min_row, target.min_col)
    cell.value = TAAHHUT_METNI
    cell.alignment = Alignment(horizontal="left", wrap_text=True)
    cell.font = Font(name="Aptos Narrow", size=11)


def _fill_body_text(ws):
    _merge_body_text(ws, "A20:I26")


def _configure_workbook_for_type(wb, tur):
    ws = wb[TAAHHUT_SHEET]
    ws.sheet_view.showGridLines = False
    for sheet in wb.worksheets:
        sheet.sheet_state = "hidden"
    ws.sheet_state = "visible"
    wb.active = wb.sheetnames.index(TAAHHUT_SHEET)

    ws.print_area = "A1:I47"
    for idx in range(10, 19):
        ws.column_dimensions[get_column_letter(idx)].hidden = True
    ws.freeze_panes = "A1"

    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = 9
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    except Exception:
        pass


def _build_workbook(veri, tur):
    wb = _load_builtin_workbook() or _setup_generated_workbook()
    ctx = taahhutname_context(veri, tur)
    _fill_ana_sayfa(wb, ctx)
    ws = wb[TAAHHUT_SHEET]
    _fill_profile_cells(ws, veri, ctx["tur"])
    _fill_project_cells(ws, ctx)
    _fill_body_text(ws)
    _configure_workbook_for_type(wb, "jeofizik" if tur == "jeofizik" else "jeoloji")
    return wb


def _export_xlsx_to_pdf(xlsx_path, pdf_path):
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise RuntimeError(f"Excel PDF aktarımı için pywin32 bulunamadı: {exc}") from exc

    excel = None
    workbook = None
    pythoncom.CoInitialize()
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(os.path.abspath(xlsx_path))
        worksheet = workbook.Worksheets(TAAHHUT_SHEET)
        worksheet.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
    except Exception as exc:
        raise RuntimeError(
            "Taahhütname PDF'e çevrilemedi. XLSX çıktı oluşturulabilir; PDF için Microsoft Excel'in "
            f"bu oturumda çalışabilir olması gerekir. Hata: {exc}"
        ) from exc
    finally:
        if workbook is not None:
            workbook.Close(False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def _patch_xlsx_reference_defaults(path):
    tmp_path = f"{path}.tmp"
    reference_cols = (
        '<cols><col width="12.44140625" customWidth="1" style="43" min="1" max="1"/>'
        '<col width="6.33203125" customWidth="1" style="43" min="2" max="2"/>'
        '<col width="9.44140625" customWidth="1" style="43" min="5" max="5"/>'
        '<col width="10" customWidth="1" style="43" min="6" max="6"/>'
        '<col width="13.88671875" customWidth="1" style="43" min="8" max="8"/>'
        '<col width="8.6640625" customWidth="1" style="43" min="9" max="9"/>'
        '<col hidden="1" width="13" customWidth="1" style="43" min="10" max="10"/>'
        '<col hidden="1" width="13" customWidth="1" style="43" min="11" max="11"/>'
        '<col hidden="1" width="13" customWidth="1" style="43" min="12" max="12"/>'
        '<col hidden="1" width="13" customWidth="1" style="43" min="13" max="13"/>'
        '<col hidden="1" width="13" customWidth="1" style="43" min="14" max="14"/>'
        '<col hidden="1" width="13" customWidth="1" style="43" min="15" max="15"/>'
        '<col hidden="1" width="13" customWidth="1" style="43" min="16" max="16"/>'
        '<col hidden="1" width="13" customWidth="1" style="43" min="17" max="17"/>'
        '<col hidden="1" width="13" customWidth="1" style="43" min="18" max="18"/></cols>'
    )
    try:
        with ZipFile(path, "r") as zin:
            entries = [(item, zin.read(item.filename)) for item in zin.infolist()]
        styles_xml = next((data.decode("utf-8") for item, data in entries if item.filename == "xl/styles.xml"), "")
        can_use_reference_cols = "Aptos Narrow" in styles_xml

        with ZipFile(tmp_path, "w", ZIP_DEFLATED) as zout:
            for item, data in entries:
                if item.filename == "xl/styles.xml":
                    text = data.decode("utf-8")
                    text = text.replace('<name val="Calibri"/>', '<name val="Aptos Narrow"/><charset val="162"/>', 1)
                    data = text.encode("utf-8")
                elif item.filename == "xl/worksheets/sheet1.xml" and can_use_reference_cols:
                    text = data.decode("utf-8")
                    start = text.find("<cols>")
                    end = text.find("</cols>")
                    if start >= 0 and end >= 0 and 'style="43"' in text[start:end]:
                        text = f"{text[:start]}{reference_cols}{text[end + len('</cols>'):]}"
                        data = text.encode("utf-8")
                zout.writestr(item, data)
        os.replace(tmp_path, path)
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass


def taahhutname_dosya_adi(veri, tur, ext=".xlsx"):
    ctx = taahhutname_context(veri, tur)
    role = "Jeofizik" if ctx["tur"] == "jeofizik" else "Jeoloji"
    owner = _safe_name(ctx["yapi_sahibi"], "Proje")
    ext = ext if str(ext).startswith(".") else f".{ext}"
    return f"{owner}_Taahhutname_{role}{ext}"


def taahhutname_olustur(veri, tur, path):
    ext = os.path.splitext(path)[1].lower()
    if ext not in (".xlsx", ".pdf"):
        path = f"{path}.xlsx"
        ext = ".xlsx"

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    wb = _build_workbook(veri, tur)
    if ext == ".xlsx":
        wb.save(path)
        _patch_xlsx_reference_defaults(path)
        return path

    with tempfile.TemporaryDirectory(prefix="raporpro_taahhut_") as tmp_dir:
        xlsx_path = os.path.join(tmp_dir, taahhutname_dosya_adi(veri, tur, ".xlsx"))
        wb.save(xlsx_path)
        _patch_xlsx_reference_defaults(xlsx_path)
        _export_xlsx_to_pdf(xlsx_path, path)
    return path


def tum_taahhutnameleri_olustur(veri, folder, ext=".xlsx"):
    os.makedirs(folder, exist_ok=True)
    paths = []
    for tur in ("jeoloji", "jeofizik"):
        path = os.path.join(folder, taahhutname_dosya_adi(veri, tur, ext))
        paths.append(taahhutname_olustur(veri, tur, path))
    return paths
