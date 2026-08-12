# Dosya: RaporPro/tests/test_performans_optimizasyonlari.py
import tempfile
import unittest
from pathlib import Path
from unittest import mock

import openpyxl
from openpyxl import Workbook

from spt_okuma_motoru import SPTKaydi, _select_spt_records_for_batch, excelden_spt_oku


class SPTPerformansOptimizasyonuTests(unittest.TestCase):
    def test_excel_akista_okunur_ve_dosya_kapatilir(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "spt.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "SPT"
            sheet.append(["Sondaj No", "Derinlik", "15", "30", "45", "N30"])
            sheet.append(["SK-1", 1.5, 4, 5, 6, 11])
            sheet.append(["SK-2", 3.0, 7, 8, 9, 17])
            workbook.save(path)
            workbook.close()

            original_load = openpyxl.load_workbook
            observed_kwargs = []

            def observed_load(*args, **kwargs):
                observed_kwargs.append(dict(kwargs))
                return original_load(*args, **kwargs)

            with mock.patch.object(openpyxl, "load_workbook", side_effect=observed_load):
                result = excelden_spt_oku(path)

            self.assertEqual(2, len(result.kayitlar))
            self.assertTrue(observed_kwargs[0]["read_only"])
            self.assertTrue(observed_kwargs[0]["data_only"])
            path.unlink()
            self.assertFalse(path.exists())

    def test_onceden_hesaplanan_fotograf_sirasi_yolu_tekrar_cozmez(self):
        records = [
            SPTKaydi(sondaj_no="SK-1", derinlik="1.50", n30="10", guven="80"),
            SPTKaydi(sondaj_no="SK-1", derinlik="3.00", n30="12", guven="90"),
        ]
        with mock.patch(
            "spt_okuma_motoru._path_unique_key",
            side_effect=AssertionError("tekrar cozulmemeli"),
        ):
            selected, removed, merged = _select_spt_records_for_batch(
                [("foto-1", records)],
                ["C:/foto-1.jpg"],
                path_order=["foto-1"],
            )

        self.assertEqual(1, len(selected))
        self.assertEqual(1, removed)
        self.assertEqual(0, merged)


if __name__ == "__main__":
    unittest.main()
