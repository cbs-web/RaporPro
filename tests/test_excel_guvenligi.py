from openpyxl import load_workbook

from excel_guvenligi import excel_hucre_degeri
from taahhutname import _set, _set_merged
from workbook_motoru import excel_workbook_yaz


def test_excel_hucre_degeri_formul_baslangiclarini_metne_cevirir():
    assert excel_hucre_degeri("=HYPERLINK(\"x\")") == "'=HYPERLINK(\"x\")"
    assert excel_hucre_degeri("+SUM(A1:A2)") == "'+SUM(A1:A2)"
    assert excel_hucre_degeri("-SUM(A1:A2)") == "'-SUM(A1:A2)"
    assert excel_hucre_degeri("@SUM(A1:A2)") == "'@SUM(A1:A2)"
    assert excel_hucre_degeri("\t=1+1") == "'\t=1+1"


def test_excel_hucre_degeri_muhendislik_sayilarini_korur():
    assert excel_hucre_degeri("-12.50") == "-12.50"
    assert excel_hucre_degeri("+1,25") == "+1,25"
    assert excel_hucre_degeri("-") == "-"
    assert excel_hucre_degeri(12.5) == 12.5


def test_workbook_yazici_formul_uretmez(tmp_path):
    path = tmp_path / "guvenli.xlsx"
    excel_workbook_yaz(
        path,
        [{
            "title": "Veri",
            "headers": ["Tanim"],
            "rows": [["=2+2"], ["-3.5"]],
            "widths": [100],
        }],
    )

    workbook = load_workbook(path, data_only=False, read_only=True)
    try:
        sheet = workbook["Veri"]
        assert sheet["A2"].value == "'=2+2"
        assert sheet["A2"].data_type == "s"
        assert sheet["A3"].value == "-3.5"
        assert sheet["A3"].data_type == "s"
    finally:
        workbook.close()


def test_taahhutname_hucre_yardimcilari_serbest_metni_kacar():
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    _set(sheet, "A1", "=1+1")
    _set_merged(sheet, "B1:C1", "@SUM(A1)")

    assert sheet["A1"].value == "'=1+1"
    assert sheet["A1"].data_type == "s"
    assert sheet["B1"].value == "'@SUM(A1)"
    assert sheet["B1"].data_type == "s"
