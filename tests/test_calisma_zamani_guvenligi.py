import os
import sys
import tempfile
import types
import unittest
from types import SimpleNamespace
from unittest import mock

from docx import Document
from openpyxl import Workbook

import ekler
import taahhutname
import tutanaklar
import ui_spt_okuma
import ui_spt_okuma_kuyruk
import ui_workbook


class _FakePdf:
    def __init__(self, output_bytes=b"%PDF-yeni"):
        self.page_count = 0
        self.output_bytes = output_bytes
        self.saved_path = None
        self.closed = False

    def save(self, path, **_kwargs):
        self.saved_path = path
        with open(path, "wb") as stream:
            stream.write(self.output_bytes)

    def close(self):
        self.closed = True


class CiktiGuvenligiTestleri(unittest.TestCase):
    def _run_fake_pdf_export(self, output_path, fake_pdf, validator):
        def add_page(pdf_doc, *_args, **_kwargs):
            pdf_doc.page_count = 1

        with (
            mock.patch.object(ekler, "ek_set_sablonu", return_value=("normal", "Normal", None)),
            mock.patch.object(ekler, "ek_basliklari", return_value=[]),
            mock.patch.object(ekler, "ek_icerik_haritasi", return_value={}),
            mock.patch.object(ekler, "_append_message_page", side_effect=add_page),
            mock.patch.object(ekler, "_pdf_ciktisini_dogrula", side_effect=validator),
            mock.patch.object(ekler.fitz, "open", return_value=fake_pdf),
        ):
            return ekler.ekler_pdf_olustur({}, output_path)

    def test_ekler_pdf_dogrulama_hatasinda_eski_cikti_korunur_ve_dokuman_kapanir(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, "ekler.pdf")
            with open(output_path, "wb") as stream:
                stream.write(b"eski")
            fake_pdf = _FakePdf()

            with self.assertRaisesRegex(RuntimeError, "gecersiz"):
                self._run_fake_pdf_export(
                    output_path,
                    fake_pdf,
                    RuntimeError("gecersiz"),
                )

            with open(output_path, "rb") as stream:
                self.assertEqual(stream.read(), b"eski")
            self.assertTrue(fake_pdf.closed)
            self.assertEqual(os.path.dirname(fake_pdf.saved_path), temp_dir)

    def test_ekler_pdf_yalnizca_dogrulamadan_sonra_atomik_degistirilir(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, "ekler.pdf")
            with open(output_path, "wb") as stream:
                stream.write(b"eski")
            fake_pdf = _FakePdf()

            self._run_fake_pdf_export(output_path, fake_pdf, None)

            with open(output_path, "rb") as stream:
                self.assertEqual(stream.read(), b"%PDF-yeni")
            self.assertTrue(fake_pdf.closed)

    def test_pdf_dogrulama_gecersiz_dokumani_da_kapatir(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            path = os.path.join(temp_dir, "bos.pdf")
            with open(path, "wb") as stream:
                stream.write(b"x")
            fake_pdf = _FakePdf()
            fake_pdf.page_count = 0
            with mock.patch.object(ekler.fitz, "open", return_value=fake_pdf):
                with self.assertRaisesRegex(RuntimeError, "sayfa"):
                    ekler._pdf_ciktisini_dogrula(path)
            self.assertTrue(fake_pdf.closed)


class ComGuvenligiTestleri(unittest.TestCase):
    def test_belge_kapatma_hatasi_quit_ve_coinitialize_temizligini_engellemez(self):
        for module in (ekler, taahhutname, tutanaklar):
            with self.subTest(module=module.__name__):
                belge = SimpleNamespace(Close=mock.Mock(side_effect=RuntimeError("close")))
                uygulama = SimpleNamespace(Quit=mock.Mock())
                pythoncom = SimpleNamespace(CoUninitialize=mock.Mock())

                module._com_guvenli_temizle(
                    pythoncom,
                    com_initialized=True,
                    belge=belge,
                    uygulama=uygulama,
                )

                belge.Close.assert_called_once_with(False)
                uygulama.Quit.assert_called_once_with()
                pythoncom.CoUninitialize.assert_called_once_with()

    @staticmethod
    def _com_modules(app):
        pythoncom = types.ModuleType("pythoncom")
        pythoncom.CoInitialize = mock.Mock()
        pythoncom.CoUninitialize = mock.Mock()
        client = types.ModuleType("win32com.client")
        client.DispatchEx = mock.Mock(return_value=app)
        win32com = types.ModuleType("win32com")
        win32com.client = client
        return pythoncom, win32com, client

    def test_legacy_word_makrolari_ve_baglanti_guncellemesi_kapali_acilir(self):
        document = SimpleNamespace(SaveAs=mock.Mock(), Close=mock.Mock())
        documents = SimpleNamespace(Open=mock.Mock(return_value=document))
        app = SimpleNamespace(
            Options=SimpleNamespace(UpdateLinksAtOpen=True),
            Documents=documents,
            Quit=mock.Mock(),
        )
        pythoncom, win32com, client = self._com_modules(app)

        with (
            tempfile.TemporaryDirectory() as temp_dir,
            mock.patch.dict(
                sys.modules,
                {
                    "pythoncom": pythoncom,
                    "win32com": win32com,
                    "win32com.client": client,
                },
            ),
        ):
            ekler._office_to_pdf("ornek.doc", temp_dir)

        self.assertEqual(app.AutomationSecurity, 3)
        self.assertFalse(app.Options.UpdateLinksAtOpen)
        self.assertFalse(documents.Open.call_args.kwargs["ConfirmConversions"])
        self.assertTrue(documents.Open.call_args.kwargs["ReadOnly"])
        self.assertFalse(documents.Open.call_args.kwargs["AddToRecentFiles"])
        pythoncom.CoUninitialize.assert_called_once_with()

    def test_legacy_excel_makrolari_ve_dis_baglantilari_kapali_acilir(self):
        workbook = SimpleNamespace(ExportAsFixedFormat=mock.Mock(), Close=mock.Mock())
        workbooks = SimpleNamespace(Open=mock.Mock(return_value=workbook))
        app = SimpleNamespace(Workbooks=workbooks, Quit=mock.Mock())
        pythoncom, win32com, client = self._com_modules(app)

        with (
            tempfile.TemporaryDirectory() as temp_dir,
            mock.patch.dict(
                sys.modules,
                {
                    "pythoncom": pythoncom,
                    "win32com": win32com,
                    "win32com.client": client,
                },
            ),
        ):
            ekler._office_to_pdf("ornek.xls", temp_dir)

        self.assertEqual(app.AutomationSecurity, 3)
        self.assertFalse(app.AskToUpdateLinks)
        self.assertFalse(app.EnableEvents)
        open_kwargs = workbooks.Open.call_args.kwargs
        self.assertEqual(open_kwargs["UpdateLinks"], 0)
        self.assertTrue(open_kwargs["ReadOnly"])
        self.assertFalse(open_kwargs["AddToMru"])
        pythoncom.CoUninitialize.assert_called_once_with()


class MetadataTestleri(unittest.TestCase):
    def test_xlsx_metadata_raporpro_olarak_notrlenir(self):
        workbook = Workbook()
        workbook.properties.creator = "Gercek Kisi"
        workbook.properties.lastModifiedBy = "Gercek Kisi"

        taahhutname._raporpro_workbook_metadata(workbook)

        self.assertEqual(workbook.properties.creator, "RaporPro")
        self.assertEqual(workbook.properties.lastModifiedBy, "RaporPro")

    def test_docx_metadata_raporpro_olarak_notrlenir(self):
        document = Document()
        document.core_properties.author = "Gercek Kisi"
        document.core_properties.last_modified_by = "Gercek Kisi"

        tutanaklar._raporpro_docx_metadata(document)

        self.assertEqual(document.core_properties.author, "RaporPro")
        self.assertEqual(document.core_properties.last_modified_by, "RaporPro")


class WorkbookOkumaTestleri(unittest.TestCase):
    def test_xlsx_salt_okunur_acilir_ve_kapanir(self):
        definitions = {
            "spt": {
                "title": "SPT",
                "columns": [("Sondaj", "sondaj_no"), ("Derinlik", "der")],
            }
        }
        sheet = SimpleNamespace(
            iter_rows=mock.Mock(
                return_value=iter(
                    [
                        ("Sondaj", "Derinlik"),
                        ("SK-1", 1.5),
                    ]
                )
            )
        )

        class FakeWorkbook:
            sheetnames = ["SPT"]

            def __init__(self):
                self.close = mock.Mock()

            def __getitem__(self, _name):
                return sheet

        workbook = FakeWorkbook()
        with mock.patch("openpyxl.load_workbook", return_value=workbook) as loader:
            result = ui_workbook._xlsx_workbook_satirlarini_oku("buyuk.xlsx", definitions)

        loader.assert_called_once_with(
            "buyuk.xlsx",
            data_only=True,
            read_only=True,
            keep_links=False,
        )
        workbook.close.assert_called_once_with()
        self.assertEqual(result["rows_by_sheet"]["spt"], [["SK-1", "1.5"]])

    def test_xlsx_iterasyon_hatasinda_da_kapanir(self):
        definitions = {
            "spt": {
                "title": "SPT",
                "columns": [("Sondaj", "sondaj_no"), ("Derinlik", "der")],
            }
        }

        def broken_rows():
            yield ("Sondaj", "Derinlik")
            raise RuntimeError("satir hatasi")

        sheet = SimpleNamespace(iter_rows=mock.Mock(return_value=broken_rows()))

        class FakeWorkbook:
            sheetnames = ["SPT"]

            def __init__(self):
                self.close = mock.Mock()

            def __getitem__(self, _name):
                return sheet

        workbook = FakeWorkbook()
        with mock.patch("openpyxl.load_workbook", return_value=workbook):
            with self.assertRaisesRegex(RuntimeError, "satir hatasi"):
                ui_workbook._xlsx_workbook_satirlarini_oku("bozuk.xlsx", definitions)
        workbook.close.assert_called_once_with()


class SptGuvenligiTestleri(unittest.TestCase):
    def test_tree_eski_item_id_degerini_kabul_etmez(self):
        record = {"kayit": object()}
        tree = SimpleNamespace(exists=lambda item_id: item_id == "yeni")
        tree_items = {"eski": record, "yeni": record}

        self.assertFalse(ui_spt_okuma._tree_item_gecerli(tree, tree_items, "eski", record))
        self.assertTrue(ui_spt_okuma._tree_item_gecerli(tree, tree_items, "yeni", record))

    def test_spt_baslangic_klasoru_proje_ayarini_kullanir(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            app = SimpleNamespace(
                veri={"ayarlar": {"spt_son_klasor": temp_dir}},
                aktif_dosya_yolu=None,
            )
            result = ui_spt_okuma.SPTOkumaMixin._spt_initial_dir(app)
        self.assertEqual(result, os.path.abspath(temp_dir))

    def test_dis_servis_reddi_islemi_durdurur_kabul_bir_kez_sorulur(self):
        app = SimpleNamespace(veri={"ayarlar": {}})
        with mock.patch.object(
            ui_spt_okuma.messagebox,
            "askyesno",
            side_effect=[False, True],
        ) as ask:
            self.assertFalse(
                ui_spt_okuma.SPTOkumaMixin._spt_dis_servis_onayi_al(
                    app,
                    {"aktif_motor": "openai"},
                )
            )
            self.assertTrue(
                ui_spt_okuma.SPTOkumaMixin._spt_dis_servis_onayi_al(
                    app,
                    {"aktif_motor": "openai"},
                )
            )
            self.assertTrue(
                ui_spt_okuma.SPTOkumaMixin._spt_dis_servis_onayi_al(
                    app,
                    {"aktif_motor": "openai"},
                )
            )
        self.assertEqual(ask.call_count, 2)
        self.assertIn("OpenAI", ask.call_args_list[0].args[1])
        self.assertEqual(app.veri["ayarlar"]["spt_dis_servis_onayi_openai"], "1")

    def test_kuyruk_hashi_dosya_degismedikce_bir_kez_hesaplanir(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            paths = []
            for name in ("a.jpg", "b.jpg"):
                path = os.path.join(temp_dir, name)
                with open(path, "wb") as stream:
                    stream.write(name.encode("ascii"))
                paths.append(path)

            queue = ui_spt_okuma_kuyruk.SPTFotografKuyrugu()
            with (
                mock.patch.object(
                    ui_spt_okuma_kuyruk,
                    "collect_image_paths",
                    return_value=paths,
                ),
                mock.patch.object(
                    ui_spt_okuma_kuyruk,
                    "source_content_key",
                    side_effect=lambda path: f"hash:{os.path.basename(path)}",
                ) as content_key,
            ):
                queue.add_sources(paths)
                queue.deduplicated_paths()
                queue.add_sources(paths)

            self.assertEqual(content_key.call_count, 2)
            self.assertEqual(queue.paths, [os.path.abspath(path) for path in paths])


if __name__ == "__main__":
    unittest.main()
